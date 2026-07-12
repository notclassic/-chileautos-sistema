#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MOTOR DE CALCULO - Capa 1 del sistema de inteligencia Chileautos
================================================================

Que hace:
  1. Lee TODAS las capturas (.xlsx) de la carpeta capturas/
  2. Limpia cada una con las mismas reglas validadas.
  3. Calcula precio justo y oportunidades sobre la captura mas reciente.
  4. Compara capturas consecutivas para medir ROTACION (avisos que
     desaparecen = proxy de venta) y CAMBIOS DE PRECIO.
  5. Cruza etiquetas de comentarios + comparables.
  6. Escribe datos.json con todo lo calculado.
  7. Guarda base_FECHA.xlsx con la base analizada.

El dashboard lee datos.json. El motor no toca el dashboard.

Uso:
    python motor.py
    (lee carpeta ./capturas, escribe ./datos.json)

Requiere: pandas, numpy, openpyxl
"""

import os, sys, json, glob, re
from datetime import datetime
import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# CONFIGURACION
# ---------------------------------------------------------------------------
CARPETA_CAPTURAS = "capturas"
SALIDA = "datos.json"

COLUMNAS_ESPERADAS = [
    'fecha','marca','modelo','version','año','traccion','condicion','precio',
    'kilometraje','transmision','combustible','id','moneda','entidad',
    'Enlace Web','telefono','whatsapp','nombre_whatsapp'
]

# Reglas de limpieza (validadas con las bases reales)
PRECIO_MIN, PRECIO_MAX = 500_000, 300_000_000
KM_MIN, KM_MAX = 0, 500_000
ANIO_MIN, ANIO_MAX = 1990, 2027

# Reglas de oportunidad
DESCUENTO_MIN = 15.0      # % bajo precio justo para "bajo precio"
KM_ANUAL_MAX = 10_000     # umbral "bajo kilometraje"
MIN_COMPARABLES = 5       # minimo de comparables para tasar


def log(msg):
    print(f"[motor] {msg}")


# ---------------------------------------------------------------------------
# 1.a EXTRACCION DE MOTOR DESDE VERSION
# ---------------------------------------------------------------------------

def extraer_motor(version_str):
    """
    Extrae la cilindrada del texto libre del campo 'version'.
    Ejemplos: '1.5 allure pack bluehdi' -> '1.5'
              '2.0t elite diesel'        -> '2.0'   (sufijo turbo)
              '5.0l xlt 4x4'            -> '5.0'   (sufijo litros)
              'gt line puretech'         -> None    (sin número)
    Devuelve str o None.
    """
    if pd.isna(version_str) or not str(version_str).strip():
        return None
    # Normalizar sufijos comunes: "1.5t" -> "1.5", "5.0l" -> "5.0", "2.0i" -> "2.0"
    s = re.sub(r'(\d+\.\d+)[tTlLiI]\b', r'\1', str(version_str).lower())
    m = re.search(r'\b(\d+\.\d+)\b', s)
    return m.group(1) if m else None


FAMILIAS_MOTOR = [
    # (nombre_canonico, variantes a buscar en el texto SIN espacios ni guiones)
    ("bluehdi",  ["bluehdi"]),          # cubre "blue hdi", "blue-hdi"
    ("hybrid4",  ["hybrid4"]),
    ("puretech", ["puretech"]),
    ("thp",      ["thp"]),
    ("ecoboost", ["ecoboost"]),
    ("tfsi",     ["tfsi"]),
    ("tsi",      ["tsi"]),
    ("tdi",      ["tdi"]),
    ("dci",      ["dci"]),
    ("crdi",     ["crdi"]),
    ("vtec",     ["vtec"]),
    ("skyactiv", ["skyactiv"]),
    ("phev",     ["phev"]),
    ("mhev",     ["mhev", "mildhybrid"]),
    ("hibrido",  ["hybrid", "hibrido", "hev"]),   # genérico, después de hybrid4/phev/mhev
    ("hdi",      ["hdi"]),              # después de bluehdi
    ("vti",      ["vti"]),
]


CORRECCIONES_DEPURADOR = "versiones_canonicas.json"


def cargar_correcciones():
    """Correcciones hechas en la hoja Depuración del dashboard:
    { 'marca|modelo|gen|firma': {'canonica': 'nombre', 'excluir': bool} }"""
    if os.path.exists(CORRECCIONES_DEPURADOR):
        try:
            with open(CORRECCIONES_DEPURADOR, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def sugerir_canonica(firma):
    """'1.6|thp|4x2' -> '1.6 THP 4x2' (sugerencia editable)."""
    if not firma:
        return ''
    partes = [p for p in str(firma).split('|') if p]
    out = []
    for p in partes:
        out.append(p.upper() if p not in ('4x2', '4x4') and not p[0].isdigit() else p)
    return ' '.join(out)


# ---------------------------------------------------------------------------
# PRE-CALIFICACION DE VERSIONES
# Un modelo suele concentrar ~90% de sus avisos en pocas versiones oficiales
# (dominantes). Cuando un grupo de avisos tiene version rara/sucia y sin
# ficha, el sistema compara sus señales duras (cilindrada, traccion,
# transmision, combustible, año, tokens del nombre y PRECIO) contra el
# perfil de cada version dominante y sugiere a cual pertenece.
# ---------------------------------------------------------------------------
import re as _re

_TRIM_STOP = {'diesel', 'bencina', 'gasolina', 'at', 'mt', 'aut', 'mec',
              '4x4', '4x2', 'awd', 'fwd', 'rwd', '2wd', '4wd', '4p', '5p',
              '2p', '3p', 'euro5', 'euro6', 'ac', 'dab', 'abs', 'll', 'hp',
              'cvt', 'dct', 'eat6', 'eat8', 'at6', 'at8', 'mt6', 'bvm', 's'}


def _cil_de_texto(t):
    mm = _re.search(r'\b(\d\.\d)\b', str(t))
    return float(mm.group(1)) if mm else None


def _trac_de_texto(t):
    t = str(t).lower()
    if any(x in t for x in ('4x4', 'awd', '4wd', 'quattro', '4motion', '4matic')):
        return '4x4'
    if any(x in t for x in ('4x2', 'fwd', 'rwd', '2wd')):
        return '4x2'
    return None


def _trans_de_texto(t):
    t = ' ' + str(t).lower() + ' '
    if any(x in t for x in (' at ', ' aut', 'cvt', 'dct', 'eat', 'tiptronic',
                            'stronic', 's-tronic', 'steptronic', 'automat')):
        return 'auto'
    if any(x in t for x in (' mt ', ' mec', 'bvm', 'manual')):
        return 'manual'
    return None


def _tokens_trim(t):
    return {w for w in _re.findall(r'[a-z]{3,}', str(t).lower())
            if w not in _TRIM_STOP}


def _moda(serie):
    try:
        s = serie.dropna().astype(str).str.strip().str.lower()
        s = s[s != '']
        return s.mode().iloc[0] if len(s) else None
    except Exception:
        return None


def construir_perfiles_versiones(dep_src):
    """Perfiles de versiones dominantes por (marca, modelo, gen).
    Dos fuentes, en orden de autoridad:
      1) 'specs': versiones OFICIALES (avisos con ficha) — la cedula.
      2) 'texto': versiones de texto que se repiten mucho (5+ avisos
         escritas igual) — dominantes de facto, utiles ANTES de tener
         fichas. Si un modelo tiene perfiles specs, mandan esos.
    """
    perfiles = {}
    # Fuente 2 primero (la fuente 1 la pisa si existe)
    txt = dep_src[dep_src['_ver'].astype(str).str.len() >= 6]
    for (ma, mo, gn, tv), g in txt.groupby(['marca', 'modelo', 'gen', '_ver']):
        if len(g) < 5:
            continue
        pre_año = {int(a): float(s.median())
                   for a, s in g.groupby('año')['precio'] if len(s) >= 1}
        perfiles.setdefault((ma, mo, gn), []).append({
            'version': str(tv).strip(), 'fuente': 'texto',
            'n': int(len(g)),
            'cil': _cil_de_texto(tv),
            'trac': _trac_de_texto(tv) or _moda(g.get('traccion')),
            'trans': _trans_de_texto(tv) or _moda(g.get('transmision')),
            'comb': _moda(g.get('combustible')),
            'años': (int(g['año'].min()), int(g['año'].max())),
            'precio_año': pre_año,
            'precio_med': float(g['precio'].median()),
            'tokens': _tokens_trim(tv),
        })
    if 'version_oficial' not in dep_src.columns:
        return perfiles
    con_vof = dep_src[dep_src['version_oficial'].notna()
                      & (dep_src['version_oficial'].astype(str).str.strip() != '')]
    claves_specs = set()
    for (ma, mo, gn, vof), g in con_vof.groupby(
            ['marca', 'modelo', 'gen', 'version_oficial']):
        if len(g) < 2:      # perfil necesita al menos 2 avisos con esa cedula
            continue
        claves_specs.add((ma, mo, gn))
        pre_año = {int(a): float(s.median())
                   for a, s in g.groupby('año')['precio'] if len(s) >= 1}
        perfiles.setdefault((ma, mo, gn), []).append({
            'version': str(vof).strip(), 'fuente': 'specs',
            'n': int(len(g)),
            'cil': _cil_de_texto(vof),
            'trac': _trac_de_texto(vof) or _moda(g.get('traccion')),
            'trans': _trans_de_texto(vof) or _moda(g.get('transmision')),
            'comb': _moda(g.get('combustible')),
            'años': (int(g['año'].min()), int(g['año'].max())),
            'precio_año': pre_año,
            'precio_med': float(g['precio'].median()),
            'tokens': _tokens_trim(vof),
        })
    # donde hay perfiles specs, los de texto se retiran (menor autoridad)
    for k in claves_specs:
        perfiles[k] = [p for p in perfiles[k] if p['fuente'] == 'specs']
    return perfiles


def precalificar_grupo(g, fv, ejemplos, candidatos):
    """Puntua el grupo raro contra cada version dominante. Devuelve
    (version, confianza, señales) o None."""
    partes = (fv or '').split('|')
    cil = None
    try:
        cil = float(partes[0]) if partes and partes[0] else None
    except ValueError:
        pass
    texto = ' '.join(ejemplos)
    señales_grupo = {
        'cil': cil or _cil_de_texto(texto),
        'trac': (partes[2] if len(partes) > 2 and partes[2] else None)
                or _trac_de_texto(texto) or _moda(g.get('traccion')),
        'trans': _trans_de_texto(texto) or _moda(g.get('transmision')),
        'comb': _moda(g.get('combustible')),
        'tokens': _tokens_trim(texto),
        'años': (int(g['año'].min()), int(g['año'].max())),
        'precio_año': {int(a): float(s.median())
                       for a, s in g.groupby('año')['precio']},
    }
    _propias = {str(e).strip().lower() for e in ejemplos}
    puntajes = []
    for p in candidatos:
        if p.get('fuente') == 'texto' and p['version'].strip().lower() in _propias:
            continue      # un grupo no se precalifica contra si mismo
        pts, razones = 0, []
        if señales_grupo['cil'] and p['cil']:
            if abs(señales_grupo['cil'] - p['cil']) < 0.05:
                pts += 3; razones.append('cilindrada')
            else:
                pts -= 4      # cilindrada distinta casi descarta
        if señales_grupo['trac'] and p['trac']:
            if señales_grupo['trac'] == p['trac']:
                pts += 2; razones.append('tracción')
            else:
                pts -= 2
        if señales_grupo['trans'] and p['trans']:
            if señales_grupo['trans'] == p['trans']:
                pts += 2; razones.append('transmisión')
            else:
                pts -= 2
        if señales_grupo['comb'] and p['comb']:
            if señales_grupo['comb'] == p['comb']:
                pts += 1; razones.append('combustible')
            else:
                pts -= 3
        a0, a1 = señales_grupo['años']
        if not (a1 < p['años'][0] - 1 or a0 > p['años'][1] + 1):
            pts += 1; razones.append('años')
        # precio: comparar mediana del mismo año cuando existe
        comunes = set(señales_grupo['precio_año']) & set(p['precio_año'])
        ref = None
        if comunes:
            a = max(comunes)
            ref = (señales_grupo['precio_año'][a], p['precio_año'][a])
        elif p['precio_med'] > 0 and señales_grupo['precio_año']:
            ref = (list(señales_grupo['precio_año'].values())[0], p['precio_med'])
        if ref and ref[1] > 0:
            dif = abs(ref[0] - ref[1]) / ref[1]
            if dif <= 0.18:
                pts += 2; razones.append('precio')
            elif dif > 0.40:
                pts -= 2
        inter = señales_grupo['tokens'] & p['tokens']
        if inter:
            pts += min(2, len(inter)); razones.append('nombre')
        puntajes.append((pts, p['n'], p['version'], razones))
    if not puntajes:
        return None
    puntajes.sort(reverse=True)
    mejor = puntajes[0]
    gap = mejor[0] - (puntajes[1][0] if len(puntajes) > 1 else -99)
    if mejor[0] >= 7 and gap >= 3:
        conf = 'alta'
    elif mejor[0] >= 5 and gap >= 2:
        conf = 'media'
    else:
        return None
    fuente = next((p.get('fuente', 'specs') for p in candidatos
                   if p['version'] == mejor[2]), 'specs')
    return (mejor[2], conf, ', '.join(mejor[3]), fuente)


def _norm_modelo_gen(s):
    # clave de modelo para generaciones: sin guiones ni espacios
    return _norm_txt(s).replace('-', '').replace(' ', '')


def _norm_txt(s):
    """minúsculas, sin espacios/guiones/puntos — para casar marca/modelo."""
    return re.sub(r'[\s\-\./]', '', str(s).lower())


_VOF_PERFIL = {}   # (ma,mo,gen,cil,acabs) -> version_oficial dominante
_ACABADOS_MULTI = None  # se arma tras definir _ACABADOS_CANON

def _acabados_multi(txt):
    """Todos los acabados presentes en el texto, en orden canónico fijo
    (así 'limited heritage' y 'heritage limited' dan lo mismo)."""
    global _ACABADOS_MULTI
    if _ACABADOS_MULTI is None:
        _ACABADOS_MULTI = ['heritage', 'xl plus'] + _ACABADOS_CANON + ['xl']
    t = ' ' + str(txt).lower() + ' '
    out = []
    for ac in _ACABADOS_MULTI:
        if ' ' + ac + ' ' in t or t.strip().endswith(' ' + ac):
            out.append(ac)
            t = t.replace(' ' + ac, ' ')   # consumir (evita 'gt' tras 'gt line')
    return out


def _trac_norm(v):
    # 4x4/4wd/awd -> 4x4 ; 4x2/2wd/rwd/fwd -> 4x2 ; sino ''
    t = str(v or '').lower()
    if any(x in t for x in ('4x4', '4wd', 'awd')):
        return '4x4'
    if any(x in t for x in ('4x2', '2wd', 'rwd', 'fwd')):
        return '4x2'
    return ''


import re as _re_mod


def _comb_norm(v):
    t = str(v or '').lower()
    if 'diesel' in t or 'diésel' in t:
        return 'diesel'
    if 'bencina' in t or 'gasolina' in t:
        return 'bencina'
    if 'hibrid' in t or 'hybrid' in t:
        return 'hibrido'
    if 'electric' in t or 'eléctric' in t:
        return 'electrico'
    if 'glp' in t or 'gnc' in t or ' gas' in t:
        return 'gas'
    return ''


def _cab_de_txt(v):
    # cabina simple vs doble (pickups); intermedia/desconocida -> ''
    t = ' ' + str(v or '').lower() + ' '
    doble = any(x in t for x in ('dob. cab', 'doble cab', 'dob cab', 'cab. doble',
                                 'cabina doble', ' d/c', ' dc ', ' cc ', 'crew',
                                 'king cab', 'super cab', 'supercab'))
    simple = any(x in t for x in ('cab. simple', 'cabina simple', 'cab simple',
                                  ' c/s', ' rc ', ' cs ', 'reg cab', 'reg. cab',
                                  'regular cab'))
    if doble and not simple:
        return 'doble'
    if simple and not doble:
        return 'simple'
    return ''


def _body_de_txt(v):
    # coupe | 2p (2-3 puertas) | 4p (4-5 puertas) | ''
    t = ' ' + str(v or '').lower() + ' '
    if 'coupe' in t or 'coupé' in t:
        return 'coupe'
    ps = _re_mod.findall(r'\b([2-5])\s*p(?:tas|uertas)?\b\.?', t)
    if ps:
        n = int(ps[0])
        return '2p' if n <= 3 else '4p'
    return ''


def _caja_de_txt(v):
    # AT/MT desde texto libre; ambiguo -> ''
    t = ' ' + str(v or '').lower() + ' '
    at = ('automatic' in t) or ('automátic' in t) or (' at ' in t) or (' aut ' in t)
    mt = ('mecanic' in t) or ('mecánic' in t) or ('manual' in t) or (' mt ' in t)
    if at and not mt:
        return 'AT'
    if mt and not at:
        return 'MT'
    return ''


def _mot_de_txt(v):
    # cilindrada desde texto libre solo si aparece UNA sola (1.0-6.9)
    hallados = set(_re_mod.findall(r'\b([1-6]\.[0-9])\b', str(v or '')))
    return hallados.pop() if len(hallados) == 1 else ''


_COMENT_TXT = {}   # id -> texto del comentario (para deducir traccion)
_UBI = {}          # id -> {region, comuna, fecha_pub} desde ubicaciones.csv


def _cargar_ubicaciones():
    if _UBI:
        return
    try:
        if os.path.exists('ubicaciones.csv'):
            import csv as _csv
            with open('ubicaciones.csv', encoding='utf-8', newline='') as _f:
                for _row in _csv.DictReader(_f):
                    _iu = str(_row.get('id', '')).lower().strip()
                    if _iu:
                        _UBI[_iu] = {'region': _row.get('region', ''),
                                     'comuna': _row.get('comuna', ''),
                                     'fecha_pub': _row.get('fecha_pub', '')}
            log(f"  Ubicaciones cargadas: {len(_UBI)}")
    except Exception as _e:
        log(f"  (ubicaciones.csv ilegible: {_e})")



def _cargar_comentarios_txt():
    # idempotente: puebla _COMENT_TXT desde comentarios.json si esta vacio
    if _COMENT_TXT:
        return
    try:
        if os.path.exists('comentarios.json'):
            with open('comentarios.json', encoding='utf-8') as _f:
                _cj = json.load(_f)
            for _k, _v in (_cj.items() if isinstance(_cj, dict) else []):
                _t = _v if isinstance(_v, str) else (_v or {}).get('comentario', '')
                if _t:
                    _COMENT_TXT[str(_k).lower().strip()] = str(_t)[:400]
    except Exception:
        pass



def _trac_unifica(s):
    # unifica sinonimos de traccion en el texto final de la sugerida
    out = str(s)
    for a, b in (('4WD', '4x4'), ('4wd', '4x4'), ('AWD', '4x4'), ('awd', '4x4'),
                 ('2WD', '4x2'), ('2wd', '4x2')):
        out = out.replace(a, b)
    return out


_CONSOLIDA = {}   # (modelo_key, gen, sugerida_cruda) -> sugerida_dominante
_MOTOR_DOM = {}   # (modelo_key, gen) -> motor dominante (>=75% de los declarados)


def _parse_sug(s):
    # componentes del formato canonico: motor, acabados, trac, comb, caja
    t = str(s or '').replace('\u2248', ' ').strip().lower()
    toks = t.split()
    mot = next((x for x in toks if _re_mod.fullmatch(r'[1-6]\.[0-9]', x)), '')
    tr = next((x for x in toks if x in ('4x4', '4x2')), '')
    cb = next((x for x in toks if x in ('diesel', 'bencina', 'hibrido',
                                        'electrico', 'gas')), '')
    cj = next((x for x in toks if x in ('at', 'mt')), '')
    cab = next((x for x in toks if x in ('cab.simple', 'cab.doble')), '')
    usados = {mot, tr, cb, cj, cab}
    acab = tuple(x for x in toks if x not in usados)
    return mot, acab, cab, tr, cb, cj


def _compatible(sub, dom):
    # cada componente PRESENTE en sub debe coincidir en dom
    for a, b in zip(sub, dom):
        if isinstance(a, tuple):
            if a and a != b:
                return False
        elif a and a != b:
            return False
    return True


_SEGMENTOS = {}

# Tabla de segmentos EMBEBIDA en el sistema (no requiere archivos externos).
# Si existe un segmentos.xlsx en la carpeta, sus filas corrigen/agregan.
_SEGMENTOS_BASE = {
    ('chevrolet', 'spark'): 'city car',
    ('kia', 'morning'): 'city car',
    ('kia', 'picanto'): 'city car',
    ('suzuki', 'alto'): 'city car',
    ('suzuki', 'celerio'): 'city car',
    ('suzuki', 's-presso'): 'city car',
    ('suzuki', 'ignis'): 'city car',
    ('hyundai', 'i10'): 'city car',
    ('hyundai', 'grand i10'): 'city car',
    ('renault', 'kwid'): 'city car',
    ('mitsubishi', 'mirage'): 'city car',
    ('mitsubishi', 'space star'): 'city car',
    ('fiat', 'mobi'): 'city car',
    ('fiat', '500'): 'city car',
    ('chery', 'qq'): 'city car',
    ('volkswagen', 'up'): 'city car',
    ('chevrolet', 'onix'): 'hatchback',
    ('toyota', 'yaris sport'): 'hatchback',
    ('hyundai', 'i20'): 'hatchback',
    ('hyundai', 'i30'): 'hatchback',
    ('kia', 'rio 5'): 'hatchback',
    ('nissan', 'tiida'): 'hatchback',
    ('nissan', 'note'): 'hatchback',
    ('nissan', 'march'): 'hatchback',
    ('suzuki', 'swift'): 'hatchback',
    ('suzuki', 'baleno'): 'hatchback',
    ('suzuki', 'sx4'): 'hatchback',
    ('ford', 'fiesta'): 'hatchback',
    ('ford', 'focus'): 'hatchback',
    ('peugeot', '208'): 'hatchback',
    ('peugeot', '206'): 'hatchback',
    ('peugeot', '207'): 'hatchback',
    ('peugeot', '308'): 'hatchback',
    ('mazda', '2'): 'hatchback',
    ('mazda', 'demio'): 'hatchback',
    ('renault', 'clio'): 'hatchback',
    ('renault', 'sandero'): 'hatchback',
    ('renault', 'stepway'): 'hatchback',
    ('volkswagen', 'gol'): 'hatchback',
    ('volkswagen', 'polo'): 'hatchback',
    ('volkswagen', 'golf'): 'hatchback',
    ('volkswagen', 'fox'): 'hatchback',
    ('volkswagen', 'escarabajo'): 'hatchback',
    ('honda', 'fit'): 'hatchback',
    ('citroën', 'c3'): 'hatchback',
    ('citroen', 'c3'): 'hatchback',
    ('citroën', 'c4'): 'hatchback',
    ('citroen', 'c4'): 'hatchback',
    ('fiat', 'uno'): 'hatchback',
    ('fiat', 'palio'): 'hatchback',
    ('fiat', 'argo'): 'hatchback',
    ('mg', '3'): 'hatchback',
    ('mg', 'mg3'): 'hatchback',
    ('audi', 'a1'): 'hatchback',
    ('audi', 'a3'): 'hatchback',
    ('volvo', 'v40'): 'hatchback',
    ('mercedes-benz', 'clase a'): 'hatchback',
    ('bmw', 'serie 1'): 'hatchback',
    ('great wall', 'florid'): 'hatchback',
    ('hyundai', 'veloster'): 'hatchback',
    ('chevrolet', 'sail'): 'sedan',
    ('chevrolet', 'cruze'): 'sedan',
    ('chevrolet', 'aveo'): 'sedan',
    ('chevrolet', 'optra'): 'sedan',
    ('chevrolet', 'corsa'): 'sedan',
    ('chevrolet', 'prisma'): 'sedan',
    ('toyota', 'corolla'): 'sedan',
    ('toyota', 'yaris'): 'sedan',
    ('toyota', 'tercel'): 'sedan',
    ('toyota', 'etios'): 'sedan',
    ('hyundai', 'accent'): 'sedan',
    ('hyundai', 'elantra'): 'sedan',
    ('hyundai', 'sonata'): 'sedan',
    ('hyundai', 'genesis'): 'sedan',
    ('kia', 'rio'): 'sedan',
    ('kia', 'rio 4'): 'sedan',
    ('kia', 'cerato'): 'sedan',
    ('kia', 'optima'): 'sedan',
    ('kia', 'k3'): 'sedan',
    ('nissan', 'v16'): 'sedan',
    ('nissan', 'versa'): 'sedan',
    ('nissan', 'sentra'): 'sedan',
    ('honda', 'city'): 'sedan',
    ('honda', 'civic'): 'sedan',
    ('honda', 'accord'): 'sedan',
    ('mazda', '3'): 'sedan',
    ('mazda', '6'): 'sedan',
    ('peugeot', '301'): 'sedan',
    ('peugeot', '408'): 'sedan',
    ('renault', 'symbol'): 'sedan',
    ('renault', 'megane'): 'sedan',
    ('renault', 'mégane'): 'sedan',
    ('renault', 'megane iii'): 'sedan',
    ('renault', 'fluence'): 'sedan',
    ('renault', 'logan'): 'sedan',
    ('volkswagen', 'virtus'): 'sedan',
    ('volkswagen', 'vento'): 'sedan',
    ('volkswagen', 'bora'): 'sedan',
    ('volkswagen', 'jetta'): 'sedan',
    ('volkswagen', 'passat'): 'sedan',
    ('mitsubishi', 'lancer'): 'sedan',
    ('subaru', 'impreza'): 'sedan',
    ('subaru', 'legacy'): 'sedan',
    ('citroën', 'c-elysee'): 'sedan',
    ('citroen', 'c-elysee'): 'sedan',
    ('citroën', 'c5'): 'sedan',
    ('citroen', 'c5'): 'sedan',
    ('fiat', 'cronos'): 'sedan',
    ('mg', '5'): 'sedan',
    ('mg', '6'): 'sedan',
    ('chery', 'arrizo'): 'sedan',
    ('changan', 'alsvin'): 'sedan',
    ('samsung', 'sm3'): 'sedan',
    ('samsung', 'sm5'): 'sedan',
    ('samsung', 'sm6'): 'sedan',
    ('bmw', '316'): 'sedan',
    ('bmw', '318'): 'sedan',
    ('bmw', '320i'): 'sedan',
    ('bmw', 'serie 3'): 'sedan',
    ('mercedes-benz', '180'): 'sedan',
    ('mercedes-benz', 'c 250'): 'sedan',
    ('mercedes-benz', 'clase c'): 'sedan',
    ('audi', 'a4'): 'sedan',
    ('audi', 'a5'): 'sedan',
    ('audi', 'a6'): 'sedan',
    ('volvo', 's60'): 'sedan',
    ('jaguar', 'xf'): 'sedan',
    ('subaru', 'outback'): 'station wagon',
    ('volvo', 'v60'): 'station wagon',
    ('peugeot', '308 sw'): 'station wagon',
    ('chevrolet', 'tracker'): 'suv',
    ('chevrolet', 'captiva'): 'suv',
    ('chevrolet', 'captiva ii'): 'suv',
    ('chevrolet', 'orlando'): 'suv',
    ('chevrolet', 'tahoe'): 'suv',
    ('chevrolet', 'blazer'): 'suv',
    ('chevrolet', 'groove'): 'suv',
    ('toyota', 'rav4'): 'suv',
    ('toyota', '4runner'): 'suv',
    ('toyota', 'land cruiser'): 'suv',
    ('toyota', 'prado'): 'suv',
    ('toyota', 'fortuner'): 'suv',
    ('toyota', 'urban cruiser'): 'suv',
    ('toyota', 'raize'): 'suv',
    ('toyota', 'corolla cross'): 'suv',
    ('toyota', 'sw4'): 'suv',
    ('hyundai', 'tucson'): 'suv',
    ('hyundai', 'santa fe'): 'suv',
    ('hyundai', 'grand santa fe'): 'suv',
    ('hyundai', 'creta'): 'suv',
    ('hyundai', 'kona'): 'suv',
    ('hyundai', 'venue'): 'suv',
    ('hyundai', 'terracan'): 'suv',
    ('hyundai', 'galloper'): 'suv',
    ('hyundai', 'veracruz'): 'suv',
    ('hyundai', 'palisade'): 'suv',
    ('kia', 'sportage'): 'suv',
    ('kia', 'sorento'): 'suv',
    ('kia', 'soul'): 'suv',
    ('kia', 'seltos'): 'suv',
    ('kia', 'stonic'): 'suv',
    ('kia', 'niro'): 'suv',
    ('kia', 'mohave'): 'suv',
    ('kia', 'carens'): 'suv',
    ('nissan', 'qashqai'): 'suv',
    ('nissan', 'x-trail'): 'suv',
    ('nissan', 'xtrail'): 'suv',
    ('nissan', 'murano'): 'suv',
    ('nissan', 'pathfinder'): 'suv',
    ('nissan', 'kicks'): 'suv',
    ('nissan', 'juke'): 'suv',
    ('nissan', 'patrol'): 'suv',
    ('nissan', 'terrano'): 'suv',
    ('suzuki', 'grand nomade'): 'suv',
    ('suzuki', 'grand vitara'): 'suv',
    ('suzuki', 'vitara'): 'suv',
    ('suzuki', 'jimny'): 'suv',
    ('suzuki', 'ertiga'): 'suv',
    ('suzuki', 'scross'): 'suv',
    ('ford', 'escape'): 'suv',
    ('ford', 'ecosport'): 'suv',
    ('ford', 'edge'): 'suv',
    ('ford', 'explorer'): 'suv',
    ('ford', 'expedition'): 'suv',
    ('ford', 'territory'): 'suv',
    ('ford', 'bronco'): 'suv',
    ('peugeot', '2008'): 'suv',
    ('peugeot', '3008'): 'suv',
    ('peugeot', '5008'): 'suv',
    ('mazda', 'cx-3'): 'suv',
    ('mazda', 'cx3'): 'suv',
    ('mazda', 'cx-30'): 'suv',
    ('mazda', 'cx30'): 'suv',
    ('mazda', 'cx-5'): 'suv',
    ('mazda', 'cx5'): 'suv',
    ('mazda', 'cx-9'): 'suv',
    ('mazda', 'cx9'): 'suv',
    ('renault', 'duster'): 'suv',
    ('renault', 'captur'): 'suv',
    ('renault', 'koleos'): 'suv',
    ('volkswagen', 'tiguan'): 'suv',
    ('volkswagen', 't-cross'): 'suv',
    ('volkswagen', 'tcross'): 'suv',
    ('volkswagen', 'taos'): 'suv',
    ('volkswagen', 'touareg'): 'suv',
    ('honda', 'hr-v'): 'suv',
    ('honda', 'hrv'): 'suv',
    ('honda', 'cr-v'): 'suv',
    ('honda', 'crv'): 'suv',
    ('honda', 'pilot'): 'suv',
    ('honda', 'wr-v'): 'suv',
    ('mitsubishi', 'asx'): 'suv',
    ('mitsubishi', 'outlander'): 'suv',
    ('mitsubishi', 'montero'): 'suv',
    ('mitsubishi', 'montero sport'): 'suv',
    ('mitsubishi', 'eclipse cross'): 'suv',
    ('ssangyong', 'korando'): 'suv',
    ('ssangyong', 'actyon'): 'suv',
    ('ssangyong', 'rexton'): 'suv',
    ('ssangyong', 'tivoli'): 'suv',
    ('subaru', 'forester'): 'suv',
    ('subaru', 'xv'): 'suv',
    ('subaru', 'tribeca'): 'suv',
    ('subaru', 'evoltis'): 'suv',
    ('citroën', 'c4 cactus'): 'suv',
    ('citroen', 'c4 cactus'): 'suv',
    ('fiat', 'pulse'): 'suv',
    ('fiat', 'fastback'): 'suv',
    ('jeep', 'renegade'): 'suv',
    ('jeep', 'compass'): 'suv',
    ('jeep', 'cherokee'): 'suv',
    ('jeep', 'grand cherokee'): 'suv',
    ('jeep', 'wrangler'): 'suv',
    ('jeep', 'commander'): 'suv',
    ('dodge', 'journey'): 'suv',
    ('dodge', 'durango'): 'suv',
    ('mg', 'zs'): 'suv',
    ('mg', 'rx5'): 'suv',
    ('mg', 'hs'): 'suv',
    ('chery', 'tiggo'): 'suv',
    ('chery', 'tiggo 2'): 'suv',
    ('chery', 'tiggo 3'): 'suv',
    ('jac', 's2'): 'suv',
    ('jac', 's3'): 'suv',
    ('great wall', 'haval'): 'suv',
    ('great wall', 'haval 3'): 'suv',
    ('great wall', 'h6'): 'suv',
    ('changan', 'cs35'): 'suv',
    ('changan', 'cs35 plus'): 'suv',
    ('changan', 'cs55'): 'suv',
    ('mahindra', 'xuv 500'): 'suv',
    ('mahindra', 'scorpio'): 'suv',
    ('samsung', 'qm5'): 'suv',
    ('samsung', 'qm6'): 'suv',
    ('bmw', 'x1'): 'suv',
    ('bmw', 'x3'): 'suv',
    ('bmw', 'x5'): 'suv',
    ('bmw', 'x6'): 'suv',
    ('mercedes-benz', 'gla'): 'suv',
    ('mercedes-benz', 'glc'): 'suv',
    ('audi', 'q2'): 'suv',
    ('audi', 'q3'): 'suv',
    ('audi', 'q5'): 'suv',
    ('audi', 'q7'): 'suv',
    ('volvo', 'xc60'): 'suv',
    ('volvo', 'xc90'): 'suv',
    ('land rover', 'discovery'): 'suv',
    ('land rover', 'range rover'): 'suv',
    ('land rover', 'defender'): 'suv',
    ('land rover', 'freelander'): 'suv',
    ('land rover', 'evoque'): 'suv',
    ('hummer', 'h2'): 'suv',
    ('hummer', 'h3'): 'suv',
    ('haval', 'h6'): 'suv',
    ('haval', 'jolion'): 'suv',
    ('fiat', 'strada'): 'camioneta pequeña',
    ('fiat', 'toro'): 'camioneta pequeña',
    ('ram', '700'): 'camioneta pequeña',
    ('ram', '1000'): 'camioneta pequeña',
    ('renault', 'oroch'): 'camioneta pequeña',
    ('volkswagen', 'saveiro'): 'camioneta pequeña',
    ('suzuki', 'carry'): 'camioneta pequeña',
    ('ford', 'courier'): 'camioneta pequeña',
    ('foton', 'tm3'): 'camioneta pequeña',
    ('changan', 'md201'): 'camioneta pequeña',
    ('chevrolet', 'montana'): 'camioneta pequeña',
    ('toyota', 'hilux'): 'camioneta mediana',
    ('nissan', 'np300'): 'camioneta mediana',
    ('nissan', 'navara'): 'camioneta mediana',
    ('nissan', 'frontier'): 'camioneta mediana',
    ('nissan', 'd21'): 'camioneta mediana',
    ('chevrolet', 'colorado'): 'camioneta mediana',
    ('chevrolet', 'dmax'): 'camioneta mediana',
    ('chevrolet', 'd-max'): 'camioneta mediana',
    ('chevrolet', 'luv'): 'camioneta mediana',
    ('ford', 'ranger'): 'camioneta mediana',
    ('mitsubishi', 'l200'): 'camioneta mediana',
    ('mitsubishi', 'katana'): 'camioneta mediana',
    ('mazda', 'bt-50'): 'camioneta mediana',
    ('mazda', 'bt50'): 'camioneta mediana',
    ('volkswagen', 'amarok'): 'camioneta mediana',
    ('ssangyong', 'actyon sports'): 'camioneta mediana',
    ('ssangyong', 'musso'): 'camioneta mediana',
    ('renault', 'alaskan'): 'camioneta mediana',
    ('peugeot', 'landtrek'): 'camioneta mediana',
    ('jac', 't6'): 'camioneta mediana',
    ('jac', 't8'): 'camioneta mediana',
    ('great wall', 'wingle'): 'camioneta mediana',
    ('great wall', 'poer'): 'camioneta mediana',
    ('maxus', 't60'): 'camioneta mediana',
    ('foton', 'tunland'): 'camioneta mediana',
    ('mahindra', 'pik up'): 'camioneta mediana',
    ('mahindra', 'genio'): 'camioneta mediana',
    ('dodge', 'dakota'): 'camioneta mediana',
    ('jeep', 'gladiator'): 'camioneta mediana',
    ('kia', 'frontier'): 'camioneta mediana',
    ('changan', 'hunter'): 'camioneta mediana',
    ('toyota', 'tacoma'): 'camioneta mediana',
    ('ford', 'f-150'): 'camioneta grande',
    ('ford', 'f150'): 'camioneta grande',
    ('ford', 'f-150 raptor'): 'camioneta grande',
    ('toyota', 'tundra'): 'camioneta grande',
    ('chevrolet', 'silverado'): 'camioneta grande',
    ('ram', '1500'): 'camioneta grande',
    ('ram', '2500'): 'camioneta grande',
    ('dodge', 'ram'): 'camioneta grande',
    ('hummer', 'h2t'): 'camioneta grande',
    ('hummer', 'h3t'): 'camioneta grande',
    ('gmc', 'sierra'): 'camioneta grande',
    ('peugeot', 'partner'): 'furgon/van',
    ('peugeot', 'boxer'): 'furgon/van',
    ('peugeot', 'expert'): 'furgon/van',
    ('peugeot', 'rifter'): 'furgon/van',
    ('citroën', 'berlingo'): 'furgon/van',
    ('citroen', 'berlingo'): 'furgon/van',
    ('citroën', 'jumpy'): 'furgon/van',
    ('citroen', 'jumpy'): 'furgon/van',
    ('citroën', 'jumper'): 'furgon/van',
    ('citroen', 'jumper'): 'furgon/van',
    ('renault', 'kangoo'): 'furgon/van',
    ('renault', 'master'): 'furgon/van',
    ('renault', 'trafic'): 'furgon/van',
    ('fiat', 'fiorino'): 'furgon/van',
    ('fiat', 'ducato'): 'furgon/van',
    ('volkswagen', 'transporter'): 'furgon/van',
    ('volkswagen', 'crafter'): 'furgon/van',
    ('hyundai', 'h1'): 'furgon/van',
    ('hyundai', 'staria'): 'furgon/van',
    ('nissan', 'urvan'): 'furgon/van',
    ('kia', 'carnival'): 'furgon/van',
    ('ssangyong', 'stavic'): 'furgon/van',
    ('ssangyong', 'istana'): 'furgon/van',
    ('suzuki', 'apv'): 'furgon/van',
    ('mercedes-benz', 'sprinter'): 'furgon/van',
    ('mercedes-benz', 'vito'): 'furgon/van',
    ('honda', 'odyssey'): 'furgon/van',
    ('dodge', 'grand caravan'): 'furgon/van',
    ('ford', 'transit'): 'furgon/van',
    ('maxus', 'g10'): 'furgon/van',
    ('jac', 'refine'): 'furgon/van',
    ('toyota', 'hiace'): 'furgon/van',
    ('chevrolet', 'n300'): 'furgon/van',
    ('chevrolet', 'n400'): 'furgon/van',
    ('hyundai', 'porter'): 'camion',
    ('kia', 'frontier ii'): 'camion',
    ('jac', 'x200'): 'camion',
    ('foton', 'aumark'): 'camion',
    ('toyota', 'coaster'): 'camion',
    ('hyundai', 'h100'): 'camion',
    ('mitsubishi', 'canter'): 'camion',
    ('chevrolet', 'nkr'): 'camion',
    ('chevrolet', 'npr'): 'camion',
    ('ford', 'mustang'): 'deportivo',
    ('dodge', 'challenger'): 'deportivo',
    ('dodge', 'charger'): 'deportivo',
    ('subaru', 'wrx'): 'deportivo',
    ('chevrolet', 'camaro'): 'deportivo',
    ('toyota', '86'): 'deportivo',
    ('nissan', '370z'): 'deportivo',
}


def cargar_segmentos():
    _SEGMENTOS.clear()
    for (ma, mo), seg in _SEGMENTOS_BASE.items():
        _SEGMENTOS[f"{_norm_txt(ma)}|{_norm_modelo_gen(mo)}"] = seg
    if os.path.exists('segmentos.xlsx'):
        try:
            gs = pd.read_excel('segmentos.xlsx')
            gs.columns = [str(c).strip().upper() for c in gs.columns]
            for _, r in gs.iterrows():
                k = f"{_norm_txt(r['MARCA'])}|{_norm_modelo_gen(str(r['MODELO']))}"
                _SEGMENTOS[k] = str(r['SEGMENTO']).strip()
        except Exception:
            pass
    log(f"  Segmentos cargados: {len(_SEGMENTOS)} modelos (tabla interna)")


def _seg_carroceria(marca, modelo):
    return _SEGMENTOS.get(
        f"{_norm_txt(marca)}|{_norm_modelo_gen(str(modelo))}", '')


def construir_motor_dominante(dfa):
    # Regla del usuario: sin motor declarado y con un motor claramente
    # dominante en el (modelo, gen), se hereda (ej: F-150 13ra -> 3.3).
    from collections import defaultdict
    _MOTOR_DOM.clear()
    cnt = defaultdict(lambda: defaultdict(int))
    _man = dfa['marca'].map(_norm_txt)
    _mon = dfa['modelo'].map(_norm_txt)
    for _i, r in dfa.iterrows():
        mot = r.get('motor')
        if pd.notna(mot) and str(mot).strip():
            cnt[(f"{_man.at[_i]}|{_mon.at[_i]}",
                 str(r.get('gen', '') or ''))][str(mot)] += 1
    for k, motores in cnt.items():
        total = sum(motores.values())
        if total < 8:
            continue
        mejor = max(motores, key=motores.get)
        if motores[mejor] / total >= 0.75:
            _MOTOR_DOM[k] = mejor
    log(f"  Motor dominante heredable: {len(_MOTOR_DOM)} modelo/generacion")


def _hereda_motor(s, modelo_key, gen):
    # antepone el motor dominante si la sugerida no lo trae (marca deduccion)
    if not s:
        return s
    crudo = s.replace('\u2248 ', '')
    if _re_mod.match(r'[1-6]\.[0-9]\b', crudo):
        return s
    dom = _MOTOR_DOM.get((modelo_key, str(gen or '')))
    if not dom:
        return s
    return ('\u2248 ' + dom + ' ' + crudo)[:40]


def construir_consolidacion(dfa):
    """Mapa de absorcion: sugeridas incompletas de poco peso hacia la
    dominante compatible del (modelo, gen). Unica compatible -> absorbe;
    varias -> solo si la mayor concentra >=75% del peso entre compatibles."""
    from collections import defaultdict
    _CONSOLIDA.clear()
    conteo = defaultdict(lambda: defaultdict(int))
    _man = dfa['marca'].map(_norm_txt)
    _mon = dfa['modelo'].map(_norm_txt)
    for _i, r in dfa.iterrows():
        s = _hereda_motor(_sugerida_de(r),
                          f"{_man.at[_i]}|{_mon.at[_i]}", r.get('gen'))
        if s:
            conteo[(f"{_man.at[_i]}|{_mon.at[_i]}", str(r.get('gen', '') or ''))][s.replace('\u2248 ', '')] += 1
    n_abs = 0
    for key, vers in conteo.items():
        doms = {v: n for v, n in vers.items() if n >= 5}
        if not doms:
            continue
        p_doms = {v: _parse_sug(v) for v in doms}
        for v, n in vers.items():
            if n > 3 or v in doms:
                continue
            pv = _parse_sug(v)
            comp_falta = (not pv[0]) or (not pv[1]) or (not pv[2]) or (not pv[3]) or (not pv[4])
            if not comp_falta:
                continue
            compat = {dv: doms[dv] for dv, pd_ in p_doms.items() if _compatible(pv, pd_)}
            if not compat:
                continue
            total = sum(compat.values())
            mejor = max(compat, key=compat.get)
            if len(compat) == 1 or compat[mejor] / total >= 0.75:
                _CONSOLIDA[(key[0], key[1], v)] = mejor
                n_abs += 1
    log(f"  Consolidacion de versiones: {n_abs} sugeridas incompletas "
        f"absorbidas por su dominante")


def _sugerida_final(p, modelo_key):
    s = _hereda_motor(_sugerida_de(p), modelo_key, p.get('gen'))
    if not s:
        return s
    crudo = s.replace('\u2248 ', '')
    dom = _CONSOLIDA.get((modelo_key, str(p.get('gen', '') or ''), crudo))
    if dom:
        return ('\u2248 ' + dom)[:36]
    return s


_KW_NO_OPORTUNIDAD = (
    # estado mecanico grave
    'para armar', 'para reparar', 'necesita reparacion', 'necesita reparación',
    'no anda', 'no enciende', 'no funciona', 'no camina', 'no da partida',
    'no parte', 'motor fundido', 'fundido', 'fundida', 'sin motor',
    'motor malo', 'caja mala', 'caja de cambios mala', 'culata mala',
    'junta quemada', 'desarmado', 'desarmada', 'en desarme', 'para desarme',
    # siniestros
    'chocado', 'chocada', 'volcado', 'volcada', 'volcamiento', 'siniestrado',
    'perdida total', 'pérdida total', 'para repuesto', 'por piezas',
    'por partes', 'en partes',
    # papeles y deudas
    'sin papeles', 'sin documentos', 'documentos atrasados', 'con prenda',
    'con embargo', 'embargado', 'con multas', 'papeles atrasados',
    'revision tecnica vencida', 'revisión técnica vencida',
    'permiso de circulacion vencido', 'permiso de circulación vencido',
    'sin llave', 'una llave rota',
)

_NEGACIONES = ('no ', 'nunca ', 'jamas ', 'jamás ', 'ni ', 'sin ', '0 ', 'cero ')


def _comentario_no_oportunidad(vid):
    # Detecta en el comentario senales de que el descuento tiene explicacion
    # (danio, para armar, papeles) -> fuera del oro. Respeta negaciones:
    # "impecable, no chocado" o "nunca chocado" NO excluyen.
    txt = _COMENT_TXT.get(str(vid).lower().strip(), '')
    if not txt:
        return ''
    t = ' ' + txt.lower() + ' '
    for kw in _KW_NO_OPORTUNIDAD:
        pos = t.find(kw)
        while pos != -1:
            previo = t[max(0, pos - 12):pos]
            if not any(previo.endswith(n) for n in _NEGACIONES):
                return kw
            pos = t.find(kw, pos + 1)
    return ''


def _sugerida_de(p):
    # Version sugerida HOMOLOGADA (regla del usuario 2026-07-10):
    # orden canonico  motor + acabado + traccion + combustible + caja,
    # sin ruido (4p, dob. cab., puntos), completando faltantes con los
    # datos efectivos del aviso. Prefijo ~ para deducciones (sin ficha).
    vo = p.get('version_oficial')
    tiene_ficha = pd.notna(vo) and str(vo).strip().lower() not in ('', 'nan')
    mot = p.get('motor')
    mot = str(mot) if pd.notna(mot) else ''
    acabs = _acabados_multi(p.get('version'))
    fuente = None
    pref = ''
    if tiene_ficha:
        fuente = str(vo)
    else:
        _k = (_norm_txt(p.get('marca')), _norm_txt(p.get('modelo')),
              str(p.get('gen', '') or ''), mot, tuple(acabs))
        _vof = _VOF_PERFIL.get(_k)
        if _vof:
            fuente = str(_vof)
            pref = '\u2248 '
    if fuente is not None:
        s = fuente.lower()
        f_mot = _mot_de_txt(s) or mot
        f_ac = _acabados_multi(s) or acabs
        f_tr = _trac_norm(s) or str(p.get('_trc', '') or '')
        f_cb = _comb_norm(s) or _comb_norm(p.get('combustible'))
        f_cj = _caja_de_txt(s) or str(p.get('_cj_ef', '') or '')
    else:
        pref = '\u2248 '
        f_mot = mot
        f_ac = acabs
        f_tr = str(p.get('_trc', '') or '')
        f_cb = _comb_norm(p.get('combustible'))
        f_cj = str(p.get('_cj_ef', '') or '')
        if not (f_mot or f_ac):
            return ''
    f_cab = ''
    if fuente is not None:
        f_cab = _cab_de_txt(fuente) or _cab_de_txt(p.get('version')) \
                or str(p.get('_cab_ef', '') or '')
    else:
        f_cab = _cab_de_txt(p.get('version')) or str(p.get('_cab_ef', '') or '')
    f_cab = {'simple': 'cab.simple', 'doble': 'cab.doble'}.get(f_cab, '')
    partes = [x for x in (f_mot, ' '.join(f_ac), f_cab, f_tr, f_cb,
                          str(f_cj).lower()) if x]
    return (pref + ' '.join(partes))[:40]

def cargar_generaciones(ruta="generaciones.xlsx"):
    """
    Lee la tabla editable de generaciones (marca, modelo, generacion,
    año_desde, año_hasta). Devuelve {(marca_norm, modelo_alias_norm):
    [(desde, hasta, etiqueta), ...]}. El campo MODELO admite alias
    separados por '/' (ej: 'Navara / NP300').
    """
    if not os.path.exists(ruta):
        log(f"  AVISO: no existe {ruta} — TODAS las generaciones saldrán "
            f"vacías ('—'). Poné el archivo en la carpeta del sistema.")
        return {}
    try:
        gdf = pd.read_excel(ruta, sheet_name="Generaciones", dtype=str)
    except ValueError:
        # el archivo existe pero no tiene la hoja "Generaciones" (p.ej. se
        # regeneró mal): probar la primera hoja antes de rendirse.
        gdf = pd.read_excel(ruta, dtype=str)
    except Exception as e:
        log(f"  (no pude leer {ruta}: {e} — sigo sin generaciones)")
        return {}
    tabla = {}
    for _, r in gdf.iterrows():
        ma = _norm_txt(r.get('MARCA', ''))
        gen = str(r.get('GENERACION', '') or '').strip()
        try:
            desde = int(float(r.get('AÑO_DESDE')))
        except (TypeError, ValueError):
            continue
        hasta_raw = str(r.get('AÑO_HASTA', '') or '').strip()
        try:
            hasta = int(float(hasta_raw)) if hasta_raw else 9999
        except ValueError:
            hasta = 9999
        for alias in str(r.get('MODELO', '')).split('/'):
            alias = _norm_txt(alias)
            if ma and alias and gen:
                tabla.setdefault((ma, _norm_modelo_gen(alias)), []).append((desde, hasta, gen))
    return tabla


def asignar_generacion(df, tabla_gen):
    """Columna 'gen': etiqueta de generación según la tabla, o '' si el
    modelo no está en la tabla / el año no cae en ningún rango."""
    if not tabla_gen:
        df['gen'] = ''
        return df
    marcas_n = df['marca'].map(_norm_txt)
    modelos_n = df['modelo'].map(_norm_modelo_gen)
    gens = []
    for ma, mo, yr in zip(marcas_n, modelos_n, df['año']):
        rangos = tabla_gen.get((ma, mo))
        etiqueta = ''
        if rangos and pd.notna(yr):
            for d, h, g in rangos:
                if d <= yr <= h:
                    etiqueta = g
                    break
        gens.append(etiqueta)
    df['gen'] = gens
    n = sum(1 for g in gens if g)
    log(f"  Generaciones asignadas: {n} de {len(gens)} avisos "
        f"({len(tabla_gen)} modelos en la tabla).")
    return df


_ACABADOS_CANON = ['raptor', 'xlt', 'elite', 'luxury', 'advance', 'pro', 'dx', 'deluxe', 'katana', 'dynamic', 'xe', 'poer', 'platinum', 'lariat', 'dlx', 'country', 'nomade', 'sel', 'xs', 'trend', 'gs', 'premium', 'sportback', 'elegance', 'sr', 'executive', 'prestige', 'ambiente', 'sahara', 'activ', 'denali', 'ghia', 'gt line', 'gt-line', 'gtline', 'gt', 'allure', 'active',
    'premier', 'premiere', 'feel', 'style', 'access', 'business', 'signature',
    'sport', 'limited', 'ltz', 'lt', 'ls', 'exclusive', 'dynamique', 'zen',
    'intens', 'life', 'trend', 'comfort', 'titanium', 'ambiente', 'highline',
    'comfortline', 'trendline', 'xei', 'xli', 'gli', 'gls', 'glx', 'gl', 'ex',
    'exl', 'lx', 'touring', 'value', 'full', 'sense']


def _acabado_txt(txt):
    t = ' ' + str(txt).lower() + ' '
    for ac in _ACABADOS_CANON:
        if ' ' + ac + ' ' in t or t.strip().endswith(' ' + ac):
            return ac
    return ''


def _aplicar_correcciones_versiones(df, corr_ver):
    """Aplica renombres/fusiones/exclusiones hechas en 'Arreglar versiones'.
    La clave de versión es cil|combustible|caja|acabado (igual que en el
    depurador). Renombrar/fusionar => escribe una version canonica comun en
    'version' (y recalcula firma). Excluir => marca el aviso para sacarlo."""
    ma = df['marca'].map(lambda s: str(s).lower())
    mo = df['modelo'].map(lambda s: str(s).lower())
    cil = df['motor'].fillna('').astype(str)
    comb = df['combustible'].fillna('').astype(str)
    caja = df['caja'].fillna('').astype(str)
    acab = df['version'].map(_acabado_txt)
    clave = cil + '|' + comb + '|' + caja + '|' + acab
    modelo_key = ma + '|' + mo

    excluir_mask = pd.Series(False, index=df.index)
    if '_ver_canon' not in df.columns:
        df['_ver_canon'] = ''
    for mk, cambios in corr_ver.items():
        sel_modelo = (modelo_key == mk)
        if not sel_modelo.any():
            continue
        renombres = cambios.get('renombres', {}) or {}
        for cl, nombre in renombres.items():
            sel = sel_modelo & (clave == cl)
            if sel.any():
                # firma canónica común => se agrupan juntos en el precio justo.
                # No tocamos 'version' para preservar la clave del depurador.
                df.loc[sel, 'firma_ver'] = 'canon:' + str(nombre).strip().lower()
                df.loc[sel, '_ver_canon'] = nombre
        for cl in (cambios.get('excluidas', []) or []):
            excluir_mask |= (sel_modelo & (clave == cl))
    if excluir_mask.any():
        n = int(excluir_mask.sum())
        df = df[~excluir_mask].copy()
        log(f"Arreglar versiones: {n} avisos excluidos por el usuario.")
    return df


def _norm_caja(transmision, version=''):
    """Devuelve 'AT', 'MT' o '' (desconocida) desde el campo transmision;
    si falta, lo infiere del texto de la versión."""
    t = (str(transmision) + ' ' + str(version)).lower()
    if any(x in t for x in ('auto', 'cvt', 'dct', 'tiptronic', 's-tronic',
                            'stronic', 'steptronic', 'eat6', 'eat8', 'at6',
                            'at8', 'edct', 'multitronic')):
        return 'AT'
    if any(x in t for x in ('manual', ' mec', 'mecan', 'bvm', ' mt ',
                            'mt5', 'mt6')):
        return 'MT'
    # 'at' / 'mt' como token suelto al final del texto
    import re as _r
    if _r.search(r'\bat\b', t):
        return 'AT'
    if _r.search(r'\bmt\b', t):
        return 'MT'
    return ''


def firma_version(version_str):
    """
    Firma de MOTOR desde el texto libre del vendedor: cilindrada + familia
    de motor + tracción. Inmune al orden de palabras, guiones y truncados.
      '1.6 gt hybrid4 4x4 300 e'   -> '1.6|hybrid4|4x4'
      '1.6 gt hybrid4 300 4x4 e'   -> '1.6|hybrid4|4x4'   (mismo grupo)
      '1.5 allure pack bluehdi'    -> '1.5|bluehdi|'
      'blue hdi 130 eat8 1.5 au'   -> '1.5|bluehdi|'      (mismo grupo)
    Devuelve None si no hay cilindrada + familia (grupo poco confiable).
    """
    if pd.isna(version_str) or not str(version_str).strip():
        return None
    s = str(version_str).lower()
    cil = extraer_motor(s)
    aplastado = re.sub(r'[\s\-\.]', '', s)
    familia = None
    for canon, variantes in FAMILIAS_MOTOR:
        if any(vv in aplastado for vv in variantes):
            familia = canon
            break
    if not cil or not familia:
        return None
    if re.search(r'4\s*x\s*4|awd|4wd', s):
        trac = '4x4'
    elif re.search(r'4\s*x\s*2|fwd|2wd', s):
        trac = '4x2'
    else:
        trac = ''
    # Línea especial: Raptor es casi otro vehículo (motor, suspensión, ancho,
    # precio 2-2.5x) -> nunca se mezcla con las versiones normales. GT/GT-Line
    # es la línea de equipamiento superior. Raptor manda sobre GT si ambas.
    if 'raptor' in s:
        linea = 'raptor'
    elif re.search(r'\bamg\b', s):
        linea = 'amg'
    elif re.search(r'\brubicon\b', s):
        linea = 'rubicon'
    elif re.search(r'\bsti\b', s):
        linea = 'sti'
    elif re.search(r'\bgti\b', s):
        linea = 'gti'
    elif re.search(r'\bgt\b|gt\s*-?\s*line|gtline', s):
        linea = 'gt'
    else:
        linea = ''
    return f"{cil}|{familia}|{trac}|{linea}"


# ---------------------------------------------------------------------------
# 1. CARGA Y LIMPIEZA
# ---------------------------------------------------------------------------
def cargar_captura(path):
    """Lee una captura .xlsx tal cual (la limpieza la hace limpiar())."""
    return pd.read_excel(path)


def fecha_de_archivo(path):
    """Fecha AAAA-MM-DD desde el nombre del archivo (base_2026-06-27.xlsx)."""
    mm = _re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(str(path)))
    if mm:
        try:
            return pd.to_datetime(mm.group(1))
        except Exception:
            pass
    return None


def fecha_de_captura(df):
    """Devuelve la fecha dominante de la captura como datetime (día)."""
    if 'fecha' not in df.columns:
        return pd.Timestamp.now().normalize()
    col = df['fecha'].astype(str)
    # Formato ISO del estado incremental: "2026-07-08" (año-mes-día). Se prueba
    # PRIMERO para no confundir 2026-07-08 con el 8 de agosto (dayfirst lo daba
    # vuelta). Luego el formato con hora del scraper, y por último algo flexible.
    s = pd.to_datetime(col, format='%Y-%m-%d', errors='coerce')
    if s.isna().all():
        s = pd.to_datetime(col, format='%d-%m-%Y %H:%M', errors='coerce')
    if s.isna().all():
        s = pd.to_datetime(col, format='%d-%m-%Y', errors='coerce')
    if s.isna().all():
        # último recurso: ISO con hora, sin dayfirst (evita invertir mes/día)
        s = pd.to_datetime(col, errors='coerce')
    if s.isna().all():
        return pd.Timestamp.now().normalize()
    return s.dt.normalize().mode().iloc[0]


def cargar_desde_incremental(estado_csv="estado_avisos.csv",
                             specs_csv="especificaciones.csv"):
    """
    Arma la tabla de trabajo desde los archivos del flujo incremental:
      - estado_avisos.csv: avisos activos con catálogo + precio + teléfono
      - especificaciones.csv: versión oficial limpia + specs (si ya se bajaron)
    Devuelve un DataFrame con las columnas que espera el motor, más
    'version_oficial' cuando hay specs. Solo incluye avisos activos (activo=1).
    """
    df = pd.read_csv(estado_csv)
    # solo activos
    if 'activo' in df.columns:
        df = df[df['activo'].astype(str) != '0'].copy()

    # unir specs por id (si existe el archivo)
    if os.path.exists(specs_csv):
        specs = pd.read_csv(specs_csv)
        # quedarse con la última fila por id (por si un id se bajó dos veces)
        specs = specs.drop_duplicates(subset='id', keep='last')
        cols_specs = [c for c in specs.columns if c not in ('url',)]
        df = df.merge(specs[cols_specs], on='id', how='left', suffixes=('', '_spec'))

    # mapear a lo que espera el motor
    df['fecha'] = df.get('ultima_vez', '')
    df['Enlace Web'] = df.get('url', '')
    df['nombre_whatsapp'] = ''
    for col in ['telefono', 'whatsapp', 'moneda', 'traccion', 'condicion',
                'transmision', 'combustible', 'entidad']:
        if col not in df.columns:
            df[col] = ''
    if 'version_oficial' not in df.columns:
        df['version_oficial'] = ''

    return df



    df = pd.read_excel(path)
    faltan = [c for c in COLUMNAS_ESPERADAS if c not in df.columns]
    if faltan:
        raise ValueError(f"{os.path.basename(path)}: faltan columnas {faltan}")

    urls = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        hdr = [c.value for c in ws[1]]
        ci_link = hdr.index('Enlace Web') + 1
        ci_id = hdr.index('id') + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci_link)
            idv = ws.cell(row=r, column=ci_id).value
            if cell.hyperlink and cell.hyperlink.target:
                urls[idv] = cell.hyperlink.target
        wb.close()
    except Exception as e:
        log(f"  aviso: no se pudieron leer hipervinculos ({e})")

    df['url'] = df['id'].map(urls)
    return df


def limpiar(df):
    """Aplica las reglas de limpieza validadas. Devuelve df limpio."""
    df = df.drop_duplicates(subset='id', keep='first').copy()
    if 'url' not in df.columns:
        df['url'] = ''
    df = df[df['moneda'] == 'CLP']
    df['año'] = pd.to_numeric(df['año'], errors='coerce')
    df = df[(df['año'] >= ANIO_MIN) & (df['año'] <= ANIO_MAX)]
    df['precio'] = pd.to_numeric(df['precio'], errors='coerce')
    df['kilometraje'] = pd.to_numeric(df['kilometraje'], errors='coerce')
    # Km abreviado (regla del usuario): muchos vendedores escriben "111"
    # por 111.000 km. Si el auto NO es del año (2+ de antiguedad) y el km
    # publicado es de 2-3 cifras, se interpreta en miles cuando el
    # resultado es plausible (<= 60.000 km/ano).
    from datetime import datetime as _dtk
    _antk = (_dtk.now().year - df['año']).clip(lower=0)
    _cand = (df['kilometraje'] >= 10) & (df['kilometraje'] <= 999) & (_antk >= 2)
    _plaus = (df['kilometraje'] * 1000 / _antk.clip(lower=1)) <= 60000
    _fixk = _cand & _plaus
    if int(_fixk.sum()):
        df.loc[_fixk, 'kilometraje'] = df.loc[_fixk, 'kilometraje'] * 1000
        log(f"  Km abreviado interpretado en miles: {int(_fixk.sum())} avisos "
            f"(ej: 111 -> 111.000)")
    df = df[(df['precio'] >= PRECIO_MIN) & (df['precio'] <= PRECIO_MAX)]
    df = df[(df['kilometraje'] >= KM_MIN) & (df['kilometraje'] <= KM_MAX)]
    df['marca'] = df['marca'].astype(str).str.strip().str.lower()
    df['modelo'] = df['modelo'].fillna('').astype(str).str.strip().str.lower()
    anio_actual = ANIO_MAX
    df['antiguedad'] = (anio_actual - df['año']).clip(lower=0)
    df['km_anual'] = np.where(df['antiguedad'] > 0,
                              df['kilometraje'] / df['antiguedad'],
                              df['kilometraje'])
    # Extraer cilindrada del campo version (mejora el agrupamiento de precio justo)
    df['motor'] = df['version'].apply(extraer_motor)
    # Normalizar transmisión a AT/MT. La caja cambia el precio ~1M o más,
    # así que se usa para no comparar un automático con un manual. Se toma
    # del campo transmision del estado y, si falta, se infiere del texto.
    df['caja'] = df.apply(lambda r: _norm_caja(r.get('transmision'), r.get('version')), axis=1)
    df['firma_ver'] = df['version'].apply(firma_version)

    # --- Correcciones de versiones por modelo (herramienta "Arreglar versiones")
    # Reconstruye la clave cil|comb|caja|acabado de cada aviso y, si el usuario
    # la renombró/fusionó/excluyó para ese modelo, aplica el cambio.
    if os.path.exists('correcciones_versiones.json'):
        try:
            with open('correcciones_versiones.json', encoding='utf-8') as f:
                corr_ver = json.load(f)
        except Exception as e:
            corr_ver = {}
            log(f"  (correcciones_versiones.json ilegible: {e})")
        if corr_ver:
            df = _aplicar_correcciones_versiones(df, corr_ver)
    tabla_gen = cargar_generaciones()
    df = asignar_generacion(df, tabla_gen)

    # --- AÑOS DE TRANSICION: si un año cae en DOS generaciones del mismo
    # modelo (ej: la gen II llega a 2020 y la III arranca en 2020), la
    # asignacion por año es ambigua -> se marca para revision por FOTOS.
    trans_map = {}
    for (ma, mo), rangos in tabla_gen.items():
        if len(rangos) < 2:
            continue
        # a) años que caen dentro de DOS rangos (solape declarado)
        for y in range(min(d for d, _, _ in rangos),
                       max(min(h, 2030) for _, h, _ in rangos) + 1):
            gens_y = sorted({g for d, h, g in rangos if d <= y <= h})
            if len(gens_y) >= 2:
                trans_map[(ma, mo, y)] = '/'.join(gens_y)
        # b) años FRONTERA entre generaciones consecutivas: el año de cambio
        # (y el siguiente) conviven ambas generaciones en el mercado real,
        # aunque la tabla no se solape (gen II hasta 2019 / gen III desde 2020)
        orden = sorted(rangos)
        for (d1, h1, g1), (d2, h2, g2) in zip(orden, orden[1:]):
            if g1 != g2 and 0 <= d2 - h1 <= 1 and h1 < 9999:
                par = '/'.join(sorted({g1, g2}))
                for y in (h1, d2):
                    trans_map.setdefault((ma, mo, y), par)
    if trans_map:
        _man = df['marca'].map(_norm_txt)
        _mon = df['modelo'].map(_norm_txt)
        df['gens_pos'] = [trans_map.get((a, b, int(y)), '') if pd.notna(y) else ''
                          for a, b, y in zip(_man, _mon, df['año'])]
    else:
        df['gens_pos'] = ''

    # --- CORRECCIONES POR AVISO (correcciones_avisos.json): lo que TU OJO
    # confirmo mirando las fotos manda sobre todo lo automatico.
    #   {id: {"gen": "III"}}      -> fija la generacion de ESE aviso
    #   {id: {"excluir": true}}   -> lo saca del analisis (comparables y medianas)
    corr_av = {}
    if os.path.exists('correcciones_avisos.json'):
        try:
            with open('correcciones_avisos.json', encoding='utf-8') as f:
                corr_av = {str(k).lower(): v for k, v in json.load(f).items()}
        except Exception as e:
            log(f"  (correcciones_avisos.json ilegible: {e})")
    if corr_av:
        ids_low = df['id'].astype(str).str.lower()
        gen_fix = ids_low.map(lambda i: (corr_av.get(i) or {}).get('gen') or None)
        n_gen = int(gen_fix.notna().sum())
        if n_gen:
            df.loc[gen_fix.notna(), 'gen'] = gen_fix[gen_fix.notna()]
            df.loc[gen_fix.notna(), 'gens_pos'] = ''   # ya no es ambiguo
        ver_fix = ids_low.map(lambda i: (corr_av.get(i) or {}).get('version') or None)
        n_ver = int(ver_fix.notna().sum())
        if n_ver:
            df.loc[ver_fix.notna(), 'version'] = ver_fix[ver_fix.notna()]
            # la firma se recalcula sobre la versión corregida
            df.loc[ver_fix.notna(), 'firma_ver'] = \
                df.loc[ver_fix.notna(), 'version'].apply(firma_version)
        excl_av = ids_low.map(lambda i: bool((corr_av.get(i) or {}).get('excluir')))
        n_exc = int(excl_av.sum())
        if n_exc:
            df = df[~excl_av].copy()
        log(f"Correcciones por aviso (fotos): {n_gen} generaciones fijadas, "
            f"{n_ver} versiones corregidas, {n_exc} avisos descartados.")

    # correcciones del depurador: la versión canónica confirmada reemplaza a la
    # firma automática como llave de agrupación; 'excluir' anula firmas basura
    corr = cargar_correcciones()
    if corr:
        claves = (df['marca'].map(_norm_txt) + '|' + df['modelo'].map(_norm_txt)
                  + '|' + df['gen'].astype(str) + '|' + df['firma_ver'].fillna(''))
        canon = claves.map(lambda k: (corr.get(k) or {}).get('canonica') or None)
        excl = claves.map(lambda k: bool((corr.get(k) or {}).get('excluir')))
        df.loc[canon.notna(), 'firma_ver'] = canon[canon.notna()]
        df.loc[excl, 'firma_ver'] = None
        log(f"  Depurador: {int(canon.notna().sum())} avisos con versión canónica, "
            f"{int(excl.sum())} excluidos de agrupación fina.")
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2. PRECIO JUSTO — cascada de 3 pasadas por especificidad de grupo
# ---------------------------------------------------------------------------
# Pasada 1 (más específica): marca + modelo + motor + combustible, ±3 años
#   → separa, p.ej., Peugeot 3008 1.2 bencina vs 1.5 diesel vs 2.0 diesel
# Pasada 2 (media):          marca + modelo + combustible, ±3 años
#   → para los sin motor extraído o con pocos comparables en pasada 1
# Pasada 3 (fallback):       marca + modelo, ±3 años
#   → comportamiento original, para los que aún no tienen precio
# ---------------------------------------------------------------------------
def calcular_precio_justo(df):
    fair   = np.full(len(df), np.nan)
    ncomp  = np.zeros(len(df), dtype=int)
    metodo = np.array([''] * len(df), dtype=object)
    dispersion = np.zeros(len(df))

    DEP_ANIO = 0.065
    DEP_KM_1000 = 0.0017
    PREF_KM = 1.6
    ANIO_ACTUAL = datetime.now().year
    km_norm = {}   # idx -> km típico usado cuando el km publicado es dudoso

    # Tracción del aviso: columna -> texto de versión -> comentario.
    # Regla del usuario: un 4x2 NO se compara con un 4x4 (se excluye la
    # tracción CONTRARIA conocida; los sin dato no se pueden condenar).
    _cargar_comentarios_txt()
    _cargar_ubicaciones()
    _ids_l = df['id'].astype(str).str.lower().str.strip().values
    _tr_col = df['traccion'].values if 'traccion' in df.columns else [''] * len(df)
    _ver_v = df['version'].values
    df['_trc'] = [(_trac_norm(_tr_col[i]) or _trac_norm(_ver_v[i])
                   or _trac_norm(_COMENT_TXT.get(_ids_l[i], '')))
                  for i in range(len(df))]
    _trac_arr = df['_trc'].values
    # caja y cilindrada efectivas: columna -> comentario (misma regla)
    _cj_col = df['caja'].fillna('').values if 'caja' in df.columns else [''] * len(df)
    _mot_col = df['motor'].values if 'motor' in df.columns else [None] * len(df)
    df['_cj_ef'] = [(_caja_de_txt(_ver_v[i]) or str(_cj_col[i])
                     or _caja_de_txt(_COMENT_TXT.get(_ids_l[i], '')))
                    for i in range(len(df))]
    df['_mot_ef'] = [((str(_mot_col[i]) if pd.notna(_mot_col[i]) else '')
                      or _mot_de_txt(_COMENT_TXT.get(_ids_l[i], '')))
                     for i in range(len(df))]
    _caja_arr = df['_cj_ef'].values
    _mot_arr = df['_mot_ef'].values
    df['_cab_ef'] = [(_cab_de_txt(_ver_v[i]) or _cab_de_txt(_COMENT_TXT.get(_ids_l[i], '')))
                     for i in range(len(df))]
    df['_body_ef'] = [(_body_de_txt(_ver_v[i]) or _body_de_txt(_COMENT_TXT.get(_ids_l[i], '')))
                      for i in range(len(df))]
    _cab_arr = df['_cab_ef'].values
    _body_arr = df['_body_ef'].values

    # Máscara de avisos que NO sirven como comparables (dañados, problema
    # legal, modificados): se tasan igual pero no ensucian el pool de otros.
    if '_excluir_pool' in df.columns:
        _raro_arr = df['_excluir_pool'].fillna(False).values.astype(bool)
    else:
        _raro_arr = np.zeros(len(df), dtype=bool)

    def _calc(yrs_pool, kms_pool, prs_pool, idxs_t, yrs_t, kms_t, sufijo,
              pool_raro=None, trac_pool=None, caja_pool=None, mot_pool=None,
              cab_pool=None, body_pool=None):
        """Precio justo: cada comparable se normaliza al año/km del target y se
        toma la mediana. pool_raro (bool array del pool) excluye del cálculo los
        avisos con situación rara, sin dejar de tasar al target."""
        pool_sano = ~pool_raro if pool_raro is not None else np.ones(len(prs_pool), dtype=bool)
        for y, km, idx in zip(yrs_t, kms_t, idxs_t):
            if not np.isnan(fair[idx]):
                continue
            mask = (np.abs(yrs_pool - y) <= 3) & pool_sano
            if trac_pool is not None:
                _tt = _trac_arr[idx]
                if _tt:
                    mask = mask & ~((trac_pool != '') & (trac_pool != _tt))
            if caja_pool is not None:
                _tc = _caja_arr[idx]
                if _tc:
                    mask = mask & ~((caja_pool != '') & (caja_pool != _tc))
            if mot_pool is not None:
                _tm = _mot_arr[idx]
                if _tm:
                    mask = mask & ~((mot_pool != '') & (mot_pool != _tm))
            if cab_pool is not None:
                _tb = _cab_arr[idx]
                if _tb:
                    mask = mask & ~((cab_pool != '') & (cab_pool != _tb))
            if body_pool is not None:
                _ty = _body_arr[idx]
                if _ty:
                    mask = mask & ~((body_pool != '') & (body_pool != _ty))
            n = int(mask.sum())
            if n < MIN_COMPARABLES:
                continue
            pp = prs_pool[mask].astype(float)
            kk = kms_pool[mask].astype(float)
            yy = yrs_pool[mask].astype(float)
            # km del target dudoso (2+ años <1000km, 5+ años <5000km): dato
            # basura -> normalizar con el km TIPICO de sus comparables
            # (mediana del pool) en vez de inflar el justo con km~0.
            edad_t = max(0, ANIO_ACTUAL - int(y))
            km_ef = km
            if (edad_t >= 2 and km < 1000) or (edad_t >= 5 and km < 5000):
                km_ef = float(np.median(kk))
                km_norm[idx] = km_ef
            ajuste_anio = np.power(1 - DEP_ANIO, -(y - yy))
            ajuste_km = np.power(1 - DEP_KM_1000 * PREF_KM, (km_ef - kk) / 1000.0)
            pp_adj = pp * ajuste_anio * ajuste_km
            pp_adj = np.clip(pp_adj, pp * 0.35, pp * 2.5)
            ncomp[idx] = n
            _q1d, _q3d = np.percentile(pp_adj, [25, 75])
            _medd = np.median(pp_adj)
            dispersion[idx] = float((_q3d - _q1d) / _medd) if _medd > 0 else 0.0
            if n >= 10:
                lo, hi = np.percentile(pp_adj, [5, 95])
            else:
                lo, hi = pp_adj.min(), pp_adj.max()
            sel = (pp_adj >= lo) & (pp_adj <= hi)
            fair[idx] = float(np.median(pp_adj[sel])) if sel.sum() else float(np.median(pp_adj))
            metodo[idx] = f'ajuste_{sufijo}'

    # --- Pasada 0: (marca, modelo, version_oficial) — la MÁS precisa ---
    # Solo corre si hay specs bajadas (columna 'version_oficial' con datos).
    # Es la mejor agrupación: misma versión canónica de fábrica = mismo auto.
    if 'version_oficial' in df.columns:
        df_v = df[df['version_oficial'].notna() &
                  (df['version_oficial'].astype(str).str.strip() != '')]
        if not df_v.empty:
            for (ma, mo, vof), g in df_v.groupby(
                    ['marca', 'modelo', 'version_oficial']):
                yrs  = g['año'].values
                kms  = g['kilometraje'].values
                prs  = g['precio'].values
                idxs = g.index.values
                _calc(yrs, kms, prs, idxs, yrs, kms, 'vof', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values, pool_raro=g['_excluir_pool'].values if '_excluir_pool' in g.columns else None)

    # --- Pasada 0.5: (marca, modelo, FIRMA de motor) ---
    # Firma = cilindrada + familia + tracción desde el texto sucio del
    # vendedor. Separa el 1.6 thp del 1.6 puretech del 1.6 hybrid4 aunque
    # las palabras vengan desordenadas o con guiones.
    if 'firma_ver' in df.columns:
        df_f = df[df['firma_ver'].notna()]
        if not df_f.empty:
            for (ma, mo, fv, gn), g in df_f.groupby(['marca', 'modelo', 'firma_ver', 'gen']):
                idxs_pend = g.index[np.isnan(fair[g.index])].values
                if len(idxs_pend) == 0:
                    continue
                _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                      idxs_pend,
                      df.loc[idxs_pend, 'año'].values,
                      df.loc[idxs_pend, 'kilometraje'].values, 'firma', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values, pool_raro=g['_excluir_pool'].values if '_excluir_pool' in g.columns else None)

    # --- Pasada 1: (marca, modelo, motor, combustible, caja) ---
    # Incluye la transmisión: un automático no se compara con un manual
    # (diferencia típica ~1M+). Los de caja desconocida ('') forman su
    # propio grupo, que es lo correcto (no se mezclan con ninguna).
    df_m = df[df['motor'].notna()]
    if not df_m.empty:
        for (ma, mo, mot, comb, cj, gn), g in df_m.groupby(
                ['marca', 'modelo', 'motor', 'combustible', 'caja', 'gen']):
            idxs_pend = g.index[np.isnan(fair[g.index])].values
            if len(idxs_pend) == 0:
                continue
            _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                  idxs_pend,
                  df.loc[idxs_pend, 'año'].values,
                  df.loc[idxs_pend, 'kilometraje'].values, 'motor', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values, pool_raro=g['_excluir_pool'].values if '_excluir_pool' in g.columns else None)

    # --- Pasada 2a: (marca, modelo, combustible, CILINDRADA) ---
    # Cuando el aviso tiene cilindrada (motor), no se compara un 1.2 con un
    # 1.6: se agrupa también por cilindrada. Evita mezclar versiones de
    # distinto valor dentro del mismo combustible/generación.
    df_c = df[df['motor'].notna()]
    if not df_c.empty:
        for (ma, mo, comb, mot, cj, gn), g in df_c.groupby(
                ['marca', 'modelo', 'combustible', 'motor', 'caja', 'gen']):
            idxs_pend = g.index[np.isnan(fair[g.index])].values
            if len(idxs_pend) == 0:
                continue
            _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                  idxs_pend,
                  df.loc[idxs_pend, 'año'].values,
                  df.loc[idxs_pend, 'kilometraje'].values, 'comb_cil', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values, pool_raro=g['_excluir_pool'].values if '_excluir_pool' in g.columns else None)

    # --- Pasada 2b: (marca, modelo, combustible, caja) ---
    # Para los sin cilindrada pero CON caja conocida: separa AT de MT.
    df_cj = df[df['caja'] != '']
    if not df_cj.empty:
        for (ma, mo, comb, cj, gn), g in df_cj.groupby(
                ['marca', 'modelo', 'combustible', 'caja', 'gen']):
            idxs_pend = g.index[np.isnan(fair[g.index])].values
            if len(idxs_pend) == 0:
                continue
            _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                  idxs_pend,
                  df.loc[idxs_pend, 'año'].values,
                  df.loc[idxs_pend, 'kilometraje'].values, 'comb', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values, pool_raro=g['_excluir_pool'].values if '_excluir_pool' in g.columns else None)

    # --- Pasada 2c: (marca, modelo, combustible) ---
    # Último recurso por combustible, para los que no tienen ni cilindrada
    # ni caja conocidas. Mejor una tasación aproximada que ninguna.
    for (ma, mo, comb, gn), g in df.groupby(['marca', 'modelo', 'combustible', 'gen']):
        idxs_pend = g.index[np.isnan(fair[g.index])].values
        if len(idxs_pend) == 0:
            continue
        _calc(
            g['año'].values, g['kilometraje'].values, g['precio'].values,
            idxs_pend,
            df.loc[idxs_pend, 'año'].values,
            df.loc[idxs_pend, 'kilometraje'].values,
            'comb', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values
        )

    # --- Pasada 3: fallback (marca, modelo) — comportamiento original ---
    for (ma, mo), g in df.groupby(['marca', 'modelo']):
        idxs_pend = g.index[np.isnan(fair[g.index])].values
        if len(idxs_pend) == 0:
            continue
        _calc(
            g['año'].values, g['kilometraje'].values, g['precio'].values,
            idxs_pend,
            df.loc[idxs_pend, 'año'].values,
            df.loc[idxs_pend, 'kilometraje'].values,
            'modelo', trac_pool=g['_trc'].values, caja_pool=g['_cj_ef'].values, mot_pool=g['_mot_ef'].values, cab_pool=g['_cab_ef'].values, body_pool=g['_body_ef'].values,
            pool_raro=g['_excluir_pool'].values if '_excluir_pool' in g.columns else None
        )

    df['dispersion'] = dispersion
    df['tasacion_confiable'] = (dispersion <= 0.45) & ~np.isnan(fair)
    df['precio_justo'] = fair
    df['comparables']  = ncomp
    df['km_norm'] = [km_norm.get(i, np.nan) for i in range(len(df))]
    df['metodo']       = metodo
    valido = df['precio_justo'].notna()
    df['descuento_pct'] = np.where(
        valido, (df['precio_justo'] - df['precio']) / df['precio_justo'] * 100, np.nan)
    # Descuento extremo (>55%): NO se oculta, se MARCA para revisar. Puede ser
    # una ganga real (venta urgente) o un error; el usuario decide. Sigue en
    # el oro, pero con la bandera 'descuento_extremo' para filtrarlo/mirarlo.
    DESCUENTO_MAX_CREIBLE = 55.0
    df['descuento_extremo'] = valido & (df['descuento_pct'] > DESCUENTO_MAX_CREIBLE)
    df['op_precio'] = ((df['descuento_pct'] >= DESCUENTO_MIN) & valido
                       & df['tasacion_confiable'])
    # km imposible (auto con años pero casi sin km): se MARCA para revisar,
    # no se saca. Puede ser error de tipeo o un auto genuinamente poco usado;
    # el usuario decide mirando el aviso. Sigue en el oro con bandera km_dudoso.
    km_imposible = ((df['antiguedad'] >= 2) & (df['kilometraje'] < 1000)) | \
                   ((df['antiguedad'] >= 5) & (df['kilometraje'] < 5000))
    df['km_dudoso_flag'] = km_imposible
    df['op_km']     = (df['km_anual'] < KM_ANUAL_MAX) & (df['antiguedad'] >= 1)
    df['op_oro']    = df['op_precio'] & df['op_km']
    return df


# ---------------------------------------------------------------------------
# 3. ROTACION (comparando capturas consecutivas)
# ---------------------------------------------------------------------------
def calcular_rotacion(capturas):
    if len(capturas) < 2:
        return None

    # Ordenar por fecha y elegir el par CONSECUTIVO más cercano en el tiempo.
    # Comparar los extremos (junio vs hoy) daba ~100% de "salida" porque son
    # fotos muy separadas y de orígenes distintos. Dos fotos cercanas dan una
    # tasa de rotación creíble. Si el par más cercano está a más de 21 días,
    # la señal es débil y se marca como poco confiable.
    caps = sorted(capturas, key=lambda c: c['fecha'])
    mejor = None
    for a, b in zip(caps, caps[1:]):
        d = (b['fecha'] - a['fecha']).days
        if d <= 0:
            continue
        if mejor is None or d < mejor[2]:
            mejor = (a, b, d)
    if mejor is None:
        return None
    prim, ult, dias = mejor
    ventana_confiable = dias <= 21

    ids_prim = prim['ids']
    ids_ult = ult['ids']
    desaparecidos = ids_prim - ids_ult
    nuevos = ids_ult - ids_prim
    sobreviven = ids_prim & ids_ult

    # VALIDACIÓN: si las dos capturas casi no comparten IDs, NO son
    # comparables (vienen de exportaciones distintas o formatos distintos).
    # Una rotación real tiene mucho solapamiento (los autos que siguen). Si el
    # solapamiento es <20% de la captura más chica, la medición no es válida:
    # se informa como "datos insuficientes" en vez de un 100% falso.
    minlen = max(1, min(len(ids_prim), len(ids_ult)))
    solapamiento = len(sobreviven) / minlen
    tasa_salida = len(desaparecidos) / max(1, len(ids_prim))
    # Inválida si: poco solapamiento (<40%) O una tasa de salida absurda
    # (>85% "salió" en pocos días = las capturas no son comparables, no ventas).
    if solapamiento < 0.40 or tasa_salida > 0.85:
        return {
            'valida': False,
            'motivo': ('Las capturas disponibles no comparten suficientes avisos '
                       'para medir rotación (probablemente son de orígenes o '
                       'formatos distintos). La rotación real se activa cuando el '
                       'sistema acumule varios días seguidos de la misma base.'),
            'global': {
                'fecha_desde': prim['fecha'].strftime('%Y-%m-%d'),
                'fecha_hasta': ult['fecha'].strftime('%Y-%m-%d'),
                'solapamiento_pct': round(solapamiento * 100, 1),
                'n_capturas': len(caps),
            },
            'por_modelo': [],
        }

    dprim = prim['df'].copy()
    dprim['id'] = dprim['id'].astype(str).str.lower().str.strip()
    dprim = dprim.set_index('id')

    glob = {
        'fecha_desde': prim['fecha'].strftime('%Y-%m-%d'),
        'fecha_hasta': ult['fecha'].strftime('%Y-%m-%d'),
        'dias': int(dias),
        'avisos_inicio': len(ids_prim),
        'avisos_fin': len(ids_ult),
        'desaparecidos': len(desaparecidos),
        'nuevos': len(nuevos),
        'tasa_salida_periodo': round(len(desaparecidos) / len(ids_prim) * 100, 1),
        'tasa_salida_diaria': round(len(desaparecidos) / len(ids_prim) / dias * 100, 2),
        'ventana_confiable': bool(ventana_confiable),
        'n_capturas': len(caps),
    }

    df_des = dprim.loc[[i for i in desaparecidos if i in dprim.index]]
    base_modelo = dprim.groupby(['marca', 'modelo']).size()
    fueron_modelo = df_des.groupby(['marca', 'modelo']).size()
    rot = pd.DataFrame({'inicio': base_modelo, 'salieron': fueron_modelo}).fillna(0)
    rot['salieron'] = rot['salieron'].astype(int)
    rot = rot[rot['inicio'] >= 1]
    rot['tasa_salida'] = (rot['salieron'] / rot['inicio'] * 100).round(1)
    rot['inventario'] = np.where(rot['inicio'] >= 10, 'alto', 'bajo')
    rot['dias_venta_est'] = np.where(
        rot['salieron'] > 0,
        (dias / (rot['salieron'] / rot['inicio'])).round(0),
        np.nan)

    def cap(s):
        return ' '.join(w.capitalize() for w in str(s).split())

    por_modelo = []
    for (ma, mo), r in rot.iterrows():
        por_modelo.append({
            'marca': cap(ma), 'modelo': cap(mo),
            'stock': int(r['inicio']),
            'salieron': int(r['salieron']),
            'tasa_salida': float(r['tasa_salida']),
            'dias_venta_est': None if np.isnan(r['dias_venta_est']) else int(r['dias_venta_est']),
            'inventario': r['inventario'],
            'confiable': bool(r['inventario'] == 'alto'),
        })
    por_modelo.sort(key=lambda x: x['tasa_salida'], reverse=True)

    dult = ult['df'].copy()
    dult['id'] = dult['id'].astype(str).str.lower().str.strip()
    dult = dult.set_index('id')
    # Solo IDs presentes en ambos índices y sin duplicados (evita KeyError).
    idx_prim = set(dprim.index[~dprim.index.duplicated()])
    idx_ult = set(dult.index[~dult.index.duplicated()])
    comun = [i for i in sobreviven if i in idx_prim and i in idx_ult]
    dprim = dprim[~dprim.index.duplicated(keep='first')]
    dult = dult[~dult.index.duplicated(keep='first')]
    p0 = dprim.loc[comun, 'precio']
    p1 = dult.loc[comun, 'precio']
    cambio = (p0 != p1)
    bajaron = (p1 < p0)
    glob['con_cambio_precio'] = int(cambio.sum())
    glob['bajaron_precio'] = int((cambio & bajaron).sum())
    glob['subieron_precio'] = int((cambio & ~bajaron).sum())

    # Detalle de BAJAS de precio (con marca/modelo) para el widget del BI
    bajas = []
    ids_baja = [i for i in comun if p1[i] < p0[i]]
    for i in ids_baja:
        r = dult.loc[i]
        antes, ahora = int(p0[i]), int(p1[i])
        bajas.append({
            'id': i,
            'marca': str(r.get('marca', '') or ''),
            'modelo': str(r.get('modelo', '') or ''),
            'año': int(r['año']) if pd.notna(r.get('año')) else None,
            'antes': antes, 'ahora': ahora,
            'delta': ahora - antes,
            'pct': round((ahora - antes) * 100.0 / antes, 1) if antes else None,
            'url': str(r.get('url') or '') if pd.notna(r.get('url')) else '',
        })
    bajas.sort(key=lambda x: x['delta'])
    glob_bajas = bajas[:120]

    return {'valida': True, 'global': glob, 'por_modelo': por_modelo,
            'bajas_precio': glob_bajas}


# ---------------------------------------------------------------------------
# ANÁLISIS DE MERCADO
# ---------------------------------------------------------------------------
def construir_depurador_modelos(dfa):
    """Para la herramienta 'Arreglar BBDD por modelo'. Por cada modelo con
    volumen, propone sus VERSIONES REALES agrupando por:
       cilindrada | combustible | caja | línea(GT) | acabado
    El acabado (active/allure/gt/etc.) se detecta del texto. Cada grupo trae
    sus avisos (año, km, precio, url), precio por año y una etiqueta propuesta
    legible. El usuario confirma/corrige/fusiona en el dashboard.
    Devuelve {clave_modelo: {marca, modelo, versiones:[...], sin_clasificar:n}}.
    """
    ACABADOS = ['raptor', 'xlt', 'elite', 'luxury', 'advance', 'pro', 'dx', 'deluxe', 'katana', 'dynamic', 'xe', 'poer', 'platinum', 'lariat', 'dlx', 'country', 'nomade', 'sel', 'xs', 'trend', 'gs', 'premium', 'sportback', 'elegance', 'sr', 'executive', 'prestige', 'ambiente', 'sahara', 'activ', 'denali', 'ghia', 'gt line', 'gt-line', 'gtline', 'gt', 'allure', 'active',
                'premier', 'premiere', 'feel', 'style', 'access', 'business',
                'signature', 'sport', 'limited', 'ltz', 'lt', 'ls', 'exclusive',
                'dynamique', 'zen', 'intens', 'life', 'trend', 'comfort',
                'titanium', 'ambiente', 'highline', 'comfortline', 'trendline',
                'xei', 'xli', 'gli', 'gls', 'glx', 'gl', 'ex', 'exl', 'lx',
                'touring', 'value', 'full', 'sense']

    def acabado_de(txt):
        t = ' ' + str(txt).lower() + ' '
        for ac in ACABADOS:                     # orden: gt line antes que gt
            if ' ' + ac + ' ' in t or t.strip().endswith(' ' + ac):
                return ac
        return ''

    def etiqueta(cil, comb, caja, acab, trac='', cab='', gen=''):
        partes = []
        if gen:
            gtxt = str(gen)
            partes.append(gtxt if 'gen' in gtxt.lower() else 'Gen ' + gtxt)
        if cil:
            partes.append(str(cil))
        if acab:
            partes.append(acab.upper() if len(acab) <= 3 else acab.title())
        if trac:
            partes.append(trac)
        cabtxt = {'reg': 'Cab.Simple', 'dcab': 'Doble Cab.'}.get(cab, '')
        if cabtxt:
            partes.append(cabtxt)
        cb = {'diesel': 'Diésel', 'bencina': 'Bencina', 'híbrido': 'Híbrido',
              'hibrido': 'Híbrido'}.get(str(comb).lower(), str(comb).title())
        if cb:
            partes.append(cb)
        if caja:
            partes.append(caja)
        return ' '.join(partes) if partes else '(sin datos)'

    out = {}
    vol = dfa.groupby(['marca', 'modelo'])['precio'].transform('size')
    grande = dfa[vol >= 12].copy()
    if grande.empty:
        return out
    def _trac_dep(txt):
        t = str(txt).lower()
        if any(x in t for x in ('4x4', '4wd', 'awd')): return '4x4'
        if any(x in t for x in ('4x2', '4wd', 'rwd', 'fwd', '2wd')): return '4x2'
        return ''
    def _cab_dep(txt):
        t = str(txt).lower()
        if any(x in t for x in ('reg cab', 'cab. sim', 'cabina simple', ' sc ', 'single', ' rc ')): return 'reg'
        if any(x in t for x in ('dob', 'doble', ' cc ', 'crew', 'dcab', ' cd ')): return 'dcab'
        return ''
    grande['_acab'] = grande['version'].apply(acabado_de)
    grande['_trac'] = grande['version'].apply(_trac_dep)
    grande['_cab'] = grande['version'].apply(_cab_dep)
    tiene_gt = grande['firma_ver'].fillna('').str.contains('gt')

    def cap(s):
        return ' '.join(w.capitalize() for w in str(s).split())

    for (ma, mo), g in grande.groupby(['marca', 'modelo']):
        versiones = []
        sin_clasif = 0
        # clave de versión: cilindrada|combustible|caja|acabado (+gt de firma)
        g = g.copy()
        g['_cil'] = g['motor'].fillna('')
        g['_cj'] = g['caja'].fillna('')
        g['_gen'] = g['gen'].fillna('').astype(str) if 'gen' in g.columns else ''
        g['_key'] = (g['_gen'].astype(str) + '|' + g['_cil'].astype(str) + '|' + g['combustible'].astype(str)
                     + '|' + g['_cj'].astype(str) + '|' + g['_acab'].astype(str)
                     + '|' + g['_trac'].astype(str) + '|' + g['_cab'].astype(str))
        # si el usuario ya fusionó/renombró, agrupar por el nombre canónico
        if '_ver_canon' in g.columns:
            g['_key'] = g.apply(lambda r: ('canon::' + str(r['_ver_canon']))
                                if str(r.get('_ver_canon') or '') else r['_key'], axis=1)
        # Fusión de duplicados sin tracción: un grupo con trac vacía que
        # coincide con otro en todo lo demás es la MISMA versión (el vendedor
        # no escribió la tracción). Se absorbe en el que sí la declara.
        _keys_presentes = set(g['_key'].unique())
        _remap = {}
        for _k in list(_keys_presentes):
            if not isinstance(_k, str):
                continue  # claves NaN (algún campo vacío): no se remapean
            partes = _k.split('|')
            if len(partes) >= 7 and partes[5] == '':  # trac vacía (posición 5: gen|cil|comb|caja|acab|trac|cab)
                for _trac_try in ('4x2', '4x4'):
                    _k2 = '|'.join(partes[:5] + [_trac_try] + partes[6:])
                    if _k2 in _keys_presentes:
                        _remap[_k] = _k2
                        break
        if _remap:
            g['_key'] = g['_key'].map(lambda k: _remap.get(k, k))
        for key, gv in g.groupby('_key'):
            if str(key).startswith('canon::'):
                nombre_canon = key.split('canon::', 1)[1]
                cil = comb = caja = acab = ''
            else:
                nombre_canon = ''
                partes_k = (key.split('|') + ['']*7)
                gen_k, cil, comb, caja, acab, trac, cab = partes_k[:7]
            if not cil and not acab:
                sin_clasif += len(gv)
                # igual lo mostramos como grupo "sin identificar"
            avisos = []
            for _, r in gv.sort_values('año', ascending=False).head(60).iterrows():
                avisos.append({
                    'año': int(r['año']) if pd.notna(r['año']) else None,
                    'km': int(r['kilometraje']) if pd.notna(r['kilometraje']) else None,
                    'precio': int(r['precio']) if pd.notna(r['precio']) else None,
                    'ent': 'P' if r['entidad'] == 'Particular' else 'A',
                    'texto': str(r['version'])[:50],
                    'url': (r.get('url') if pd.notna(r.get('url')) else ''),
                    'id': str(r['id']).lower(),
                })
            precio_año = {}
            for a2, gg in gv.groupby('año'):
                if pd.notna(a2):
                    precio_año[int(a2)] = int(gg['precio'].median())
            versiones.append({
                'clave': key,
                'etiqueta': nombre_canon or etiqueta(cil, comb, caja, acab, trac, cab, gen_k),
                'gen': gen_k,
                'cil': cil, 'combustible': comb, 'caja': caja, 'acabado': acab,
                'n': int(len(gv)),
                'precio_med': int(gv['precio'].median()),
                'años': [int(gv['año'].min()), int(gv['año'].max())] if gv['año'].notna().any() else None,
                'precio_año': dict(sorted(precio_año.items())),
                'identificada': bool(cil or acab),
                'avisos': avisos,
            })
        versiones.sort(key=lambda x: -x['n'])
        out[f"{ma}|{mo}"] = {
            'marca': cap(ma), 'modelo': cap(mo),
            'total': int(len(g)),
            'n_versiones': len(versiones),
            'sin_clasificar': int(sin_clasif),
            'versiones': versiones,
        }
    return out


def construir_analisis(dfa):
    """Análisis tipo analista de mercado sobre la base activa. Devuelve dict
    con: resumen por marca, depreciación/retención, señales de alerta
    (modelos muy nuevos sobre-publicados) y oportunidad por segmento."""
    ANIO = ANIO_MAX
    d = dfa[(dfa['precio'] > 0) & dfa['año'].notna()].copy()
    d['año'] = d['año'].astype(int)

    def cap(s):
        return ' '.join(w.capitalize() for w in str(s).split())

    # --- 1. RESUMEN POR MARCA (volumen, precio medio, antigüedad típica) ---
    por_marca = []
    for ma, g in d.groupby('marca'):
        if len(g) < 20:
            continue
        antig = (ANIO - g['año']).clip(lower=0)
        por_marca.append({
            'marca': cap(ma),
            'avisos': int(len(g)),
            'precio_med': int(g['precio'].median()),
            'antiguedad_med': round(float(antig.median()), 1),
            'km_med': int(g['kilometraje'].median()) if g['kilometraje'].notna().any() else 0,
            'pct_particular': round(float((g['entidad'] == 'Particular').mean() * 100), 0),
        })
    por_marca.sort(key=lambda x: -x['avisos'])

    # --- 2. DEPRECIACIÓN / RETENCIÓN DE VALOR POR MARCA ---
    # Cuánto pierde el valor entre 0 y 3 años (mediana de precio por año).
    deprec = []
    for ma, g in d.groupby('marca'):
        if len(g) < 40:
            continue
        p0 = g[g['año'] >= ANIO - 1]['precio'].median()
        p3 = g[g['año'].between(ANIO - 4, ANIO - 3)]['precio'].median()
        if pd.notna(p0) and pd.notna(p3) and p0 > 0:
            perdida = round((1 - p3 / p0) * 100, 0)
            deprec.append({
                'marca': cap(ma), 'avisos': int(len(g)),
                'perdida_3a': perdida,
                'retiene_3a': round(100 - perdida, 0),
                'precio_0a': int(p0), 'precio_3a': int(p3),
            })
    deprec.sort(key=lambda x: x['perdida_3a'], reverse=True)

    # --- 3. SEÑALES DE ALERTA: modelos con demasiados avisos muy nuevos ---
    # Alto % de 0-1 año publicados = mucha gente revende casi nuevo
    # (posible problema del auto, arrepentimiento, o flota/rent-a-car).
    alertas = []
    for (ma, mo), g in d.groupby(['marca', 'modelo']):
        if len(g) < 15:
            continue
        nuevos = int((g['año'] >= ANIO - 1).sum())
        pct = nuevos / len(g) * 100
        if pct >= 45 and nuevos >= 8:
            alertas.append({
                'marca': cap(ma), 'modelo': cap(mo),
                'avisos': int(len(g)), 'nuevos_0_1a': nuevos,
                'pct_nuevos': round(pct, 0),
                'precio_med': int(g['precio'].median()),
            })
    alertas.sort(key=lambda x: (-x['pct_nuevos'], -x['avisos']))

    # --- 4. OPORTUNIDAD POR SEGMENTO ---
    # Segmento = derivado de la carrocería si hay ficha; si no, heurística
    # por modelo. Medimos % de avisos que son oportunidad (oro) por segmento.
    seg = []
    if 'op_oro' in d.columns:
        d['_seg'] = d.apply(_segmento_de, axis=1)
        for s, g in d.groupby('_seg'):
            if s == '?' or len(g) < 30:
                continue
            seg.append({
                'segmento': s, 'avisos': int(len(g)),
                'precio_med': int(g['precio'].median()),
                'pct_oro': round(float(g['op_oro'].mean() * 100), 1),
                'oro': int(g['op_oro'].sum()),
            })
        seg.sort(key=lambda x: -x['pct_oro'])

    return {
        'por_marca': por_marca,
        'depreciacion': deprec,
        'alertas': alertas,
        'segmentos': seg,
        'anio_ref': ANIO,
    }


_SEG_MAP = [
    ('pickup', ['hilux', 'ranger', 'amarok', 'l200', 'dmax', 'd-max', 'frontier',
                'navara', 'np300', 'colorado', 's10', 'gladiator', 'maverick',
                'sail', 'partner', 'kangoo', 'berlingo', 'saveiro']),
    ('suv',    ['tucson', 'sportage', 'santa fe', 'kuga', 'cr-v', 'crv', 'rav4',
                'tiguan', 'q3', 'q5', 'x1', 'x3', 'captur', 'duster', 'tracker',
                'creta', 'kicks', 'seltos', 'cx-5', 'cx5', 'forester', 'outlander',
                'ecosport', 'territory', 'omoda', 'jaecoo', 'tiggo']),
    ('city',   ['sail', 'yaris', 'swift', 'rio', 'accent', 'onix', 'clio', 'sandero',
                'kwid', 'picanto', 'morning', 'spark', 'march', 'versa', 'city',
                'polo', 'gol', 'mobi', 'k3', 'i10', 'i20']),
]


def _segmento_de(r):
    ficha = None
    mo = str(r.get('modelo', '')).lower()
    for seg, modelos in _SEG_MAP:
        if any(x in mo for x in modelos):
            return seg
    return '?'


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    # --- Fuente de datos: incremental (estado_avisos.csv) si existe;
    #     si no, las capturas .xlsx de siempre. ---
    def elegir_capturas_xlsx():
        """Un archivo por FECHA: si conviven 'Chileautos 2026-07-08 AM.xlsx' y
        '... PM.xlsx' (o cualquier otro sufijo), para el análisis por días se
        usa el más reciente de esa fecha (por hora de modificación). Los demás
        quedan en la carpeta como respaldo, sin molestar."""
        grupos = {}
        for path in glob.glob(os.path.join(CARPETA_CAPTURAS, "*.xlsx")):
            mfecha = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
            clave = mfecha.group(1) if mfecha else os.path.basename(path)
            if clave not in grupos or os.path.getmtime(path) > os.path.getmtime(grupos[clave]):
                grupos[clave] = path
        return [grupos[k] for k in sorted(grupos)]

    usa_incremental = os.path.exists("estado_avisos.csv")

    capturas = []
    if usa_incremental:
        log("Usando datos INCREMENTALES (estado_avisos.csv).")
        df = cargar_desde_incremental()
        dfl = limpiar(df)
        fecha = pd.Timestamp.today().normalize()
        try:
            fecha = fecha_de_captura(df)
        except Exception:
            pass
        capturas.append({'fecha': fecha, 'df': dfl, 'ids': set(dfl['id'].astype(str).str.lower().str.strip())})
        con_specs = (dfl.get('version_oficial', pd.Series(dtype=str)).astype(str).str.strip() != '').sum()
        log(f"  {len(df)} avisos activos, {len(dfl)} limpios, {con_specs} con versión oficial (specs).")
        # Si además hay capturas .xlsx viejas, se suman para poder medir rotación.
        if os.path.isdir(CARPETA_CAPTURAS):
            for path in elegir_capturas_xlsx():
                try:
                    dfx = cargar_captura(path)
                    f_cap = fecha_de_archivo(path) or fecha_de_captura(dfx)
                    if f_cap == fecha:
                        # hoy ya está representado por el estado (más completo)
                        continue
                    capturas.append({'fecha': f_cap,
                                     'df': limpiar(dfx), 'ids': set(limpiar(dfx)['id'].astype(str).str.lower().str.strip())})
                except Exception as e:
                    log(f"  (salteo {os.path.basename(path)}: {e})")
    else:
        if not os.path.isdir(CARPETA_CAPTURAS):
            sys.exit(f"No existe la carpeta '{CARPETA_CAPTURAS}' ni estado_avisos.csv. "
                     f"Corré primero la descarga de datos.")
        archivos = elegir_capturas_xlsx()
        if not archivos:
            sys.exit(f"No hay archivos .xlsx en '{CARPETA_CAPTURAS}'.")
        log(f"Capturas encontradas: {len(archivos)}")
        for path in archivos:
            df = cargar_captura(path)
            fecha = fecha_de_archivo(path)
            if fecha is None:
                fecha = fecha_de_captura(df)
            dfl = limpiar(df)
            capturas.append({'fecha': fecha, 'df': dfl, 'ids': set(dfl['id'].astype(str).str.lower().str.strip())})
            log(f"  {os.path.basename(path)} -> {fecha.date()} | "
                f"{len(df)} filas, {len(dfl)} limpias")

    capturas.sort(key=lambda c: c['fecha'])

    reciente = capturas[-1]

    # Cargar etiquetas de comentarios ANTES de tasar, para excluir del pool de
    # comparables los autos con situación rara (dañado, problema legal, etc.):
    # un chocado barato no debe arrastrar el precio justo hacia abajo.
    etiquetas_por_id = {}
    if os.path.exists('comentarios_clasificados.json'):
        try:
            with open('comentarios_clasificados.json', encoding='utf-8') as f:
                clasif = json.load(f)
            for vid, dd in clasif.items():
                etiquetas_por_id[str(vid).lower()] = {
                    'etiquetas': dd.get('etiquetas', []),
                    'comentario': dd.get('comentario', ''),
                    'urgencia': dd.get('urgencia', 0),
                    'resumen_ia': dd.get('resumen_ia', ''),
                }
            # SANEO contra negaciones: el clasificador marca por keyword sin
            # ver el "no" de adelante ("precio NO conversable" -> conversable).
            # Se verifica contra el texto real y se corrige la etiqueta.
            _saneadas = 0
            for vid, dd in etiquetas_por_id.items():
                txt = ' ' + str(dd.get('comentario', '') or
                                _COMENT_TXT.get(vid, '')).lower() + ' '
                ets = dd.get('etiquetas', [])
                if 'precio_conversable' in ets and any(
                        x in txt for x in ('no conversable', 'no es conversable',
                                           'precio firme', 'no negociable',
                                           'precio fijo')):
                    ets.remove('precio_conversable')
                    ets.append('precio_firme')
                    _saneadas += 1
                if 'daniado' in ets and any(
                        x in txt for x in ('sin danos', 'sin daños', 'sin detalles',
                                           'ningun dano', 'ningún daño',
                                           'nada que hacerle')) and not any(
                        x in txt for x in ('chocado', 'volcado', 'para armar',
                                           'fundido', 'golpe')):
                    ets.remove('daniado')
                    _saneadas += 1
            if _saneadas:
                log(f"  Etiquetas saneadas contra negaciones: {_saneadas}")
            log(f"Etiquetas de comentarios cargadas: {len(etiquetas_por_id)}")
        except Exception as e:
            log(f"  aviso: no se pudo leer comentarios_clasificados.json ({e})")
    else:
        log("Sin comentarios_clasificados.json todavia: las oportunidades van sin etiquetas.")

    # Marcar en el df qué avisos NO deben servir de comparables (situación rara).
    EXCLUIR_DE_POOL = {'daniado', 'problema_legal', 'modificado'}
    _rdf = reciente['df']
    _ids_low = _rdf['id'].astype(str).str.lower()
    _rara = _ids_low.map(lambda i: bool(
        set(etiquetas_por_id.get(i, {}).get('etiquetas', [])) & EXCLUIR_DE_POOL))
    reciente['df'] = _rdf.assign(_excluir_pool=_rara.values)
    if _rara.any():
        log(f"  {int(_rara.sum())} avisos con situación rara (dañado/legal/modificado) "
            f"excluidos del pool de comparables.")

    log(f"Calculando precio justo sobre captura {reciente['fecha'].date()} ...")
    _dfx = reciente['df'].copy()
    # Garante de columnas: si la captura analizada viene cruda (sin specs,
    # fechas u otras columnas opcionales), primero se RESCATAN por id desde
    # las otras capturas cargadas (la del estado viene enriquecida); solo lo
    # irrescatable se crea vacio, con aviso.
    _cols_op = ('url', 'version_oficial', 'Cilindrada', 'Combustible',
                'Transmisión', 'Tracción', 'traccion', 'caja',
                'combustible', 'motor', 'entidad', 'primera_vez')
    _por_rescatar = [c for c in _cols_op if c not in _dfx.columns]
    if _por_rescatar:
        _idx = _dfx['id'].astype(str).str.lower()
        for _cap in reversed(capturas[:-1]):
            _cdf = _cap.get('df')
            if _cdf is None or 'id' not in getattr(_cdf, 'columns', []):
                continue
            _tiene = [c for c in _por_rescatar if c in _cdf.columns]
            if not _tiene:
                continue
            _aux = _cdf.drop_duplicates(subset='id').copy()
            _aux['_k'] = _aux['id'].astype(str).str.lower()
            _aux = _aux.set_index('_k')
            for _c in _tiene:
                _dfx[_c] = _idx.map(_aux[_c]).values
            log("  Columnas rescatadas por id desde la captura "
                f"{_cap['fecha'].date()}: {', '.join(_tiene)}")
            _por_rescatar = [c for c in _por_rescatar if c not in _tiene]
            if not _por_rescatar:
                break
    _faltantes = []
    for _col in _por_rescatar:
        _dfx[_col] = pd.NA
        _faltantes.append(_col)
    if _faltantes:
        log("  AVISO: columnas ausentes creadas vacias: " + ', '.join(_faltantes)
            + " (revisar especificaciones.csv / fuentes)")
    dfa = calcular_precio_justo(_dfx)

    n_vof    = (dfa['metodo'].str.endswith('vof')).sum()
    n_motor  = (dfa['metodo'].str.endswith('motor')).sum()
    n_comb   = (dfa['metodo'].str.endswith('comb')).sum()
    n_modelo = (dfa['metodo'].str.endswith('modelo')).sum()
    n_sin    = (dfa['metodo'] == '').sum()
    log(f"  Pasada 0 (versión oficial): {n_vof} | Pasada 1 (motor+comb): {n_motor} | "
        f"Pasada 2 (comb): {n_comb} | Fallback (modelo): {n_modelo} | Sin tasar: {n_sin}")

    log("Calculando rotacion entre capturas ...")
    rotacion = calcular_rotacion(capturas)

    def cap2(s):
        return ' '.join(w.capitalize() for w in str(s).split())
    vol_modelo = dfa.groupby(['marca', 'modelo'])['precio'].size()


    # Fichas técnicas completas (especificaciones_full.json), si el goteo
    # de specs ya bajó algunas: se adjuntan a las oportunidades para verlas
    # con el botón 📋 del radar.
    fichas_full = {}
    if os.path.exists("especificaciones_full.json"):
        try:
            with open("especificaciones_full.json", encoding="utf-8") as f:
                for fid, fdata in json.load(f).items():
                    fichas_full[str(fid).lower()] = (fdata or {}).get("campos") or {}
            log(f"Fichas técnicas completas cargadas: {len(fichas_full)}")
        except Exception as e:
            log(f"  aviso: no se pudo leer especificaciones_full.json ({e})")

    BANDERAS_ROJAS = {'daniado', 'problema_legal', 'km_dudoso',
                       'precio_condicionado', 'precio_mas_iva'}

    # ---- ORO POR COMENTARIO -------------------------------------------
    # Señales de negociacion en el texto (venta urgente, precio conversable,
    # se aceptan ofertas) ELEVAN un aviso a oportunidad de oro, aunque su
    # descuento no llegue al umbral clasico. Condiciones: vendedor Particular
    # (en automotoras estas frases son marketing), descuento real >= 8% y
    # ninguna bandera roja en el comentario.
    SENALES_NEGOCIO = {'venta_urgente', 'precio_conversable'}
    ids_negociables = set()
    for _vid, _inf in etiquetas_por_id.items():
        _ets = set(_inf.get('etiquetas', []))
        if (_ets & SENALES_NEGOCIO) and not (_ets & BANDERAS_ROJAS):
            ids_negociables.add(_vid)
    if ids_negociables and 'op_oro' in dfa.columns:
        _mask_neg = (
            dfa['id'].astype(str).str.lower().isin(ids_negociables)
            & (dfa['entidad'].astype(str).str.strip().str.lower() == 'particular')
            & (dfa['descuento_pct'] >= 8)
            & dfa['precio_justo'].notna()
        )
        _nuevos_oro = int((_mask_neg & ~dfa['op_oro']).sum())
        dfa.loc[_mask_neg, 'op_oro'] = True
        log(f"Oro por comentario (particular + señal de negociacion, sin "
            f"banderas): {int(_mask_neg.sum())} avisos ({_nuevos_oro} elevados)")

    _hoy_ts = pd.Timestamp.now().normalize()

    def _fecha_iso(v):
        if v is None or pd.isna(v):
            return ''
        txt = str(v).strip()
        d = pd.to_datetime(txt[:10], format='%Y-%m-%d', errors='coerce')
        if pd.isna(d):
            d = pd.to_datetime(txt.split()[0] if txt else txt, dayfirst=True, errors='coerce')
        return d.strftime('%Y-%m-%d') if pd.notna(d) else '' 

    def _dias_pub(pv):
        if pv is None or pd.isna(pv):
            return None
        txt = str(pv).strip()
        d = pd.to_datetime(txt[:10], format='%Y-%m-%d', errors='coerce')
        if pd.isna(d):
            d = pd.to_datetime(txt.split()[0] if txt else txt,
                               dayfirst=True, errors='coerce')
        if pd.isna(d):
            return None
        dias = int((_hoy_ts - d.normalize()).days)
        return dias if dias >= 0 else 0

    def fila_op(r):
        vid = str(r['id']).lower()
        info = etiquetas_por_id.get(vid, {})
        etiquetas = list(info.get('etiquetas', []))
        _kw_no = _comentario_no_oportunidad(vid)
        if _kw_no:
            etiquetas.append('no oportunidad: ' + _kw_no)
        ant = int(r['antiguedad'])
        km = int(r['kilometraje'])
        km_imposible = (ant >= 2 and km < 1000) or (ant >= 5 and km < 5000)
        if km_imposible and 'km_dudoso' not in etiquetas:
            etiquetas.append('km_dudoso')
        return {
            'id': vid,
            'marca': cap2(r['marca']), 'modelo': cap2(r['modelo']),
            'version': (str(r['version'])[:42] if pd.notna(r['version']) else ''),
            'version_original': (str(r['version']) if pd.notna(r['version']) else ''),
            'tasada_como': (str(r['firma_ver']) if ('firma_ver' in r.index and pd.notna(r['firma_ver'])) else ''),
            'version_oficial': (str(r['version_oficial'])
                                if ('version_oficial' in r.index and pd.notna(r['version_oficial'])
                                    and str(r['version_oficial']).strip()) else ''),
            'generacion': (str(r['gen']) if 'gen' in r.index else ''),
            'año': int(r['año']), 'km': int(r['kilometraje']), 'precio': int(r['precio']),
            'precio_justo': int(r['precio_justo']) if pd.notna(r['precio_justo']) else None,
            'descuento': round(float(r['descuento_pct']), 1) if pd.notna(r['descuento_pct']) else None,
            'km_anual': int(r['km_anual']),
            'entidad': r['entidad'], 'comparables': int(r['comparables']),
            'metodo': r['metodo'],
            'mercado': int(vol_modelo.get((r['marca'], r['modelo']), 0)),
            'url': (r.get('url') if pd.notna(r.get('url')) else ''),
            'f_oro': (bool(r['op_oro'])) and not _kw_no, 'f_precio': bool(r['op_precio']), 'f_km': bool(r['op_km']),
            'desc_extremo': bool(r.get('descuento_extremo', False)),
            'km_norm': (int(r['km_norm']) if pd.notna(r.get('km_norm')) else None),
            'etiquetas': etiquetas,
            'region': _UBI.get(vid, {}).get('region', ''),
            'comuna': _UBI.get(vid, {}).get('comuna', ''),
            'fecha_pub': _UBI.get(vid, {}).get('fecha_pub', ''),
            'bandera_roja': bool(set(etiquetas) & BANDERAS_ROJAS),
            'tiene_comentario': bool(info.get('comentario')),
            'comentario': info.get('comentario', ''),
            'gen': str(r.get('gen', '') or ''),
            'gens_pos': str(r.get('gens_pos', '') or ''),
            'urgencia': int(info.get('urgencia', 0) or 0),
            'resumen_ia': info.get('resumen_ia', ''),
            'ficha': fichas_full.get(vid) or None,
            'primera_vez': _fecha_iso(r['primera_vez'] if 'primera_vez' in r.index else None),
            'ultima_vez': _fecha_iso(r['ultima_vez'] if 'ultima_vez' in r.index else None),
            'dias_publicado': _dias_pub(r['primera_vez'] if 'primera_vez' in r.index else None),
        }

    valido = dfa[dfa['comparables'] >= MIN_COMPARABLES]
    pool = pd.concat([
        valido[valido['op_oro']].sort_values('descuento_pct', ascending=False).head(200),
        valido[valido['op_precio']].sort_values('descuento_pct', ascending=False).head(200),
        valido[valido['op_km']].sort_values('km_anual').head(200),
    ]).drop_duplicates(subset='id')
    oportunidades = [fila_op(r) for _, r in pool.iterrows()]

    # Comparables: por cada oportunidad, los autos del mismo marca+modelo a +-3 años
    CAP_COMP = 250
    _COMENT_TXT.clear()
    try:
        if os.path.exists('comentarios.json'):
            with open('comentarios.json', encoding='utf-8') as _f:
                _cj0 = json.load(_f)
            for _k0, _v0 in (_cj0.items() if isinstance(_cj0, dict) else []):
                _t0 = _v0 if isinstance(_v0, str) else (_v0 or {}).get('comentario', '')
                if _t0:
                    _COMENT_TXT[str(_k0).lower().strip()] = str(_t0)[:400]
    except Exception:
        pass

    # perfiles oficiales por (modelo, gen, motor, acabados) para que la
    # sugerida deducida adopte el nombre de la ficha cuando coincide
    _VOF_PERFIL.clear()
    _cvof = dfa[dfa['version_oficial'].notna()
                & (dfa['version_oficial'].astype(str).str.strip() != '')
                & (dfa['version_oficial'].astype(str).str.lower() != 'nan')]
    _tmp_perfil = {}
    for _, _rv in _cvof.iterrows():
        _kk = (_norm_txt(_rv.get('marca')), _norm_txt(_rv.get('modelo')),
               str(_rv.get('gen', '') or ''),
               (str(_rv.get('motor')) if pd.notna(_rv.get('motor')) else ''),
               tuple(_acabados_multi(_rv.get('version'))))
        _tmp_perfil.setdefault(_kk, {}).setdefault(str(_rv['version_oficial']), 0)
        _tmp_perfil[_kk][str(_rv['version_oficial'])] += 1
    for _kk, _cnt in _tmp_perfil.items():
        _VOF_PERFIL[_kk] = max(_cnt, key=_cnt.get)
    log(f"  Perfiles ficha para sugeridas: {len(_VOF_PERFIL)}")
    cargar_segmentos()
    construir_motor_dominante(dfa)
    construir_consolidacion(dfa)

    _comp_estricto = {}
    comps = {}
    pool_ids = set(pool['id'].astype(str).str.lower())
    # Detectar tracción y cabina del texto para comparar peras con peras
    def _trac(t):
        t = str(t).lower()
        if any(x in t for x in ('4x4', '4wd', 'awd')): return '4x4'
        if any(x in t for x in ('4x2', '4wd', 'rwd', 'fwd')): return '4x2'
        return ''
    def _cabina(t):
        t = str(t).lower()
        if any(x in t for x in ('reg cab', 'cab. sim', 'cabina simple', 'single')): return 'reg'
        if any(x in t for x in ('dob', 'doble', 'cc ', 'crew', 'dcab', 'd cab')): return 'dcab'
        return ''
    dfa = dfa.copy()
    dfa['_trac_c'] = dfa['version'].map(_trac)
    dfa['_cab_c'] = dfa['version'].map(_cabina)

    for _, r in pool.iterrows():
        vid = str(r['id']).lower()
        g = dfa[(dfa['marca'] == r['marca']) & (dfa['modelo'] == r['modelo'])]
        # Filtro de GENERACIÓN: no comparar un 3008 gen I (2013) con gen II
        # (2018). PERO en años de transición (una gen se lanzó a mitad de año,
        # conviven dos), la asignación por año es ambigua -> NO se filtra por
        # gen, se muestran ambas y el usuario destilda mirando las fotos.
        r_gen = str(r.get('gen', '') or '').strip()
        r_trans = str(r.get('gens_pos', '') or '').strip()
        if r_gen and not r_trans:
            # aviso con gen clara: comparar solo con su gen y con los de años
            # de transición (que podrían ser de su gen).
            g = g[(g['gen'].astype(str).str.strip() == r_gen)
                  | (g['gens_pos'].astype(str).str.strip() != '')]
        # Comparables de VERDAD: mismo motor, combustible, caja, tracción y
        # cabina cuando el aviso los declara. Si el aviso no declara alguno,
        # no se filtra por ese (para no quedarse sin comparables).
        r_mot, r_comb = r.get('motor'), r.get('combustible')
        r_caja = r.get('caja', '')
        r_trac = _trac(r['version']); r_cab = _cabina(r['version'])
        # Versiones bestia: separación dura en ambos sentidos.
        BESTIAS = ['raptor', 'amg', 'rubicon', 'sti', 'gti']
        rv = str(r['version']).lower()
        r_bestia = next((x for x in BESTIAS if re.search(r'\b'+x+r'\b', rv)), '')
        gv_low = g['version'].astype(str).str.lower()
        if r_bestia:
            g = g[gv_low.str.contains(r'\b'+r_bestia+r'\b', regex=True)]
        else:
            patt = '|'.join(r'\b'+x+r'\b' for x in BESTIAS)
            g = g[~gv_low.str.contains(patt, regex=True)]
        # traccion estricta: fuera los comparables de traccion CONTRARIA
        _rtrc = str(r.get('_trc', '') or '')
        if _rtrc and '_trc' in g.columns:
            g = g[(g['_trc'] == '') | (g['_trc'] == _rtrc)]
        _rcj = str(r.get('_cj_ef', '') or '')
        if _rcj and '_cj_ef' in g.columns:
            g = g[(g['_cj_ef'] == '') | (g['_cj_ef'] == _rcj)]
        _rmo = str(r.get('_mot_ef', '') or '')
        if _rmo and '_mot_ef' in g.columns:
            g = g[(g['_mot_ef'] == '') | (g['_mot_ef'] == _rmo)]
        _rcb = str(r.get('_cab_ef', '') or '')
        if _rcb and '_cab_ef' in g.columns:
            g = g[(g['_cab_ef'] == '') | (g['_cab_ef'] == _rcb)]
        _rby = str(r.get('_body_ef', '') or '')
        if _rby and '_body_ef' in g.columns:
            g = g[(g['_body_ef'] == '') | (g['_body_ef'] == _rby)]
        r_acab = _acabado_txt(r['version'])
        if pd.notna(r_mot) and str(r_mot):
            # Aviso CON motor: comparar solo con mismo motor. Los comparables
            # sin motor declarado pero con el MISMO acabado también entran
            # (probablemente son la misma versión, solo que no lo escribieron).
            if r_acab:
                g = g[(g['motor'] == r_mot)
                      | ((g['motor'].isna() | (g['motor'].astype(str) == ''))
                         & (g['version'].map(_acabado_txt) == r_acab))]
            else:
                g = g[g['motor'] == r_mot]
        elif r_acab:
            # Aviso SIN motor pero CON acabado: filtrar por acabado. Un "allure"
            # (1.6) no se compara con un "gt" (2.0). Además, si hay comparables
            # con el mismo acabado que SÍ declaran motor, tomar el motor
            # dominante de ese acabado y filtrar por él (separa 1.6 de 2.0).
            g = g[g['version'].map(_acabado_txt) == r_acab]
            _con_mot = g[g['motor'].notna() & (g['motor'].astype(str) != '')]
            if len(_con_mot) >= 5:
                _mot_dom = _con_mot['motor'].mode()
                if len(_mot_dom):
                    md = _mot_dom.iloc[0]
                    g = g[(g['motor'] == md)
                          | (g['motor'].isna() | (g['motor'].astype(str) == ''))]
        # Si el aviso no tiene ni motor ni acabado ("—"), no hay cómo separar:
        # queda el grupo del modelo/gen. Esos solo los resuelve el ojo humano.
        if pd.notna(r_comb) and str(r_comb):
            g = g[g['combustible'] == r_comb]
        if r_caja:
            g = g[(g['caja'] == r_caja) | (g['caja'] == '')]
        if r_trac:
            g = g[(g['_trac_c'] == r_trac) | (g['_trac_c'] == '')]
        if r_cab:
            g = g[(g['_cab_c'] == r_cab) | (g['_cab_c'] == '')]
        peers = g[(g['año'] - r['año']).abs() <= 3].copy()
        full = len(peers)
        if full > CAP_COMP:
            self_row = peers[peers['id'].str.lower() == vid]
            otros = peers[peers['id'].str.lower() != vid].sort_values('precio')
            idx = np.linspace(0, len(otros) - 1, CAP_COMP - 1).astype(int)
            peers = pd.concat([self_row, otros.iloc[idx]]).drop_duplicates(subset='id')
        peers = peers.sort_values('precio', ascending=False)
        filas = []
        for _, p in peers.iterrows():
            _ant = int(p['antiguedad']) if pd.notna(p.get('antiguedad')) else 0
            _kmp = p['kilometraje'] if pd.notna(p['kilometraje']) else 0
            _kmdud = (_ant >= 2 and _kmp < 1000) or (_ant >= 5 and _kmp < 5000)
            filas.append([
                int(p['año']), int(round(p['kilometraje'] / 1000)), int(p['precio']),
                (str(p['version'])[:24] if pd.notna(p['version']) else ''),
                ('P' if p['entidad'] == 'Particular' else 'A'),
                1 if str(p['id']).lower() == vid else 0,
                str(p['id']).lower(),
                str(p.get('gen', '') or ''),
                str(p.get('gens_pos', '') or ''),
                1 if _kmdud else 0,
                _sugerida_final(p, f"{_norm_txt(p.get('marca'))}|{_norm_txt(p.get('modelo'))}"),
                (str(p['url']) if pd.notna(p.get('url')) else ''),
            ])
        comps[vid] = {'n': full, 'rows': filas}
        _comp_estricto[vid] = full

    # Coherencia radar<->modal: el nº de comparables del radar = filtro estricto.
    for _o in oportunidades:
        _fe = _comp_estricto.get(_o['id'])
        if _fe is not None:
            _o['comparables'] = _fe
            if _fe < MIN_COMPARABLES:
                _o['tasacion_pobre'] = True
                # sin base real no hay tasación defendible: retirar el número
                _o['fair'] = None
                _o['desc'] = None
                _o['f_oro'] = False
                _o['f_precio'] = False

    # --- Hoja de depuración: sugerencias por marca/modelo/gen/firma ---
    corr_dep = cargar_correcciones()
    dep_src = dfa.copy()
    dep_src['_firma'] = dep_src['firma_ver'].fillna('')
    dep_src['_ver'] = dep_src['version'].fillna('').astype(str).str.strip()
    vol_mod = dep_src.groupby(['marca', 'modelo'])['precio'].transform('size')
    dep_src = dep_src[vol_mod >= 15]  # modelos con mercado real
    perfiles_ver = construir_perfiles_versiones(dep_src)
    if perfiles_ver:
        log(f"Perfiles de versiones dominantes (con ficha): "
            f"{sum(len(v) for v in perfiles_ver.values())} en "
            f"{len(perfiles_ver)} modelo/generación")
    depuracion = []
    for (ma, mo, gn, fv), g in dep_src.groupby(['marca', 'modelo', 'gen', '_firma']):
        if len(g) < 3:
            continue
        clave = f"{_norm_txt(ma)}|{_norm_txt(mo)}|{gn}|{fv}"
        guardado = corr_dep.get(clave) or {}
        ejemplos = [x for x in g['_ver'].drop_duplicates().head(3) if x]
        specs_vistas = ''
        if 'version_oficial' in g.columns:
            vofs = [str(x).strip() for x in g['version_oficial'].dropna().unique()
                    if str(x).strip()]
            specs_vistas = ' | '.join(vofs[:3])
        # Dispersión de precio ajustada por año: si dentro del MISMO año del
        # mismo grupo los precios varían demasiado, probablemente conviven
        # dos versiones que la firma no distingue -> revisar primero.
        disp = 0.0
        por_año = g.groupby('año')['precio']
        años_validos = [a for a, s in por_año if len(s) >= 3]
        if años_validos:
            ratios = []
            for a, s in por_año:
                if len(s) >= 3 and s.median() > 0:
                    q1, q3 = s.quantile([0.25, 0.75])
                    ratios.append(float((q3 - q1) / s.median()))
            if ratios:
                disp = round(max(ratios), 2)
        depuracion.append({
            'clave': clave, 'marca': ma, 'modelo': mo, 'gen': gn, 'firma': fv,
            'avisos': int(len(g)),
            'precio_mediano': int(g['precio'].median()),
            'años': f"{int(g['año'].min())}-{int(g['año'].max())}",
            'ejemplos': ' | '.join(ejemplos)[:120],
            'sugerida': sugerir_canonica(fv),
            'canonica': guardado.get('canonica', ''),
            'excluir': bool(guardado.get('excluir')),
            'dispersion': disp,
            'revisar': disp >= 0.35,  # rango intercuartil > 35% de la mediana en un mismo año
            'specs': specs_vistas,
            'precal': None,
        })
        # Pre-calificacion: solo para grupos SIN canonica y SIN ficha propia,
        # cuando el modelo tiene versiones dominantes conocidas
        if (not guardado.get('canonica') and not specs_vistas
                and len(g) <= 30 and (ma, mo, gn) in perfiles_ver):
            r = precalificar_grupo(g, fv, ejemplos, perfiles_ver[(ma, mo, gn)])
            if r:
                depuracion[-1]['precal'] = {
                    'version': r[0], 'conf': r[1], 'senales': r[2],
                    'fuente': r[3] if len(r) > 3 else 'specs'}
    depuracion.sort(key=lambda d: (-d['avisos']))

    # ---- TASADOR: tablas de precio historico por nivel de detalle ----
    # Nivel 'oficial' (version de ficha), 'firma' (version del texto
    # normalizada) y 'modelo' (todo el modelo/gen). El dashboard tasa
    # eligiendo el nivel mas fino con datos.
    tasador = []
    _src_t = dfa[(dfa['precio'] > 0) & dfa['año'].notna()].copy()
    _src_t['_fv'] = _src_t.get('firma_ver', pd.Series('', index=_src_t.index)).fillna('')

    def _agrega_tasa(gb, nivel, min_n):
        for llav, g in gb:
            if len(g) < min_n:
                continue
            *cab, an = llav
            ma, mo, gn = cab[0], cab[1], cab[2]
            ver = cab[3] if len(cab) > 3 else ''
            tasador.append({
                'ma': ma, 'mo': mo, 'gen': str(gn), 'ver': str(ver),
                'nivel': nivel, 'año': int(an), 'n': int(len(g)),
                'p50': int(g['precio'].median()),
                'p25': int(g['precio'].quantile(0.25)),
                'p75': int(g['precio'].quantile(0.75)),
                'km50': int(g['kilometraje'].median()) if g['kilometraje'].notna().any() else 0,
            })

    if 'version_oficial' in _src_t.columns:
        _cv = _src_t[_src_t['version_oficial'].fillna('').astype(str).str.strip() != '']
        _agrega_tasa(_cv.groupby(['marca', 'modelo', 'gen', 'version_oficial', 'año']),
                     'oficial', 2)
    _cf = _src_t[_src_t['_fv'] != '']
    _agrega_tasa(_cf.groupby(['marca', 'modelo', 'gen', '_fv', 'año']), 'firma', 3)
    _agrega_tasa(_src_t.groupby(['marca', 'modelo', 'gen', 'año']), 'modelo', 3)
    log(f"Tasador: {len(tasador)} celdas de precio historico exportadas")

    from collections import Counter as _Counter
    cont_etq = _Counter()
    for o in oportunidades:
        for e in o['etiquetas']:
            cont_etq[e] += 1

    # ---- ANÁLISIS DE MERCADO (visión de analista) ----
    # Trabaja sobre toda la base limpia y activa (dfa), no solo el radar.
    analisis = construir_analisis(dfa)
    depurador_modelos = construir_depurador_modelos(dfa)
    log(f"Depurador por modelo: {len(depurador_modelos)} modelos con versiones propuestas")

    # --- Tabla de generaciones para la pestaña de corrección (dashboard) ---
    # Cortes actuales por modelo + conteo de avisos por generación asignada,
    # para que el usuario vea/edite los rangos y detecte errores de origen.
    gen_edit = {}
    _tabla_gen_export = cargar_generaciones()
    for (ma, mo), rangos in (_tabla_gen_export or {}).items():
        key = f"{ma}|{mo}"
        gen_edit[key] = {
            'marca': ma, 'modelo': mo,
            'rangos': [{'gen': g, 'desde': int(d),
                        'hasta': (int(h) if h < 9999 else None)}
                       for d, h, g in sorted(rangos)],
        }
    # conteo real de avisos por modelo/gen (para mostrar cuántos hay en cada una)
    _cnt = (dfa.groupby([dfa['marca'].map(_norm_txt),
                         dfa['modelo'].map(_norm_txt),
                         dfa['gen'].astype(str)]).size())
    for (ma, mo, g), n in _cnt.items():
        key = f"{ma}|{mo}"
        if key in gen_edit:
            gen_edit[key].setdefault('conteo', {})[g or '(sin gen)'] = int(n)
    # modelos con avisos pero SIN tabla de generación (candidatos a agregar)
    _modelos_con_avisos = sorted(set(
        (a, b) for a, b in zip(dfa['marca'].map(_norm_txt), dfa['modelo'].map(_norm_txt))))
    for ma, mo in _modelos_con_avisos:
        key = f"{ma}|{mo}"
        if key not in gen_edit:
            gen_edit[key] = {'marca': ma, 'modelo': mo, 'rangos': [], 'conteo': {}}

    # --- Depuración BBDD: tabla plana por modelo (spec del usuario 2026-07-10) ---
    _coment_txt = {k: v[:160] for k, v in _COMENT_TXT.items()}

    depuracion_avisos = {}
    _man_d = dfa['marca'].map(_norm_txt)
    _mon_d = dfa['modelo'].map(_norm_txt)
    for _i, r in dfa.iterrows():
        key = f"{_man_d.at[_i]}|{_mon_d.at[_i]}"
        _id = str(r.get('id', '') or '').lower().strip()
        def _s(v, n=24):
            return ('' if pd.isna(v) else str(v))[:n]
        depuracion_avisos.setdefault(key, []).append([
            _id,
            _s(r.get('gen'), 16),
            int(r['año']) if pd.notna(r.get('año')) else 0,
            _s(r.get('version'), 40),
            _sugerida_final(r, key),
            int(round((r.get('kilometraje') or 0) / 1000)) if pd.notna(r.get('kilometraje')) else 0,
            int(r['precio']) if pd.notna(r.get('precio')) else 0,
            _s(r.get('Cilindrada'), 8),
            _s(r.get('Combustible'), 12),
            _s(r.get('Transmisión'), 12),
            _s(r.get('Tracción'), 8),
            _coment_txt.get(_id, ''),
            ', '.join(t.replace('_', ' ') for t in
                      (etiquetas_por_id.get(_id, {}) or {}).get('etiquetas', [])),
            (str(r['url']) if pd.notna(r.get('url')) else ''),
            _s(r.get('entidad'), 12),
            (lambda _d: _d if _d is not None else -1)(_dias_pub(r['primera_vez'] if 'primera_vez' in r.index else None)),
            (int(round(r['descuento'])) if pd.notna(r.get('descuento')) else None),
            _seg_carroceria(r.get('marca'), r.get('modelo')),
        ])
    # ordenar cada modelo por gen y año (como el Excel del usuario)
    for _k in depuracion_avisos:
        depuracion_avisos[_k].sort(key=lambda x: (x[1] or 'zzz', x[2]))
    log(f"Depuración BBDD: {sum(len(v) for v in depuracion_avisos.values())} avisos "
        f"en {len(depuracion_avisos)} modelos exportados.")

    # --- Fichas tecnicas completas para la pestania Depuracion ---
    fichas_full = {}
    try:
        if os.path.exists('especificaciones_full.json'):
            with open('especificaciones_full.json', encoding='utf-8') as _f:
                _ff = json.load(_f)
            _ids_base = set(dfa['id'].astype(str).str.lower().str.strip())
            for _k, _v in (_ff.items() if isinstance(_ff, dict) else []):
                _kl = str(_k).lower().strip()
                if _kl in _ids_base and isinstance(_v, dict):
                    fichas_full[_kl] = {str(a): str(b)[:80] for a, b in _v.items()
                                        if b not in (None, '', 'nan')}
            log(f"  Fichas completas para depuracion: {len(fichas_full)}")
    except Exception as _e:
        log(f"  (especificaciones_full.json ilegible: {_e})")

    salida = {
        'generado': datetime.now().isoformat(timespec='seconds'),
        'capturas': [c['fecha'].strftime('%Y-%m-%d') for c in capturas],
        'captura_analizada': reciente['fecha'].strftime('%Y-%m-%d'),
        'resumen': {
            'total': int(len(dfa)),
            'precio_mediano': int(dfa['precio'].median()),
            'op_oro': int(dfa['op_oro'].sum()),
            'op_precio': int(dfa['op_precio'].sum()),
            'op_km': int(dfa['op_km'].sum()),
        },
        'etiquetas_conteo': dict(cont_etq),
        'oportunidades': oportunidades,
        'comps': comps,
        'rotacion': rotacion,
        'depuracion': depuracion,
        'tasador': tasador,
        'analisis': analisis,
        'depurador_modelos': depurador_modelos,
        'generaciones_edit': gen_edit,
        'depuracion_avisos': depuracion_avisos,
        'fichas_full': fichas_full,
    }

    with open(SALIDA, 'w', encoding='utf-8') as f:
        json.dump(salida, f, ensure_ascii=False, indent=2)

    # --- Excel enriquecido del dia: base completa + generaciones anexadas ---
    try:
        from datetime import date as _date
        _man_x = dfa['marca'].map(_norm_txt)
        _mon_x = dfa['modelo'].map(_norm_txt)
        _exp = pd.DataFrame({
            'id': dfa['id'],
            'marca': dfa['marca'],
            'modelo': dfa['modelo'],
            'generacion': dfa['gen'],
            'año': dfa['año'],
            'version_vendedor': dfa['version'],
            'version_ajustada': [
                _sugerida_final(r, f"{_man_x.at[_i]}|{_mon_x.at[_i]}")
                for _i, r in dfa.iterrows()],
            'kilometraje': dfa['kilometraje'],
            'precio': dfa['precio'],
            'traccion': dfa.get('_trc', ''),
            'caja': dfa.get('_cj_ef', ''),
            'motor': dfa.get('_mot_ef', ''),
            'cabina': dfa.get('_cab_ef', ''),
            'segmento': [_seg_carroceria(a, b) for a, b in
                         zip(dfa['marca'], dfa['modelo'])],
            'vendedor': dfa.get('entidad', ''),
            'version_oficial_ficha': dfa.get('version_oficial', ''),
            'cilindrada_ficha': dfa.get('Cilindrada', ''),
            'combustible_ficha': dfa.get('Combustible', ''),
            'transmision_ficha': dfa.get('Transmisión', ''),
            'traccion_ficha': dfa.get('Tracción', ''),
            'dias_en_mercado': [
                (lambda _d: _d if _d is not None else '')(
                    _dias_pub(r['primera_vez'] if 'primera_vez' in r.index else None))
                for _i, r in dfa.iterrows()],
            'precio_justo': dfa.get('precio_justo', ''),
            'descuento_pct': dfa.get('descuento', ''),
            'etiquetas': [
                ', '.join((etiquetas_por_id.get(str(_v).lower().strip(), {}) or {})
                          .get('etiquetas', []))
                for _v in dfa['id']],
            'comentario_vendedor': [
                str(_coment_txt.get(str(_v).lower().strip(), ''))[:500]
                for _v in dfa['id']],
            'url': dfa.get('url', ''),
        })
        _fx = f"base_depurada_{_date.today().isoformat()}.xlsx"
        _exp.to_excel(_fx, index=False, sheet_name='Base')
        log(f"Excel del dia exportado: {_fx} ({len(_exp)} avisos, "
            f"{int(( _exp['generacion'].astype(str).str.strip() != '').sum())} con generacion)")
    except Exception as _e:
        log(f"(no pude exportar el Excel del dia: {_e})")

    log(f"Listo. Escrito '{SALIDA}'.")

    # Guardar Excel con la base analizada
    fecha_str = reciente['fecha'].strftime('%Y-%m-%d')
    nombre_excel = f"base_{fecha_str}.xlsx"
    cols_excel = [
        'marca', 'modelo', 'version', 'motor', 'año', 'kilometraje', 'precio',
        'precio_justo', 'descuento_pct', 'km_anual', 'comparables',
        'metodo', 'entidad', 'transmision', 'combustible', 'traccion',
        'condicion', 'op_oro', 'op_precio', 'op_km', 'url'
    ]
    cols_presentes = [c for c in cols_excel if c in dfa.columns]
    dfa[cols_presentes].to_excel(nombre_excel, index=False)
    log(f"Excel guardado: '{nombre_excel}' ({len(dfa)} filas).")

    print("\n--- RESUMEN ---")
    print(f"Capturas: {', '.join(salida['capturas'])}")
    r = salida['resumen']
    print(f"Captura analizada: {salida['captura_analizada']} ({r['total']} avisos limpios)")
    print(f"Oportunidades -> oro: {r['op_oro']} | bajo precio: {r['op_precio']} | bajo km: {r['op_km']}")
    if rotacion and rotacion.get('valida'):
        g = rotacion['global']
        print(f"\nRotacion {g['fecha_desde']} -> {g['fecha_hasta']} ({g['dias']} dias):")
        print(f"  Salieron del mercado: {g['desaparecidos']} de {g['avisos_inicio']} "
              f"({g['tasa_salida_periodo']}% | {g['tasa_salida_diaria']}%/dia)")
        print(f"  Cambios de precio: {g['con_cambio_precio']} "
              f"(bajaron {g['bajaron_precio']}, subieron {g['subieron_precio']})")
        alto = [m for m in rotacion['por_modelo'] if m['inventario'] == 'alto']
        bajo = [m for m in rotacion['por_modelo'] if m['inventario'] == 'bajo']
        print(f"\n  Modelos por inventario: {len(alto)} alto (>=10) | {len(bajo)} bajo (<10)")
        print("\n  ALTO INVENTARIO - rotacion confiable, mas rapidos:")
        for m in alto[:8]:
            dv = f"~{m['dias_venta_est']}d" if m['dias_venta_est'] else "s/d"
            print(f"    {m['marca']} {m['modelo']}: {m['tasa_salida']}% "
                  f"({m['salieron']}/{m['stock']}, venta est. {dv})")


if __name__ == '__main__':
    main()
