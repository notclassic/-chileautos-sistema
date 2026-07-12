#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MOTOR DE CALCULO - Capa 1 del sistema de inteligencia Chileautos
================================================================

Que hace:
  1. Lee TODAS las capturas (.xlsx) de la carpeta capturas/
  2. Limpia cada una con las mismas reglas validadas.
  3. Calcula precio justo y oportunidades sobre la captura mas reciente.
  4. Compara capturas consecutivas para medir ROTACION (avisos que
     desaparecen = proxy de venta) y CAMBIOS DE PRECIO.
  5. Cruza etiquetas de comentarios + comparables.
  6. Escribe datos.json con todo lo calculado.
  7. Guarda base_FECHA.xlsx con la base analizada.

El dashboard lee datos.json. El motor no toca el dashboard.

Uso:
    python motor.py
    (lee carpeta ./capturas, escribe ./datos.json)

Requiere: pandas, numpy, openpyxl
"""

import os, sys, json, glob, re
from datetime import datetime
import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# CONFIGURACION
# ---------------------------------------------------------------------------
CARPETA_CAPTURAS = "capturas"
SALIDA = "datos.json"

COLUMNAS_ESPERADAS = [
    'fecha','marca','modelo','version','año','traccion','condicion','precio',
    'kilometraje','transmision','combustible','id','moneda','entidad',
    'Enlace Web','telefono','whatsapp','nombre_whatsapp'
]

# Reglas de limpieza (validadas con las bases reales)
PRECIO_MIN, PRECIO_MAX = 500_000, 300_000_000
KM_MIN, KM_MAX = 0, 500_000
ANIO_MIN, ANIO_MAX = 1990, 2027

# Reglas de oportunidad
DESCUENTO_MIN = 15.0      # % bajo precio justo para "bajo precio"
KM_ANUAL_MAX = 10_000     # umbral "bajo kilometraje"
MIN_COMPARABLES = 5       # minimo de comparables para tasar


def log(msg):
    print(f"[motor] {msg}")


# ---------------------------------------------------------------------------
# 1.a EXTRACCION DE MOTOR DESDE VERSION
# ---------------------------------------------------------------------------

def extraer_motor(version_str):
    """
    Extrae la cilindrada del texto libre del campo 'version'.
    Ejemplos: '1.5 allure pack bluehdi' -> '1.5'
              '2.0t elite diesel'        -> '2.0'   (sufijo turbo)
              '5.0l xlt 4x4'            -> '5.0'   (sufijo litros)
              'gt line puretech'         -> None    (sin número)
    Devuelve str o None.
    """
    if pd.isna(version_str) or not str(version_str).strip():
        return None
    # Normalizar sufijos comunes: "1.5t" -> "1.5", "5.0l" -> "5.0", "2.0i" -> "2.0"
    s = re.sub(r'(\d+\.\d+)[tTlLiI]\b', r'\1', str(version_str).lower())
    m = re.search(r'\b(\d+\.\d+)\b', s)
    return m.group(1) if m else None


FAMILIAS_MOTOR = [
    # (nombre_canonico, variantes a buscar en el texto SIN espacios ni guiones)
    ("bluehdi",  ["bluehdi"]),          # cubre "blue hdi", "blue-hdi"
    ("hybrid4",  ["hybrid4"]),
    ("puretech", ["puretech"]),
    ("thp",      ["thp"]),
    ("ecoboost", ["ecoboost"]),
    ("tfsi",     ["tfsi"]),
    ("tsi",      ["tsi"]),
    ("tdi",      ["tdi"]),
    ("dci",      ["dci"]),
    ("crdi",     ["crdi"]),
    ("vtec",     ["vtec"]),
    ("skyactiv", ["skyactiv"]),
    ("phev",     ["phev"]),
    ("mhev",     ["mhev", "mildhybrid"]),
    ("hibrido",  ["hybrid", "hibrido", "hev"]),   # genérico, después de hybrid4/phev/mhev
    ("hdi",      ["hdi"]),              # después de bluehdi
    ("vti",      ["vti"]),
]


CORRECCIONES_DEPURADOR = "versiones_canonicas.json"


def cargar_correcciones():
    """Correcciones hechas en la hoja Depuración del dashboard:
    { 'marca|modelo|gen|firma': {'canonica': 'nombre', 'excluir': bool} }"""
    if os.path.exists(CORRECCIONES_DEPURADOR):
        try:
            with open(CORRECCIONES_DEPURADOR, encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def sugerir_canonica(firma):
    """'1.6|thp|4x2' -> '1.6 THP 4x2' (sugerencia editable)."""
    if not firma:
        return ''
    partes = [p for p in str(firma).split('|') if p]
    out = []
    for p in partes:
        out.append(p.upper() if p not in ('4x2', '4x4') and not p[0].isdigit() else p)
    return ' '.join(out)


# ---------------------------------------------------------------------------
# PRE-CALIFICACION DE VERSIONES
# Un modelo suele concentrar ~90% de sus avisos en pocas versiones oficiales
# (dominantes). Cuando un grupo de avisos tiene version rara/sucia y sin
# ficha, el sistema compara sus señales duras (cilindrada, traccion,
# transmision, combustible, año, tokens del nombre y PRECIO) contra el
# perfil de cada version dominante y sugiere a cual pertenece.
# ---------------------------------------------------------------------------
import re as _re

_TRIM_STOP = {'diesel', 'bencina', 'gasolina', 'at', 'mt', 'aut', 'mec',
              '4x4', '4x2', 'awd', 'fwd', 'rwd', '2wd', '4wd', '4p', '5p',
              '2p', '3p', 'euro5', 'euro6', 'ac', 'dab', 'abs', 'll', 'hp',
              'cvt', 'dct', 'eat6', 'eat8', 'at6', 'at8', 'mt6', 'bvm', 's'}


def _cil_de_texto(t):
    mm = _re.search(r'\b(\d\.\d)\b', str(t))
    return float(mm.group(1)) if mm else None


def _trac_de_texto(t):
    t = str(t).lower()
    if any(x in t for x in ('4x4', 'awd', '4wd', 'quattro', '4motion', '4matic')):
        return '4x4'
    if any(x in t for x in ('4x2', 'fwd', 'rwd', '2wd')):
        return '4x2'
    return None


def _trans_de_texto(t):
    t = ' ' + str(t).lower() + ' '
    if any(x in t for x in (' at ', ' aut', 'cvt', 'dct', 'eat', 'tiptronic',
                            'stronic', 's-tronic', 'steptronic', 'automat')):
        return 'auto'
    if any(x in t for x in (' mt ', ' mec', 'bvm', 'manual')):
        return 'manual'
    return None


def _tokens_trim(t):
    return {w for w in _re.findall(r'[a-z]{3,}', str(t).lower())
            if w not in _TRIM_STOP}


def _moda(serie):
    try:
        s = serie.dropna().astype(str).str.strip().str.lower()
        s = s[s != '']
        return s.mode().iloc[0] if len(s) else None
    except Exception:
        return None


def construir_perfiles_versiones(dep_src):
    """Perfiles de versiones dominantes por (marca, modelo, gen).
    Dos fuentes, en orden de autoridad:
      1) 'specs': versiones OFICIALES (avisos con ficha) — la cedula.
      2) 'texto': versiones de texto que se repiten mucho (5+ avisos
         escritas igual) — dominantes de facto, utiles ANTES de tener
         fichas. Si un modelo tiene perfiles specs, mandan esos.
    """
    perfiles = {}
    # Fuente 2 primero (la fuente 1 la pisa si existe)
    txt = dep_src[dep_src['_ver'].astype(str).str.len() >= 6]
    for (ma, mo, gn, tv), g in txt.groupby(['marca', 'modelo', 'gen', '_ver']):
        if len(g) < 5:
            continue
        pre_año = {int(a): float(s.median())
                   for a, s in g.groupby('año')['precio'] if len(s) >= 1}
        perfiles.setdefault((ma, mo, gn), []).append({
            'version': str(tv).strip(), 'fuente': 'texto',
            'n': int(len(g)),
            'cil': _cil_de_texto(tv),
            'trac': _trac_de_texto(tv) or _moda(g.get('traccion')),
            'trans': _trans_de_texto(tv) or _moda(g.get('transmision')),
            'comb': _moda(g.get('combustible')),
            'años': (int(g['año'].min()), int(g['año'].max())),
            'precio_año': pre_año,
            'precio_med': float(g['precio'].median()),
            'tokens': _tokens_trim(tv),
        })
    if 'version_oficial' not in dep_src.columns:
        return perfiles
    con_vof = dep_src[dep_src['version_oficial'].notna()
                      & (dep_src['version_oficial'].astype(str).str.strip() != '')]
    claves_specs = set()
    for (ma, mo, gn, vof), g in con_vof.groupby(
            ['marca', 'modelo', 'gen', 'version_oficial']):
        if len(g) < 2:      # perfil necesita al menos 2 avisos con esa cedula
            continue
        claves_specs.add((ma, mo, gn))
        pre_año = {int(a): float(s.median())
                   for a, s in g.groupby('año')['precio'] if len(s) >= 1}
        perfiles.setdefault((ma, mo, gn), []).append({
            'version': str(vof).strip(), 'fuente': 'specs',
            'n': int(len(g)),
            'cil': _cil_de_texto(vof),
            'trac': _trac_de_texto(vof) or _moda(g.get('traccion')),
            'trans': _trans_de_texto(vof) or _moda(g.get('transmision')),
            'comb': _moda(g.get('combustible')),
            'años': (int(g['año'].min()), int(g['año'].max())),
            'precio_año': pre_año,
            'precio_med': float(g['precio'].median()),
            'tokens': _tokens_trim(vof),
        })
    # donde hay perfiles specs, los de texto se retiran (menor autoridad)
    for k in claves_specs:
        perfiles[k] = [p for p in perfiles[k] if p['fuente'] == 'specs']
    return perfiles


def precalificar_grupo(g, fv, ejemplos, candidatos):
    """Puntua el grupo raro contra cada version dominante. Devuelve
    (version, confianza, señales) o None."""
    partes = (fv or '').split('|')
    cil = None
    try:
        cil = float(partes[0]) if partes and partes[0] else None
    except ValueError:
        pass
    texto = ' '.join(ejemplos)
    señales_grupo = {
        'cil': cil or _cil_de_texto(texto),
        'trac': (partes[2] if len(partes) > 2 and partes[2] else None)
                or _trac_de_texto(texto) or _moda(g.get('traccion')),
        'trans': _trans_de_texto(texto) or _moda(g.get('transmision')),
        'comb': _moda(g.get('combustible')),
        'tokens': _tokens_trim(texto),
        'años': (int(g['año'].min()), int(g['año'].max())),
        'precio_año': {int(a): float(s.median())
                       for a, s in g.groupby('año')['precio']},
    }
    _propias = {str(e).strip().lower() for e in ejemplos}
    puntajes = []
    for p in candidatos:
        if p.get('fuente') == 'texto' and p['version'].strip().lower() in _propias:
            continue      # un grupo no se precalifica contra si mismo
        pts, razones = 0, []
        if señales_grupo['cil'] and p['cil']:
            if abs(señales_grupo['cil'] - p['cil']) < 0.05:
                pts += 3; razones.append('cilindrada')
            else:
                pts -= 4      # cilindrada distinta casi descarta
        if señales_grupo['trac'] and p['trac']:
            if señales_grupo['trac'] == p['trac']:
                pts += 2; razones.append('tracción')
            else:
                pts -= 2
        if señales_grupo['trans'] and p['trans']:
            if señales_grupo['trans'] == p['trans']:
                pts += 2; razones.append('transmisión')
            else:
                pts -= 2
        if señales_grupo['comb'] and p['comb']:
            if señales_grupo['comb'] == p['comb']:
                pts += 1; razones.append('combustible')
            else:
                pts -= 3
        a0, a1 = señales_grupo['años']
        if not (a1 < p['años'][0] - 1 or a0 > p['años'][1] + 1):
            pts += 1; razones.append('años')
        # precio: comparar mediana del mismo año cuando existe
        comunes = set(señales_grupo['precio_año']) & set(p['precio_año'])
        ref = None
        if comunes:
            a = max(comunes)
            ref = (señales_grupo['precio_año'][a], p['precio_año'][a])
        elif p['precio_med'] > 0 and señales_grupo['precio_año']:
            ref = (list(señales_grupo['precio_año'].values())[0], p['precio_med'])
        if ref and ref[1] > 0:
            dif = abs(ref[0] - ref[1]) / ref[1]
            if dif <= 0.18:
                pts += 2; razones.append('precio')
            elif dif > 0.40:
                pts -= 2
        inter = señales_grupo['tokens'] & p['tokens']
        if inter:
            pts += min(2, len(inter)); razones.append('nombre')
        puntajes.append((pts, p['n'], p['version'], razones))
    if not puntajes:
        return None
    puntajes.sort(reverse=True)
    mejor = puntajes[0]
    gap = mejor[0] - (puntajes[1][0] if len(puntajes) > 1 else -99)
    if mejor[0] >= 7 and gap >= 3:
        conf = 'alta'
    elif mejor[0] >= 5 and gap >= 2:
        conf = 'media'
    else:
        return None
    fuente = next((p.get('fuente', 'specs') for p in candidatos
                   if p['version'] == mejor[2]), 'specs')
    return (mejor[2], conf, ', '.join(mejor[3]), fuente)


def _norm_txt(s):
    """minúsculas, sin espacios/guiones/puntos — para casar marca/modelo."""
    return re.sub(r'[\s\-\./]', '', str(s).lower())


def cargar_generaciones(ruta="generaciones.xlsx"):
    """
    Lee la tabla editable de generaciones (marca, modelo, generacion,
    año_desde, año_hasta). Devuelve {(marca_norm, modelo_alias_norm):
    [(desde, hasta, etiqueta), ...]}. El campo MODELO admite alias
    separados por '/' (ej: 'Navara / NP300').
    """
    if not os.path.exists(ruta):
        return {}
    try:
        gdf = pd.read_excel(ruta, sheet_name="Generaciones", dtype=str)
    except Exception as e:
        log(f"  (no pude leer {ruta}: {e} — sigo sin generaciones)")
        return {}
    tabla = {}
    for _, r in gdf.iterrows():
        ma = _norm_txt(r.get('MARCA', ''))
        gen = str(r.get('GENERACION', '') or '').strip()
        try:
            desde = int(float(r.get('AÑO_DESDE')))
        except (TypeError, ValueError):
            continue
        hasta_raw = str(r.get('AÑO_HASTA', '') or '').strip()
        try:
            hasta = int(float(hasta_raw)) if hasta_raw else 9999
        except ValueError:
            hasta = 9999
        for alias in str(r.get('MODELO', '')).split('/'):
            alias = _norm_txt(alias)
            if ma and alias and gen:
                tabla.setdefault((ma, alias), []).append((desde, hasta, gen))
    return tabla


def asignar_generacion(df, tabla_gen):
    """Columna 'gen': etiqueta de generación según la tabla, o '' si el
    modelo no está en la tabla / el año no cae en ningún rango."""
    if not tabla_gen:
        df['gen'] = ''
        return df
    marcas_n = df['marca'].map(_norm_txt)
    modelos_n = df['modelo'].map(_norm_txt)
    gens = []
    for ma, mo, yr in zip(marcas_n, modelos_n, df['año']):
        rangos = tabla_gen.get((ma, mo))
        etiqueta = ''
        if rangos and pd.notna(yr):
            for d, h, g in rangos:
                if d <= yr <= h:
                    etiqueta = g
                    break
        gens.append(etiqueta)
    df['gen'] = gens
    n = sum(1 for g in gens if g)
    log(f"  Generaciones asignadas: {n} de {len(gens)} avisos "
        f"({len(tabla_gen)} modelos en la tabla).")
    return df


_ACABADOS_CANON = ['gt line', 'gt-line', 'gtline', 'gt', 'allure', 'active',
    'premier', 'premiere', 'feel', 'style', 'access', 'business', 'signature',
    'sport', 'limited', 'ltz', 'lt', 'ls', 'exclusive', 'dynamique', 'zen',
    'intens', 'life', 'trend', 'comfort', 'titanium', 'ambiente', 'highline',
    'comfortline', 'trendline', 'xei', 'xli', 'gli', 'gls', 'glx', 'gl', 'ex',
    'exl', 'lx', 'touring', 'value', 'full', 'sense']


def _acabado_txt(txt):
    t = ' ' + str(txt).lower() + ' '
    for ac in _ACABADOS_CANON:
        if ' ' + ac + ' ' in t or t.strip().endswith(' ' + ac):
            return ac
    return ''


def _aplicar_correcciones_versiones(df, corr_ver):
    """Aplica renombres/fusiones/exclusiones hechas en 'Arreglar versiones'.
    La clave de versión es cil|combustible|caja|acabado (igual que en el
    depurador). Renombrar/fusionar => escribe una version canonica comun en
    'version' (y recalcula firma). Excluir => marca el aviso para sacarlo."""
    ma = df['marca'].map(lambda s: str(s).lower())
    mo = df['modelo'].map(lambda s: str(s).lower())
    cil = df['motor'].fillna('').astype(str)
    comb = df['combustible'].fillna('').astype(str)
    caja = df['caja'].fillna('').astype(str)
    acab = df['version'].map(_acabado_txt)
    clave = cil + '|' + comb + '|' + caja + '|' + acab
    modelo_key = ma + '|' + mo

    excluir_mask = pd.Series(False, index=df.index)
    if '_ver_canon' not in df.columns:
        df['_ver_canon'] = ''
    for mk, cambios in corr_ver.items():
        sel_modelo = (modelo_key == mk)
        if not sel_modelo.any():
            continue
        renombres = cambios.get('renombres', {}) or {}
        for cl, nombre in renombres.items():
            sel = sel_modelo & (clave == cl)
            if sel.any():
                # firma canónica común => se agrupan juntos en el precio justo.
                # No tocamos 'version' para preservar la clave del depurador.
                df.loc[sel, 'firma_ver'] = 'canon:' + str(nombre).strip().lower()
                df.loc[sel, '_ver_canon'] = nombre
        for cl in (cambios.get('excluidas', []) or []):
            excluir_mask |= (sel_modelo & (clave == cl))
    if excluir_mask.any():
        n = int(excluir_mask.sum())
        df = df[~excluir_mask].copy()
        log(f"Arreglar versiones: {n} avisos excluidos por el usuario.")
    return df


def _norm_caja(transmision, version=''):
    """Devuelve 'AT', 'MT' o '' (desconocida) desde el campo transmision;
    si falta, lo infiere del texto de la versión."""
    t = (str(transmision) + ' ' + str(version)).lower()
    if any(x in t for x in ('auto', 'cvt', 'dct', 'tiptronic', 's-tronic',
                            'stronic', 'steptronic', 'eat6', 'eat8', 'at6',
                            'at8', 'edct', 'multitronic')):
        return 'AT'
    if any(x in t for x in ('manual', ' mec', 'mecan', 'bvm', ' mt ',
                            'mt5', 'mt6')):
        return 'MT'
    # 'at' / 'mt' como token suelto al final del texto
    import re as _r
    if _r.search(r'\bat\b', t):
        return 'AT'
    if _r.search(r'\bmt\b', t):
        return 'MT'
    return ''


def firma_version(version_str):
    """
    Firma de MOTOR desde el texto libre del vendedor: cilindrada + familia
    de motor + tracción. Inmune al orden de palabras, guiones y truncados.
      '1.6 gt hybrid4 4x4 300 e'   -> '1.6|hybrid4|4x4'
      '1.6 gt hybrid4 300 4x4 e'   -> '1.6|hybrid4|4x4'   (mismo grupo)
      '1.5 allure pack bluehdi'    -> '1.5|bluehdi|'
      'blue hdi 130 eat8 1.5 au'   -> '1.5|bluehdi|'      (mismo grupo)
    Devuelve None si no hay cilindrada + familia (grupo poco confiable).
    """
    if pd.isna(version_str) or not str(version_str).strip():
        return None
    s = str(version_str).lower()
    cil = extraer_motor(s)
    aplastado = re.sub(r'[\s\-\.]', '', s)
    familia = None
    for canon, variantes in FAMILIAS_MOTOR:
        if any(vv in aplastado for vv in variantes):
            familia = canon
            break
    if not cil or not familia:
        return None
    if re.search(r'4\s*x\s*4|awd|4wd', s):
        trac = '4x4'
    elif re.search(r'4\s*x\s*2|fwd|2wd', s):
        trac = '4x2'
    else:
        trac = ''
    # Línea GT / GT-Line: versión especial (equipamiento y a veces motor
    # superior). Se marca en la firma para no mezclarla con las versiones
    # normales, que valen bastante menos. "gt" como palabra suelta o pegada
    # (gtline, gt-line), evitando falsos positivos dentro de otras palabras.
    linea = 'gt' if re.search(r'\bgt\b|gt\s*-?\s*line|gtline', s) else ''
    return f"{cil}|{familia}|{trac}|{linea}"


# ---------------------------------------------------------------------------
# 1. CARGA Y LIMPIEZA
# ---------------------------------------------------------------------------
def cargar_captura(path):
    """Lee una captura .xlsx tal cual (la limpieza la hace limpiar())."""
    return pd.read_excel(path)


def fecha_de_archivo(path):
    """Fecha AAAA-MM-DD desde el nombre del archivo (base_2026-06-27.xlsx)."""
    mm = _re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(str(path)))
    if mm:
        try:
            return pd.to_datetime(mm.group(1))
        except Exception:
            pass
    return None


def fecha_de_captura(df):
    """Devuelve la fecha dominante de la captura como datetime (día)."""
    if 'fecha' not in df.columns:
        return pd.Timestamp.now().normalize()
    col = df['fecha'].astype(str)
    # Formato ISO del estado incremental: "2026-07-08" (año-mes-día). Se prueba
    # PRIMERO para no confundir 2026-07-08 con el 8 de agosto (dayfirst lo daba
    # vuelta). Luego el formato con hora del scraper, y por último algo flexible.
    s = pd.to_datetime(col, format='%Y-%m-%d', errors='coerce')
    if s.isna().all():
        s = pd.to_datetime(col, format='%d-%m-%Y %H:%M', errors='coerce')
    if s.isna().all():
        s = pd.to_datetime(col, format='%d-%m-%Y', errors='coerce')
    if s.isna().all():
        # último recurso: ISO con hora, sin dayfirst (evita invertir mes/día)
        s = pd.to_datetime(col, errors='coerce')
    if s.isna().all():
        return pd.Timestamp.now().normalize()
    return s.dt.normalize().mode().iloc[0]


def cargar_desde_incremental(estado_csv="estado_avisos.csv",
                             specs_csv="especificaciones.csv"):
    """
    Arma la tabla de trabajo desde los archivos del flujo incremental:
      - estado_avisos.csv: avisos activos con catálogo + precio + teléfono
      - especificaciones.csv: versión oficial limpia + specs (si ya se bajaron)
    Devuelve un DataFrame con las columnas que espera el motor, más
    'version_oficial' cuando hay specs. Solo incluye avisos activos (activo=1).
    """
    df = pd.read_csv(estado_csv)
    # solo activos
    if 'activo' in df.columns:
        df = df[df['activo'].astype(str) != '0'].copy()

    # unir specs por id (si existe el archivo)
    if os.path.exists(specs_csv):
        specs = pd.read_csv(specs_csv)
        # quedarse con la última fila por id (por si un id se bajó dos veces)
        specs = specs.drop_duplicates(subset='id', keep='last')
        cols_specs = [c for c in specs.columns if c not in ('url',)]
        df = df.merge(specs[cols_specs], on='id', how='left', suffixes=('', '_spec'))

    # mapear a lo que espera el motor
    df['fecha'] = df.get('ultima_vez', '')
    df['Enlace Web'] = df.get('url', '')
    df['nombre_whatsapp'] = ''
    for col in ['telefono', 'whatsapp', 'moneda', 'traccion', 'condicion',
                'transmision', 'combustible', 'entidad']:
        if col not in df.columns:
            df[col] = ''
    if 'version_oficial' not in df.columns:
        df['version_oficial'] = ''

    return df



    df = pd.read_excel(path)
    faltan = [c for c in COLUMNAS_ESPERADAS if c not in df.columns]
    if faltan:
        raise ValueError(f"{os.path.basename(path)}: faltan columnas {faltan}")

    urls = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        hdr = [c.value for c in ws[1]]
        ci_link = hdr.index('Enlace Web') + 1
        ci_id = hdr.index('id') + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=ci_link)
            idv = ws.cell(row=r, column=ci_id).value
            if cell.hyperlink and cell.hyperlink.target:
                urls[idv] = cell.hyperlink.target
        wb.close()
    except Exception as e:
        log(f"  aviso: no se pudieron leer hipervinculos ({e})")

    df['url'] = df['id'].map(urls)
    return df


def limpiar(df):
    """Aplica las reglas de limpieza validadas. Devuelve df limpio."""
    df = df.drop_duplicates(subset='id', keep='first').copy()
    df = df[df['moneda'] == 'CLP']
    df['año'] = pd.to_numeric(df['año'], errors='coerce')
    df = df[(df['año'] >= ANIO_MIN) & (df['año'] <= ANIO_MAX)]
    df['precio'] = pd.to_numeric(df['precio'], errors='coerce')
    df['kilometraje'] = pd.to_numeric(df['kilometraje'], errors='coerce')
    df = df[(df['precio'] >= PRECIO_MIN) & (df['precio'] <= PRECIO_MAX)]
    df = df[(df['kilometraje'] >= KM_MIN) & (df['kilometraje'] <= KM_MAX)]
    df['marca'] = df['marca'].astype(str).str.strip().str.lower()
    df['modelo'] = df['modelo'].fillna('').astype(str).str.strip().str.lower()
    anio_actual = ANIO_MAX
    df['antiguedad'] = (anio_actual - df['año']).clip(lower=0)
    df['km_anual'] = np.where(df['antiguedad'] > 0,
                              df['kilometraje'] / df['antiguedad'],
                              df['kilometraje'])
    # Extraer cilindrada del campo version (mejora el agrupamiento de precio justo)
    df['motor'] = df['version'].apply(extraer_motor)
    # Normalizar transmisión a AT/MT. La caja cambia el precio ~1M o más,
    # así que se usa para no comparar un automático con un manual. Se toma
    # del campo transmision del estado y, si falta, se infiere del texto.
    df['caja'] = df.apply(lambda r: _norm_caja(r.get('transmision'), r.get('version')), axis=1)
    df['firma_ver'] = df['version'].apply(firma_version)

    # --- Correcciones de versiones por modelo (herramienta "Arreglar versiones")
    # Reconstruye la clave cil|comb|caja|acabado de cada aviso y, si el usuario
    # la renombró/fusionó/excluyó para ese modelo, aplica el cambio.
    if os.path.exists('correcciones_versiones.json'):
        try:
            with open('correcciones_versiones.json', encoding='utf-8') as f:
                corr_ver = json.load(f)
        except Exception as e:
            corr_ver = {}
            log(f"  (correcciones_versiones.json ilegible: {e})")
        if corr_ver:
            df = _aplicar_correcciones_versiones(df, corr_ver)
    tabla_gen = cargar_generaciones()
    df = asignar_generacion(df, tabla_gen)

    # --- AÑOS DE TRANSICION: si un año cae en DOS generaciones del mismo
    # modelo (ej: la gen II llega a 2020 y la III arranca en 2020), la
    # asignacion por año es ambigua -> se marca para revision por FOTOS.
    trans_map = {}
    for (ma, mo), rangos in tabla_gen.items():
        if len(rangos) < 2:
            continue
        # a) años que caen dentro de DOS rangos (solape declarado)
        for y in range(min(d for d, _, _ in rangos),
                       max(min(h, 2030) for _, h, _ in rangos) + 1):
            gens_y = sorted({g for d, h, g in rangos if d <= y <= h})
            if len(gens_y) >= 2:
                trans_map[(ma, mo, y)] = '/'.join(gens_y)
        # b) años FRONTERA entre generaciones consecutivas: el año de cambio
        # (y el siguiente) conviven ambas generaciones en el mercado real,
        # aunque la tabla no se solape (gen II hasta 2019 / gen III desde 2020)
        orden = sorted(rangos)
        for (d1, h1, g1), (d2, h2, g2) in zip(orden, orden[1:]):
            if g1 != g2 and 0 <= d2 - h1 <= 1 and h1 < 9999:
                par = '/'.join(sorted({g1, g2}))
                for y in (h1, d2):
                    trans_map.setdefault((ma, mo, y), par)
    if trans_map:
        _man = df['marca'].map(_norm_txt)
        _mon = df['modelo'].map(_norm_txt)
        df['gens_pos'] = [trans_map.get((a, b, int(y)), '') if pd.notna(y) else ''
                          for a, b, y in zip(_man, _mon, df['año'])]
    else:
        df['gens_pos'] = ''

    # --- CORRECCIONES POR AVISO (correcciones_avisos.json): lo que TU OJO
    # confirmo mirando las fotos manda sobre todo lo automatico.
    #   {id: {"gen": "III"}}      -> fija la generacion de ESE aviso
    #   {id: {"excluir": true}}   -> lo saca del analisis (comparables y medianas)
    corr_av = {}
    if os.path.exists('correcciones_avisos.json'):
        try:
            with open('correcciones_avisos.json', encoding='utf-8') as f:
                corr_av = {str(k).lower(): v for k, v in json.load(f).items()}
        except Exception as e:
            log(f"  (correcciones_avisos.json ilegible: {e})")
    if corr_av:
        ids_low = df['id'].astype(str).str.lower()
        gen_fix = ids_low.map(lambda i: (corr_av.get(i) or {}).get('gen') or None)
        n_gen = int(gen_fix.notna().sum())
        if n_gen:
            df.loc[gen_fix.notna(), 'gen'] = gen_fix[gen_fix.notna()]
            df.loc[gen_fix.notna(), 'gens_pos'] = ''   # ya no es ambiguo
        ver_fix = ids_low.map(lambda i: (corr_av.get(i) or {}).get('version') or None)
        n_ver = int(ver_fix.notna().sum())
        if n_ver:
            df.loc[ver_fix.notna(), 'version'] = ver_fix[ver_fix.notna()]
            # la firma se recalcula sobre la versión corregida
            df.loc[ver_fix.notna(), 'firma_ver'] = \
                df.loc[ver_fix.notna(), 'version'].apply(firma_version)
        excl_av = ids_low.map(lambda i: bool((corr_av.get(i) or {}).get('excluir')))
        n_exc = int(excl_av.sum())
        if n_exc:
            df = df[~excl_av].copy()
        log(f"Correcciones por aviso (fotos): {n_gen} generaciones fijadas, "
            f"{n_ver} versiones corregidas, {n_exc} avisos descartados.")

    # correcciones del depurador: la versión canónica confirmada reemplaza a la
    # firma automática como llave de agrupación; 'excluir' anula firmas basura
    corr = cargar_correcciones()
    if corr:
        claves = (df['marca'].map(_norm_txt) + '|' + df['modelo'].map(_norm_txt)
                  + '|' + df['gen'].astype(str) + '|' + df['firma_ver'].fillna(''))
        canon = claves.map(lambda k: (corr.get(k) or {}).get('canonica') or None)
        excl = claves.map(lambda k: bool((corr.get(k) or {}).get('excluir')))
        df.loc[canon.notna(), 'firma_ver'] = canon[canon.notna()]
        df.loc[excl, 'firma_ver'] = None
        log(f"  Depurador: {int(canon.notna().sum())} avisos con versión canónica, "
            f"{int(excl.sum())} excluidos de agrupación fina.")
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2. PRECIO JUSTO — cascada de 3 pasadas por especificidad de grupo
# ---------------------------------------------------------------------------
# Pasada 1 (más específica): marca + modelo + motor + combustible, ±3 años
#   → separa, p.ej., Peugeot 3008 1.2 bencina vs 1.5 diesel vs 2.0 diesel
# Pasada 2 (media):          marca + modelo + combustible, ±3 años
#   → para los sin motor extraído o con pocos comparables en pasada 1
# Pasada 3 (fallback):       marca + modelo, ±3 años
#   → comportamiento original, para los que aún no tienen precio
# ---------------------------------------------------------------------------
def calcular_precio_justo(df):
    fair   = np.full(len(df), np.nan)
    ncomp  = np.zeros(len(df), dtype=int)
    metodo = np.array([''] * len(df), dtype=object)

    def _calc(yrs_pool, kms_pool, prs_pool, idxs_t, yrs_t, kms_t, sufijo):
        """Calcula precio justo para cada target usando el pool como comparables."""
        for y, km, idx in zip(yrs_t, kms_t, idxs_t):
            if not np.isnan(fair[idx]):
                continue  # ya asignado en pasada anterior
            mask = np.abs(yrs_pool - y) <= 3
            n = int(mask.sum())
            if n < MIN_COMPARABLES:
                continue
            pp, kk = prs_pool[mask], kms_pool[mask]
            ncomp[idx] = n
            p10, p90 = np.percentile(pp, [10, 90])
            pmed = float(np.median(pp))
            asignado = False
            if n >= 10:
                lo, hi = np.percentile(pp, [5, 95])
                sel = (pp >= lo) & (pp <= hi)
                if sel.sum() >= 8 and np.ptp(kk[sel]) > 0:
                    b, a = np.polyfit(kk[sel], pp[sel], 1)
                    if b < 0:
                        fair[idx] = float(np.clip(a + b * km, p10, p90))
                        metodo[idx] = f'km_adj_{sufijo}'
                        asignado = True
            if not asignado:
                fair[idx] = pmed
                metodo[idx] = f'mediana_{sufijo}'

    # --- Pasada 0: (marca, modelo, version_oficial) — la MÁS precisa ---
    # Solo corre si hay specs bajadas (columna 'version_oficial' con datos).
    # Es la mejor agrupación: misma versión canónica de fábrica = mismo auto.
    if 'version_oficial' in df.columns:
        df_v = df[df['version_oficial'].notna() &
                  (df['version_oficial'].astype(str).str.strip() != '')]
        if not df_v.empty:
            for (ma, mo, vof), g in df_v.groupby(
                    ['marca', 'modelo', 'version_oficial']):
                yrs  = g['año'].values
                kms  = g['kilometraje'].values
                prs  = g['precio'].values
                idxs = g.index.values
                _calc(yrs, kms, prs, idxs, yrs, kms, 'vof')

    # --- Pasada 0.5: (marca, modelo, FIRMA de motor) ---
    # Firma = cilindrada + familia + tracción desde el texto sucio del
    # vendedor. Separa el 1.6 thp del 1.6 puretech del 1.6 hybrid4 aunque
    # las palabras vengan desordenadas o con guiones.
    if 'firma_ver' in df.columns:
        df_f = df[df['firma_ver'].notna()]
        if not df_f.empty:
            for (ma, mo, fv, gn), g in df_f.groupby(['marca', 'modelo', 'firma_ver', 'gen']):
                idxs_pend = g.index[np.isnan(fair[g.index])].values
                if len(idxs_pend) == 0:
                    continue
                _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                      idxs_pend,
                      df.loc[idxs_pend, 'año'].values,
                      df.loc[idxs_pend, 'kilometraje'].values, 'firma')

    # --- Pasada 1: (marca, modelo, motor, combustible, caja) ---
    # Incluye la transmisión: un automático no se compara con un manual
    # (diferencia típica ~1M+). Los de caja desconocida ('') forman su
    # propio grupo, que es lo correcto (no se mezclan con ninguna).
    df_m = df[df['motor'].notna()]
    if not df_m.empty:
        for (ma, mo, mot, comb, cj, gn), g in df_m.groupby(
                ['marca', 'modelo', 'motor', 'combustible', 'caja', 'gen']):
            idxs_pend = g.index[np.isnan(fair[g.index])].values
            if len(idxs_pend) == 0:
                continue
            _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                  idxs_pend,
                  df.loc[idxs_pend, 'año'].values,
                  df.loc[idxs_pend, 'kilometraje'].values, 'motor')

    # --- Pasada 2a: (marca, modelo, combustible, CILINDRADA) ---
    # Cuando el aviso tiene cilindrada (motor), no se compara un 1.2 con un
    # 1.6: se agrupa también por cilindrada. Evita mezclar versiones de
    # distinto valor dentro del mismo combustible/generación.
    df_c = df[df['motor'].notna()]
    if not df_c.empty:
        for (ma, mo, comb, mot, cj, gn), g in df_c.groupby(
                ['marca', 'modelo', 'combustible', 'motor', 'caja', 'gen']):
            idxs_pend = g.index[np.isnan(fair[g.index])].values
            if len(idxs_pend) == 0:
                continue
            _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                  idxs_pend,
                  df.loc[idxs_pend, 'año'].values,
                  df.loc[idxs_pend, 'kilometraje'].values, 'comb_cil')

    # --- Pasada 2b: (marca, modelo, combustible, caja) ---
    # Para los sin cilindrada pero CON caja conocida: separa AT de MT.
    df_cj = df[df['caja'] != '']
    if not df_cj.empty:
        for (ma, mo, comb, cj, gn), g in df_cj.groupby(
                ['marca', 'modelo', 'combustible', 'caja', 'gen']):
            idxs_pend = g.index[np.isnan(fair[g.index])].values
            if len(idxs_pend) == 0:
                continue
            _calc(g['año'].values, g['kilometraje'].values, g['precio'].values,
                  idxs_pend,
                  df.loc[idxs_pend, 'año'].values,
                  df.loc[idxs_pend, 'kilometraje'].values, 'comb')

    # --- Pasada 2c: (marca, modelo, combustible) ---
    # Último recurso por combustible, para los que no tienen ni cilindrada
    # ni caja conocidas. Mejor una tasación aproximada que ninguna.
    for (ma, mo, comb, gn), g in df.groupby(['marca', 'modelo', 'combustible', 'gen']):
        idxs_pend = g.index[np.isnan(fair[g.index])].values
        if len(idxs_pend) == 0:
            continue
        _calc(
            g['año'].values, g['kilometraje'].values, g['precio'].values,
            idxs_pend,
            df.loc[idxs_pend, 'año'].values,
            df.loc[idxs_pend, 'kilometraje'].values,
            'comb'
        )

    # --- Pasada 3: fallback (marca, modelo) — comportamiento original ---
    for (ma, mo), g in df.groupby(['marca', 'modelo']):
        idxs_pend = g.index[np.isnan(fair[g.index])].values
        if len(idxs_pend) == 0:
            continue
        _calc(
            g['año'].values, g['kilometraje'].values, g['precio'].values,
            idxs_pend,
            df.loc[idxs_pend, 'año'].values,
            df.loc[idxs_pend, 'kilometraje'].values,
            'modelo'
        )

    df['precio_justo'] = fair
    df['comparables']  = ncomp
    df['metodo']       = metodo
    valido = df['precio_justo'].notna()
    df['descuento_pct'] = np.where(
        valido, (df['precio_justo'] - df['precio']) / df['precio_justo'] * 100, np.nan)
    df['op_precio'] = (df['descuento_pct'] >= DESCUENTO_MIN) & valido
    df['op_km']     = (df['km_anual'] < KM_ANUAL_MAX) & (df['antiguedad'] >= 1)
    df['op_oro']    = df['op_precio'] & df['op_km']
    return df


# ---------------------------------------------------------------------------
# 3. ROTACION (comparando capturas consecutivas)
# ---------------------------------------------------------------------------
def calcular_rotacion(capturas):
    if len(capturas) < 2:
        return None

    # Ordenar por fecha y elegir el par CONSECUTIVO más cercano en el tiempo.
    # Comparar los extremos (junio vs hoy) daba ~100% de "salida" porque son
    # fotos muy separadas y de orígenes distintos. Dos fotos cercanas dan una
    # tasa de rotación creíble. Si el par más cercano está a más de 21 días,
    # la señal es débil y se marca como poco confiable.
    caps = sorted(capturas, key=lambda c: c['fecha'])
    mejor = None
    for a, b in zip(caps, caps[1:]):
        d = (b['fecha'] - a['fecha']).days
        if d <= 0:
            continue
        if mejor is None or d < mejor[2]:
            mejor = (a, b, d)
    if mejor is None:
        return None
    prim, ult, dias = mejor
    ventana_confiable = dias <= 21

    ids_prim = prim['ids']
    ids_ult = ult['ids']
    desaparecidos = ids_prim - ids_ult
    nuevos = ids_ult - ids_prim
    sobreviven = ids_prim & ids_ult

    dprim = prim['df'].set_index('id')

    glob = {
        'fecha_desde': prim['fecha'].strftime('%Y-%m-%d'),
        'fecha_hasta': ult['fecha'].strftime('%Y-%m-%d'),
        'dias': int(dias),
        'avisos_inicio': len(ids_prim),
        'avisos_fin': len(ids_ult),
        'desaparecidos': len(desaparecidos),
        'nuevos': len(nuevos),
        'tasa_salida_periodo': round(len(desaparecidos) / len(ids_prim) * 100, 1),
        'tasa_salida_diaria': round(len(desaparecidos) / len(ids_prim) / dias * 100, 2),
        'ventana_confiable': bool(ventana_confiable),
        'n_capturas': len(caps),
    }

    df_des = dprim.loc[list(desaparecidos)]
    base_modelo = dprim.groupby(['marca', 'modelo']).size()
    fueron_modelo = df_des.groupby(['marca', 'modelo']).size()
    rot = pd.DataFrame({'inicio': base_modelo, 'salieron': fueron_modelo}).fillna(0)
    rot['salieron'] = rot['salieron'].astype(int)
    rot = rot[rot['inicio'] >= 1]
    rot['tasa_salida'] = (rot['salieron'] / rot['inicio'] * 100).round(1)
    rot['inventario'] = np.where(rot['inicio'] >= 10, 'alto', 'bajo')
    rot['dias_venta_est'] = np.where(
        rot['salieron'] > 0,
        (dias / (rot['salieron'] / rot['inicio'])).round(0),
        np.nan)

    def cap(s):
        return ' '.join(w.capitalize() for w in str(s).split())

    por_modelo = []
    for (ma, mo), r in rot.iterrows():
        por_modelo.append({
            'marca': cap(ma), 'modelo': cap(mo),
            'stock': int(r['inicio']),
            'salieron': int(r['salieron']),
            'tasa_salida': float(r['tasa_salida']),
            'dias_venta_est': None if np.isnan(r['dias_venta_est']) else int(r['dias_venta_est']),
            'inventario': r['inventario'],
            'confiable': bool(r['inventario'] == 'alto'),
        })
    por_modelo.sort(key=lambda x: x['tasa_salida'], reverse=True)

    dult = ult['df'].set_index('id')
    comun = list(sobreviven)
    p0 = dprim.loc[comun, 'precio']
    p1 = dult.loc[comun, 'precio']
    cambio = (p0 != p1)
    bajaron = (p1 < p0)
    glob['con_cambio_precio'] = int(cambio.sum())
    glob['bajaron_precio'] = int((cambio & bajaron).sum())
    glob['subieron_precio'] = int((cambio & ~bajaron).sum())

    return {'global': glob, 'por_modelo': por_modelo}


# ---------------------------------------------------------------------------
# ANÁLISIS DE MERCADO
# ---------------------------------------------------------------------------
def construir_depurador_modelos(dfa):
    """Para la herramienta 'Arreglar BBDD por modelo'. Por cada modelo con
    volumen, propone sus VERSIONES REALES agrupando por:
       cilindrada | combustible | caja | línea(GT) | acabado
    El acabado (active/allure/gt/etc.) se detecta del texto. Cada grupo trae
    sus avisos (año, km, precio, url), precio por año y una etiqueta propuesta
    legible. El usuario confirma/corrige/fusiona en el dashboard.
    Devuelve {clave_modelo: {marca, modelo, versiones:[...], sin_clasificar:n}}.
    """
    ACABADOS = ['gt line', 'gt-line', 'gtline', 'gt', 'allure', 'active',
                'premier', 'premiere', 'feel', 'style', 'access', 'business',
                'signature', 'sport', 'limited', 'ltz', 'lt', 'ls', 'exclusive',
                'dynamique', 'zen', 'intens', 'life', 'trend', 'comfort',
                'titanium', 'ambiente', 'highline', 'comfortline', 'trendline',
                'xei', 'xli', 'gli', 'gls', 'glx', 'gl', 'ex', 'exl', 'lx',
                'touring', 'value', 'full', 'sense']

    def acabado_de(txt):
        t = ' ' + str(txt).lower() + ' '
        for ac in ACABADOS:                     # orden: gt line antes que gt
            if ' ' + ac + ' ' in t or t.strip().endswith(' ' + ac):
                return ac
        return ''

    def etiqueta(cil, comb, caja, acab):
        partes = []
        if cil:
            partes.append(str(cil))
        if acab:
            partes.append(acab.upper() if len(acab) <= 3 else acab.title())
        cb = {'diesel': 'Diésel', 'bencina': 'Bencina', 'híbrido': 'Híbrido',
              'hibrido': 'Híbrido'}.get(str(comb).lower(), str(comb).title())
        if cb:
            partes.append(cb)
        if caja:
            partes.append(caja)
        return ' '.join(partes) if partes else '(sin datos)'

    out = {}
    vol = dfa.groupby(['marca', 'modelo'])['precio'].transform('size')
    grande = dfa[vol >= 12].copy()
    if grande.empty:
        return out
    grande['_acab'] = grande['version'].apply(acabado_de)
    tiene_gt = grande['firma_ver'].fillna('').str.contains('gt')

    def cap(s):
        return ' '.join(w.capitalize() for w in str(s).split())

    for (ma, mo), g in grande.groupby(['marca', 'modelo']):
        versiones = []
        sin_clasif = 0
        # clave de versión: cilindrada|combustible|caja|acabado (+gt de firma)
        g = g.copy()
        g['_cil'] = g['motor'].fillna('')
        g['_cj'] = g['caja'].fillna('')
        g['_key'] = (g['_cil'].astype(str) + '|' + g['combustible'].astype(str)
                     + '|' + g['_cj'].astype(str) + '|' + g['_acab'].astype(str))
        # si el usuario ya fusionó/renombró, agrupar por el nombre canónico
        if '_ver_canon' in g.columns:
            g['_key'] = g.apply(lambda r: ('canon::' + str(r['_ver_canon']))
                                if str(r.get('_ver_canon') or '') else r['_key'], axis=1)
        for key, gv in g.groupby('_key'):
            if str(key).startswith('canon::'):
                nombre_canon = key.split('canon::', 1)[1]
                cil = comb = caja = acab = ''
            else:
                nombre_canon = ''
                cil, comb, caja, acab = (key.split('|') + ['', '', '', ''])[:4]
            if not cil and not acab:
                sin_clasif += len(gv)
                # igual lo mostramos como grupo "sin identificar"
            avisos = []
            for _, r in gv.sort_values('año', ascending=False).head(60).iterrows():
                avisos.append({
                    'año': int(r['año']) if pd.notna(r['año']) else None,
                    'km': int(r['kilometraje']) if pd.notna(r['kilometraje']) else None,
                    'precio': int(r['precio']) if pd.notna(r['precio']) else None,
                    'ent': 'P' if r['entidad'] == 'Particular' else 'A',
                    'texto': str(r['version'])[:50],
                    'url': r['url'] if pd.notna(r['url']) else '',
                    'id': str(r['id']).lower(),
                })
            precio_año = {}
            for a2, gg in gv.groupby('año'):
                if pd.notna(a2):
                    precio_año[int(a2)] = int(gg['precio'].median())
            versiones.append({
                'clave': key,
                'etiqueta': nombre_canon or etiqueta(cil, comb, caja, acab),
                'cil': cil, 'combustible': comb, 'caja': caja, 'acabado': acab,
                'n': int(len(gv)),
                'precio_med': int(gv['precio'].median()),
                'años': [int(gv['año'].min()), int(gv['año'].max())] if gv['año'].notna().any() else None,
                'precio_año': dict(sorted(precio_año.items())),
                'identificada': bool(cil or acab),
                'avisos': avisos,
            })
        versiones.sort(key=lambda x: -x['n'])
        out[f"{ma}|{mo}"] = {
            'marca': cap(ma), 'modelo': cap(mo),
            'total': int(len(g)),
            'n_versiones': len(versiones),
            'sin_clasificar': int(sin_clasif),
            'versiones': versiones,
        }
    return out


def construir_analisis(dfa):
    """Análisis tipo analista de mercado sobre la base activa. Devuelve dict
    con: resumen por marca, depreciación/retención, señales de alerta
    (modelos muy nuevos sobre-publicados) y oportunidad por segmento."""
    ANIO = ANIO_MAX
    d = dfa[(dfa['precio'] > 0) & dfa['año'].notna()].copy()
    d['año'] = d['año'].astype(int)

    def cap(s):
        return ' '.join(w.capitalize() for w in str(s).split())

    # --- 1. RESUMEN POR MARCA (volumen, precio medio, antigüedad típica) ---
    por_marca = []
    for ma, g in d.groupby('marca'):
        if len(g) < 20:
            continue
        antig = (ANIO - g['año']).clip(lower=0)
        por_marca.append({
            'marca': cap(ma),
            'avisos': int(len(g)),
            'precio_med': int(g['precio'].median()),
            'antiguedad_med': round(float(antig.median()), 1),
            'km_med': int(g['kilometraje'].median()) if g['kilometraje'].notna().any() else 0,
            'pct_particular': round(float((g['entidad'] == 'Particular').mean() * 100), 0),
        })
    por_marca.sort(key=lambda x: -x['avisos'])

    # --- 2. DEPRECIACIÓN / RETENCIÓN DE VALOR POR MARCA ---
    # Cuánto pierde el valor entre 0 y 3 años (mediana de precio por año).
    deprec = []
    for ma, g in d.groupby('marca'):
        if len(g) < 40:
            continue
        p0 = g[g['año'] >= ANIO - 1]['precio'].median()
        p3 = g[g['año'].between(ANIO - 4, ANIO - 3)]['precio'].median()
        if pd.notna(p0) and pd.notna(p3) and p0 > 0:
            perdida = round((1 - p3 / p0) * 100, 0)
            deprec.append({
                'marca': cap(ma), 'avisos': int(len(g)),
                'perdida_3a': perdida,
                'retiene_3a': round(100 - perdida, 0),
                'precio_0a': int(p0), 'precio_3a': int(p3),
            })
    deprec.sort(key=lambda x: x['perdida_3a'], reverse=True)

    # --- 3. SEÑALES DE ALERTA: modelos con demasiados avisos muy nuevos ---
    # Alto % de 0-1 año publicados = mucha gente revende casi nuevo
    # (posible problema del auto, arrepentimiento, o flota/rent-a-car).
    alertas = []
    for (ma, mo), g in d.groupby(['marca', 'modelo']):
        if len(g) < 15:
            continue
        nuevos = int((g['año'] >= ANIO - 1).sum())
        pct = nuevos / len(g) * 100
        if pct >= 45 and nuevos >= 8:
            alertas.append({
                'marca': cap(ma), 'modelo': cap(mo),
                'avisos': int(len(g)), 'nuevos_0_1a': nuevos,
                'pct_nuevos': round(pct, 0),
                'precio_med': int(g['precio'].median()),
            })
    alertas.sort(key=lambda x: (-x['pct_nuevos'], -x['avisos']))

    # --- 4. OPORTUNIDAD POR SEGMENTO ---
    # Segmento = derivado de la carrocería si hay ficha; si no, heurística
    # por modelo. Medimos % de avisos que son oportunidad (oro) por segmento.
    seg = []
    if 'op_oro' in d.columns:
        d['_seg'] = d.apply(_segmento_de, axis=1)
        for s, g in d.groupby('_seg'):
            if s == '?' or len(g) < 30:
                continue
            seg.append({
                'segmento': s, 'avisos': int(len(g)),
                'precio_med': int(g['precio'].median()),
                'pct_oro': round(float(g['op_oro'].mean() * 100), 1),
                'oro': int(g['op_oro'].sum()),
            })
        seg.sort(key=lambda x: -x['pct_oro'])

    return {
        'por_marca': por_marca,
        'depreciacion': deprec,
        'alertas': alertas,
        'segmentos': seg,
        'anio_ref': ANIO,
    }


_SEG_MAP = [
    ('pickup', ['hilux', 'ranger', 'amarok', 'l200', 'dmax', 'd-max', 'frontier',
                'navara', 'np300', 'colorado', 's10', 'gladiator', 'maverick',
                'sail', 'partner', 'kangoo', 'berlingo', 'saveiro']),
    ('suv',    ['tucson', 'sportage', 'santa fe', 'kuga', 'cr-v', 'crv', 'rav4',
                'tiguan', 'q3', 'q5', 'x1', 'x3', 'captur', 'duster', 'tracker',
                'creta', 'kicks', 'seltos', 'cx-5', 'cx5', 'forester', 'outlander',
                'ecosport', 'territory', 'omoda', 'jaecoo', 'tiggo']),
    ('city',   ['sail', 'yaris', 'swift', 'rio', 'accent', 'onix', 'clio', 'sandero',
                'kwid', 'picanto', 'morning', 'spark', 'march', 'versa', 'city',
                'polo', 'gol', 'mobi', 'k3', 'i10', 'i20']),
]


def _segmento_de(r):
    ficha = None
    mo = str(r.get('modelo', '')).lower()
    for seg, modelos in _SEG_MAP:
        if any(x in mo for x in modelos):
            return seg
    return '?'


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    # --- Fuente de datos: incremental (estado_avisos.csv) si existe;
    #     si no, las capturas .xlsx de siempre. ---
    def elegir_capturas_xlsx():
        """Un archivo por FECHA: si conviven 'Chileautos 2026-07-08 AM.xlsx' y
        '... PM.xlsx' (o cualquier otro sufijo), para el análisis por días se
        usa el más reciente de esa fecha (por hora de modificación). Los demás
        quedan en la carpeta como respaldo, sin molestar."""
        grupos = {}
        for path in glob.glob(os.path.join(CARPETA_CAPTURAS, "*.xlsx")):
            mfecha = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
            clave = mfecha.group(1) if mfecha else os.path.basename(path)
            if clave not in grupos or os.path.getmtime(path) > os.path.getmtime(grupos[clave]):
                grupos[clave] = path
        return [grupos[k] for k in sorted(grupos)]

    usa_incremental = os.path.exists("estado_avisos.csv")

    capturas = []
    if usa_incremental:
        log("Usando datos INCREMENTALES (estado_avisos.csv).")
        df = cargar_desde_incremental()
        dfl = limpiar(df)
        fecha = pd.Timestamp.today().normalize()
        try:
            fecha = fecha_de_captura(df)
        except Exception:
            pass
        capturas.append({'fecha': fecha, 'df': dfl, 'ids': set(dfl['id'])})
        con_specs = (dfl.get('version_oficial', pd.Series(dtype=str)).astype(str).str.strip() != '').sum()
        log(f"  {len(df)} avisos activos, {len(dfl)} limpios, {con_specs} con versión oficial (specs).")
        # Si además hay capturas .xlsx viejas, se suman para poder medir rotación.
        if os.path.isdir(CARPETA_CAPTURAS):
            for path in elegir_capturas_xlsx():
                try:
                    dfx = cargar_captura(path)
                    f_cap = fecha_de_archivo(path) or fecha_de_captura(dfx)
                    if f_cap == fecha:
                        # hoy ya está representado por el estado (más completo)
                        continue
                    capturas.append({'fecha': f_cap,
                                     'df': limpiar(dfx), 'ids': set(limpiar(dfx)['id'])})
                except Exception as e:
                    log(f"  (salteo {os.path.basename(path)}: {e})")
    else:
        if not os.path.isdir(CARPETA_CAPTURAS):
            sys.exit(f"No existe la carpeta '{CARPETA_CAPTURAS}' ni estado_avisos.csv. "
                     f"Corré primero la descarga de datos.")
        archivos = elegir_capturas_xlsx()
        if not archivos:
            sys.exit(f"No hay archivos .xlsx en '{CARPETA_CAPTURAS}'.")
        log(f"Capturas encontradas: {len(archivos)}")
        for path in archivos:
            df = cargar_captura(path)
            fecha = fecha_de_archivo(path)
            if fecha is None:
                fecha = fecha_de_captura(df)
            dfl = limpiar(df)
            capturas.append({'fecha': fecha, 'df': dfl, 'ids': set(dfl['id'])})
            log(f"  {os.path.basename(path)} -> {fecha.date()} | "
                f"{len(df)} filas, {len(dfl)} limpias")

    capturas.sort(key=lambda c: c['fecha'])

    reciente = capturas[-1]
    log(f"Calculando precio justo sobre captura {reciente['fecha'].date()} ...")
    dfa = calcular_precio_justo(reciente['df'].copy())

    n_vof    = (dfa['metodo'].str.endswith('vof')).sum()
    n_motor  = (dfa['metodo'].str.endswith('motor')).sum()
    n_comb   = (dfa['metodo'].str.endswith('comb')).sum()
    n_modelo = (dfa['metodo'].str.endswith('modelo')).sum()
    n_sin    = (dfa['metodo'] == '').sum()
    log(f"  Pasada 0 (versión oficial): {n_vof} | Pasada 1 (motor+comb): {n_motor} | "
        f"Pasada 2 (comb): {n_comb} | Fallback (modelo): {n_modelo} | Sin tasar: {n_sin}")

    log("Calculando rotacion entre capturas ...")
    rotacion = calcular_rotacion(capturas)

    def cap2(s):
        return ' '.join(w.capitalize() for w in str(s).split())
    vol_modelo = dfa.groupby(['marca', 'modelo'])['precio'].size()

    # Cargar etiquetas de comentarios si existen (comentarios_clasificados.json).
    etiquetas_por_id = {}
    if os.path.exists('comentarios_clasificados.json'):
        try:
            with open('comentarios_clasificados.json', encoding='utf-8') as f:
                clasif = json.load(f)
            for vid, d in clasif.items():
                etiquetas_por_id[str(vid).lower()] = {
                    'etiquetas': d.get('etiquetas', []),
                    'comentario': d.get('comentario', ''),
                    'urgencia': d.get('urgencia', 0),
                    'resumen_ia': d.get('resumen_ia', ''),
                }
            log(f"Etiquetas de comentarios cargadas: {len(etiquetas_por_id)}")
        except Exception as e:
            log(f"  aviso: no se pudo leer comentarios_clasificados.json ({e})")
    else:
        log("Sin comentarios_clasificados.json todavia: las oportunidades van sin etiquetas.")

    # Fichas técnicas completas (especificaciones_full.json), si el goteo
    # de specs ya bajó algunas: se adjuntan a las oportunidades para verlas
    # con el botón 📋 del radar.
    fichas_full = {}
    if os.path.exists("especificaciones_full.json"):
        try:
            with open("especificaciones_full.json", encoding="utf-8") as f:
                for fid, fdata in json.load(f).items():
                    fichas_full[str(fid).lower()] = (fdata or {}).get("campos") or {}
            log(f"Fichas técnicas completas cargadas: {len(fichas_full)}")
        except Exception as e:
            log(f"  aviso: no se pudo leer especificaciones_full.json ({e})")

    BANDERAS_ROJAS = {'daniado', 'problema_legal', 'km_dudoso',
                       'precio_condicionado', 'precio_mas_iva'}

    # ---- ORO POR COMENTARIO -------------------------------------------
    # Señales de negociacion en el texto (venta urgente, precio conversable,
    # se aceptan ofertas) ELEVAN un aviso a oportunidad de oro, aunque su
    # descuento no llegue al umbral clasico. Condiciones: vendedor Particular
    # (en automotoras estas frases son marketing), descuento real >= 8% y
    # ninguna bandera roja en el comentario.
    SENALES_NEGOCIO = {'venta_urgente', 'precio_conversable'}
    ids_negociables = set()
    for _vid, _inf in etiquetas_por_id.items():
        _ets = set(_inf.get('etiquetas', []))
        if (_ets & SENALES_NEGOCIO) and not (_ets & BANDERAS_ROJAS):
            ids_negociables.add(_vid)
    if ids_negociables and 'op_oro' in dfa.columns:
        _mask_neg = (
            dfa['id'].astype(str).str.lower().isin(ids_negociables)
            & (dfa['entidad'].astype(str).str.strip().str.lower() == 'particular')
            & (dfa['descuento_pct'] >= 8)
            & dfa['precio_justo'].notna()
        )
        _nuevos_oro = int((_mask_neg & ~dfa['op_oro']).sum())
        dfa.loc[_mask_neg, 'op_oro'] = True
        log(f"Oro por comentario (particular + señal de negociacion, sin "
            f"banderas): {int(_mask_neg.sum())} avisos ({_nuevos_oro} elevados)")

    _hoy_ts = pd.Timestamp.now().normalize()

    def _fecha_iso(v):
        if v is None or pd.isna(v):
            return ''
        txt = str(v).strip()
        d = pd.to_datetime(txt[:10], format='%Y-%m-%d', errors='coerce')
        if pd.isna(d):
            d = pd.to_datetime(txt.split()[0] if txt else txt, dayfirst=True, errors='coerce')
        return d.strftime('%Y-%m-%d') if pd.notna(d) else '' 

    def _dias_pub(pv):
        if pv is None or pd.isna(pv):
            return None
        txt = str(pv).strip()
        d = pd.to_datetime(txt[:10], format='%Y-%m-%d', errors='coerce')
        if pd.isna(d):
            d = pd.to_datetime(txt.split()[0] if txt else txt,
                               dayfirst=True, errors='coerce')
        if pd.isna(d):
            return None
        dias = int((_hoy_ts - d.normalize()).days)
        return dias if dias >= 0 else 0

    def fila_op(r):
        vid = str(r['id']).lower()
        info = etiquetas_por_id.get(vid, {})
        etiquetas = list(info.get('etiquetas', []))
        ant = int(r['antiguedad'])
        km = int(r['kilometraje'])
        km_imposible = (ant >= 2 and km < 1000) or (ant >= 5 and km < 5000)
        if km_imposible and 'km_dudoso' not in etiquetas:
            etiquetas.append('km_dudoso')
        return {
            'id': vid,
            'marca': cap2(r['marca']), 'modelo': cap2(r['modelo']),
            'version': (str(r['version'])[:42] if pd.notna(r['version']) else ''),
            'version_original': (str(r['version']) if pd.notna(r['version']) else ''),
            'tasada_como': (str(r['firma_ver']) if ('firma_ver' in r.index and pd.notna(r['firma_ver'])) else ''),
            'version_oficial': (str(r['version_oficial'])
                                if ('version_oficial' in r.index and pd.notna(r['version_oficial'])
                                    and str(r['version_oficial']).strip()) else ''),
            'generacion': (str(r['gen']) if 'gen' in r.index else ''),
            'año': int(r['año']), 'km': int(r['kilometraje']), 'precio': int(r['precio']),
            'precio_justo': int(r['precio_justo']) if pd.notna(r['precio_justo']) else None,
            'descuento': round(float(r['descuento_pct']), 1) if pd.notna(r['descuento_pct']) else None,
            'km_anual': int(r['km_anual']),
            'entidad': r['entidad'], 'comparables': int(r['comparables']),
            'metodo': r['metodo'],
            'mercado': int(vol_modelo.get((r['marca'], r['modelo']), 0)),
            'url': r['url'] if pd.notna(r['url']) else '',
            'f_oro': bool(r['op_oro']), 'f_precio': bool(r['op_precio']), 'f_km': bool(r['op_km']),
            'etiquetas': etiquetas,
            'bandera_roja': bool(set(etiquetas) & BANDERAS_ROJAS),
            'tiene_comentario': bool(info.get('comentario')),
            'comentario': info.get('comentario', ''),
            'gen': str(r.get('gen', '') or ''),
            'gens_pos': str(r.get('gens_pos', '') or ''),
            'urgencia': int(info.get('urgencia', 0) or 0),
            'resumen_ia': info.get('resumen_ia', ''),
            'ficha': fichas_full.get(vid) or None,
            'primera_vez': _fecha_iso(r['primera_vez'] if 'primera_vez' in r.index else None),
            'ultima_vez': _fecha_iso(r['ultima_vez'] if 'ultima_vez' in r.index else None),
            'dias_publicado': _dias_pub(r['primera_vez'] if 'primera_vez' in r.index else None),
        }

    valido = dfa[dfa['comparables'] >= MIN_COMPARABLES]
    pool = pd.concat([
        valido[valido['op_oro']].sort_values('descuento_pct', ascending=False).head(200),
        valido[valido['op_precio']].sort_values('descuento_pct', ascending=False).head(200),
        valido[valido['op_km']].sort_values('km_anual').head(200),
    ]).drop_duplicates(subset='id')
    oportunidades = [fila_op(r) for _, r in pool.iterrows()]

    # Comparables: por cada oportunidad, los autos del mismo marca+modelo a +-3 años
    CAP_COMP = 40
    comps = {}
    pool_ids = set(pool['id'].astype(str).str.lower())
    # Detectar tracción y cabina del texto para comparar peras con peras
    def _trac(t):
        t = str(t).lower()
        if any(x in t for x in ('4x4', '4wd', 'awd')): return '4x4'
        if any(x in t for x in ('4x2', '4wd', 'rwd', 'fwd')): return '4x2'
        return ''
    def _cabina(t):
        t = str(t).lower()
        if any(x in t for x in ('reg cab', 'cab. sim', 'cabina simple', 'single')): return 'reg'
        if any(x in t for x in ('dob', 'doble', 'cc ', 'crew', 'dcab', 'd cab')): return 'dcab'
        return ''
    dfa = dfa.copy()
    dfa['_trac_c'] = dfa['version'].map(_trac)
    dfa['_cab_c'] = dfa['version'].map(_cabina)

    for _, r in pool.iterrows():
        vid = str(r['id']).lower()
        g = dfa[(dfa['marca'] == r['marca']) & (dfa['modelo'] == r['modelo'])]
        # Comparables de VERDAD: mismo motor, combustible, caja, tracción y
        # cabina cuando el aviso los declara. Si el aviso no declara alguno,
        # no se filtra por ese (para no quedarse sin comparables).
        r_mot, r_comb = r.get('motor'), r.get('combustible')
        r_caja = r.get('caja', '')
        r_trac = _trac(r['version']); r_cab = _cabina(r['version'])
        if pd.notna(r_mot) and str(r_mot):
            g = g[g['motor'] == r_mot]
        if pd.notna(r_comb) and str(r_comb):
            g = g[g['combustible'] == r_comb]
        if r_caja:
            g = g[(g['caja'] == r_caja) | (g['caja'] == '')]
        if r_trac:
            g = g[(g['_trac_c'] == r_trac) | (g['_trac_c'] == '')]
        if r_cab:
            g = g[(g['_cab_c'] == r_cab) | (g['_cab_c'] == '')]
        peers = g[(g['año'] - r['año']).abs() <= 3].copy()
        full = len(peers)
        if full > CAP_COMP:
            self_row = peers[peers['id'].str.lower() == vid]
            otros = peers[peers['id'].str.lower() != vid].sort_values('precio')
            idx = np.linspace(0, len(otros) - 1, CAP_COMP - 1).astype(int)
            peers = pd.concat([self_row, otros.iloc[idx]]).drop_duplicates(subset='id')
        peers = peers.sort_values('precio', ascending=False)
        filas = []
        for _, p in peers.iterrows():
            filas.append([
                int(p['año']), int(round(p['kilometraje'] / 1000)), int(p['precio']),
                (str(p['version'])[:24] if pd.notna(p['version']) else ''),
                ('P' if p['entidad'] == 'Particular' else 'A'),
                1 if str(p['id']).lower() == vid else 0,
                str(p['id']).lower(),
                str(p.get('gen', '') or ''),
                str(p.get('gens_pos', '') or ''),
            ])
        comps[vid] = {'n': full, 'rows': filas}

    # --- Hoja de depuración: sugerencias por marca/modelo/gen/firma ---
    corr_dep = cargar_correcciones()
    dep_src = dfa.copy()
    dep_src['_firma'] = dep_src['firma_ver'].fillna('')
    dep_src['_ver'] = dep_src['version'].fillna('').astype(str).str.strip()
    vol_mod = dep_src.groupby(['marca', 'modelo'])['precio'].transform('size')
    dep_src = dep_src[vol_mod >= 15]  # modelos con mercado real
    perfiles_ver = construir_perfiles_versiones(dep_src)
    if perfiles_ver:
        log(f"Perfiles de versiones dominantes (con ficha): "
            f"{sum(len(v) for v in perfiles_ver.values())} en "
            f"{len(perfiles_ver)} modelo/generación")
    depuracion = []
    for (ma, mo, gn, fv), g in dep_src.groupby(['marca', 'modelo', 'gen', '_firma']):
        if len(g) < 3:
            continue
        clave = f"{_norm_txt(ma)}|{_norm_txt(mo)}|{gn}|{fv}"
        guardado = corr_dep.get(clave) or {}
        ejemplos = [x for x in g['_ver'].drop_duplicates().head(3) if x]
        specs_vistas = ''
        if 'version_oficial' in g.columns:
            vofs = [str(x).strip() for x in g['version_oficial'].dropna().unique()
                    if str(x).strip()]
            specs_vistas = ' | '.join(vofs[:3])
        # Dispersión de precio ajustada por año: si dentro del MISMO año del
        # mismo grupo los precios varían demasiado, probablemente conviven
        # dos versiones que la firma no distingue -> revisar primero.
        disp = 0.0
        por_año = g.groupby('año')['precio']
        años_validos = [a for a, s in por_año if len(s) >= 3]
        if años_validos:
            ratios = []
            for a, s in por_año:
                if len(s) >= 3 and s.median() > 0:
                    q1, q3 = s.quantile([0.25, 0.75])
                    ratios.append(float((q3 - q1) / s.median()))
            if ratios:
                disp = round(max(ratios), 2)
        depuracion.append({
            'clave': clave, 'marca': ma, 'modelo': mo, 'gen': gn, 'firma': fv,
            'avisos': int(len(g)),
            'precio_mediano': int(g['precio'].median()),
            'años': f"{int(g['año'].min())}-{int(g['año'].max())}",
            'ejemplos': ' | '.join(ejemplos)[:120],
            'sugerida': sugerir_canonica(fv),
            'canonica': guardado.get('canonica', ''),
            'excluir': bool(guardado.get('excluir')),
            'dispersion': disp,
            'revisar': disp >= 0.35,  # rango intercuartil > 35% de la mediana en un mismo año
            'specs': specs_vistas,
            'precal': None,
        })
        # Pre-calificacion: solo para grupos SIN canonica y SIN ficha propia,
        # cuando el modelo tiene versiones dominantes conocidas
        if (not guardado.get('canonica') and not specs_vistas
                and len(g) <= 30 and (ma, mo, gn) in perfiles_ver):
            r = precalificar_grupo(g, fv, ejemplos, perfiles_ver[(ma, mo, gn)])
            if r:
                depuracion[-1]['precal'] = {
                    'version': r[0], 'conf': r[1], 'senales': r[2],
                    'fuente': r[3] if len(r) > 3 else 'specs'}
    depuracion.sort(key=lambda d: (-d['avisos']))

    # ---- TASADOR: tablas de precio historico por nivel de detalle ----
    # Nivel 'oficial' (version de ficha), 'firma' (version del texto
    # normalizada) y 'modelo' (todo el modelo/gen). El dashboard tasa
    # eligiendo el nivel mas fino con datos.
    tasador = []
    _src_t = dfa[(dfa['precio'] > 0) & dfa['año'].notna()].copy()
    _src_t['_fv'] = _src_t.get('firma_ver', pd.Series('', index=_src_t.index)).fillna('')

    def _agrega_tasa(gb, nivel, min_n):
        for llav, g in gb:
            if len(g) < min_n:
                continue
            *cab, an = llav
            ma, mo, gn = cab[0], cab[1], cab[2]
            ver = cab[3] if len(cab) > 3 else ''
            tasador.append({
                'ma': ma, 'mo': mo, 'gen': str(gn), 'ver': str(ver),
                'nivel': nivel, 'año': int(an), 'n': int(len(g)),
                'p50': int(g['precio'].median()),
                'p25': int(g['precio'].quantile(0.25)),
                'p75': int(g['precio'].quantile(0.75)),
                'km50': int(g['kilometraje'].median()) if g['kilometraje'].notna().any() else 0,
            })

    if 'version_oficial' in _src_t.columns:
        _cv = _src_t[_src_t['version_oficial'].fillna('').astype(str).str.strip() != '']
        _agrega_tasa(_cv.groupby(['marca', 'modelo', 'gen', 'version_oficial', 'año']),
                     'oficial', 2)
    _cf = _src_t[_src_t['_fv'] != '']
    _agrega_tasa(_cf.groupby(['marca', 'modelo', 'gen', '_fv', 'año']), 'firma', 3)
    _agrega_tasa(_src_t.groupby(['marca', 'modelo', 'gen', 'año']), 'modelo', 3)
    log(f"Tasador: {len(tasador)} celdas de precio historico exportadas")

    from collections import Counter as _Counter
    cont_etq = _Counter()
    for o in oportunidades:
        for e in o['etiquetas']:
            cont_etq[e] += 1

    # ---- ANÁLISIS DE MERCADO (visión de analista) ----
    # Trabaja sobre toda la base limpia y activa (dfa), no solo el radar.
    analisis = construir_analisis(dfa)
    depurador_modelos = construir_depurador_modelos(dfa)
    log(f"Depurador por modelo: {len(depurador_modelos)} modelos con versiones propuestas")

    salida = {
        'generado': datetime.now().isoformat(timespec='seconds'),
        'capturas': [c['fecha'].strftime('%Y-%m-%d') for c in capturas],
        'captura_analizada': reciente['fecha'].strftime('%Y-%m-%d'),
        'resumen': {
            'total': int(len(dfa)),
            'precio_mediano': int(dfa['precio'].median()),
            'op_oro': int(dfa['op_oro'].sum()),
            'op_precio': int(dfa['op_precio'].sum()),
            'op_km': int(dfa['op_km'].sum()),
        },
        'etiquetas_conteo': dict(cont_etq),
        'oportunidades': oportunidades,
        'comps': comps,
        'rotacion': rotacion,
        'depuracion': depuracion,
        'tasador': tasador,
        'analisis': analisis,
        'depurador_modelos': depurador_modelos,
    }

    with open(SALIDA, 'w', encoding='utf-8') as f:
        json.dump(salida, f, ensure_ascii=False, indent=2)

    log(f"Listo. Escrito '{SALIDA}'.")

    # Guardar Excel con la base analizada
    fecha_str = reciente['fecha'].strftime('%Y-%m-%d')
    nombre_excel = f"base_{fecha_str}.xlsx"
    cols_excel = [
        'marca', 'modelo', 'version', 'motor', 'año', 'kilometraje', 'precio',
        'precio_justo', 'descuento_pct', 'km_anual', 'comparables',
        'metodo', 'entidad', 'transmision', 'combustible', 'traccion',
        'condicion', 'op_oro', 'op_precio', 'op_km', 'url'
    ]
    cols_presentes = [c for c in cols_excel if c in dfa.columns]
    dfa[cols_presentes].to_excel(nombre_excel, index=False)
    log(f"Excel guardado: '{nombre_excel}' ({len(dfa)} filas).")

    print("\n--- RESUMEN ---")
    print(f"Capturas: {', '.join(salida['capturas'])}")
    r = salida['resumen']
    print(f"Captura analizada: {salida['captura_analizada']} ({r['total']} avisos limpios)")
    print(f"Oportunidades -> oro: {r['op_oro']} | bajo precio: {r['op_precio']} | bajo km: {r['op_km']}")
    if rotacion:
        g = rotacion['global']
        print(f"\nRotacion {g['fecha_desde']} -> {g['fecha_hasta']} ({g['dias']} dias):")
        print(f"  Salieron del mercado: {g['desaparecidos']} de {g['avisos_inicio']} "
              f"({g['tasa_salida_periodo']}% | {g['tasa_salida_diaria']}%/dia)")
        print(f"  Cambios de precio: {g['con_cambio_precio']} "
              f"(bajaron {g['bajaron_precio']}, subieron {g['subieron_precio']})")
        alto = [m for m in rotacion['por_modelo'] if m['inventario'] == 'alto']
        bajo = [m for m in rotacion['por_modelo'] if m['inventario'] == 'bajo']
        print(f"\n  Modelos por inventario: {len(alto)} alto (>=10) | {len(bajo)} bajo (<10)")
        print("\n  ALTO INVENTARIO - rotacion confiable, mas rapidos:")
        for m in alto[:8]:
            dv = f"~{m['dias_venta_est']}d" if m['dias_venta_est'] else "s/d"
            print(f"    {m['marca']} {m['modelo']}: {m['tasa_salida']}% "
                  f"({m['salieron']}/{m['stock']}, venta est. {dv})")


if __name__ == '__main__':
    main()
