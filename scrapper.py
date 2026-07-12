from curl_cffi import requests
from curl_cffi.requests import AsyncSession
import asyncio
import time
from bs4 import BeautifulSoup
import random
import logging
from datetime import datetime

import re
import unicodedata
import os
import pandas as pd
import json

from colorama import init, Fore, Style
import sys

# Configure stdout to handle utf-8 characters properly on Windows
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')

init(autoreset=True)


max_workers = min(32, (os.cpu_count() or 1) + 4)

# Headers generales
headers_base = {
    "accept": "application/json",
    "accept-encoding": "gzip, deflate, br, zstd",
    "accept-language": "es,es-ES;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "cookie": "csnclientid=7C71FBBC-EEDB-4946-A51F-B7AD6785DE0C; ...",  
    "user-agent": "Mozilla/5.0 (Linux; Android 13; SM-G981B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36 Edg/132.0.0.0",  
    "referer": "https://www.chileautos.cl/vehiculos/autos-veh%C3%ADculo/",
    "sec-ch-ua": "\"Not A(Brand\";v=\"8\", \"Chromium\";v=\"132\", \"Microsoft Edge\";v=\"132\"",
    "sec-ch-ua-mobile": "?1",
    "sec-ch-ua-platform": "\"Android\"",
}

cookies_manual = {
    "_fbp": "fb.1.1773063990118.383977275651180737",
    "_ga": "GA1.1.1977719151.1773063989",
    "_ga_2S06LYQ1G3": "GS2.1.s1773063994$o1$g0$t1773063994$j60$l0$h0",
    "_ga_KPC4YK8R9E": "GS2.1.s1773063990$o1$g1$t1773064000$j50$l0$h0",
    "_gat": "1",
    "_gcl_au": "1.1.1643385582.1773063990",
    "_gid": "GA1.2.304303509.1773063989",
    "_hjSession_5130356": "eyJpZCI6ImQ5N2FhMmU4LTY1MWUtNDIxMC04NDdmLTgxMDE0ZTVkZWM1ZSIsImMiOjE3NzMwNjM5ODk3MDIsInMiOjAsInIiOjAsInNiIjowLCJzciI6MCwic2UiOjAsImZzIjoxLCJzcCI6MH0=",
    "_hjSessionUser_5130356": "eyJpZCI6ImE5OWU1NjFlLTVkMTEtNWUwNi1iYWJmLWZlOTA3NWY0YmM0MSIsImNyZWF0ZWQiOjE3NzMwNjM5ODk3MDAsImV4aXN0aW5nIjp0cnVlfQ==",
    "_sharedID": "2090ff63-d7f0-42e9-9b97-7b76587e0b62",
    "_sharedID_cst": "kSylLAssaw%3D%3D",
    "afx_csid": "a05d9a01506d42ca953fd932606912558de7de2415b4b6f",
    "afx_pbd": "bTr8Z%2BOo2atUeFkcfcuzBQXmCoBnGOQaQv1ekd9n4As%2Fh0fnh%2BqbJwJ4zesPw3TSD2WVafGdsCDbEw56nFPTbDcDabcKAHI3sEDzEM0lHZUJOSjg8uu4LLCRhrfbL0ZfWbrMxWSIIO6LIh15yzXkvHqbR0hoNU%2Fg05G7jfVE31N%2F3VNmipUxKbKgbLhNQTMro9SsUBXQr16qvhRD%2B9xFpQ%3D%3D",
    "afx_profile": "%7B%22acid%22%3A%221453175ce84559ff66b51fae4e92ce30eb793203%22%2C%22prid%22%3A%22W7M1SNgdpafXL53SRwo2fe5ArnKdxtgOu7qyI1aQMZ9UjeoigxDpOL2jNd8wjnsdgUqqa7bpNa5%2FOEuL1ilq00%2Ft5czf0oAaocnTwqJ88n4MfGh6VZPGxzO1EfKl58H7%2B3jqTYaj%2FYj%2FGaF3iOVGrGWxs7EyI8T9df%2FNIF8X9i2ly6f4mX9GtPEZ9HKQXoOnppsN6jq4bqJcCmUr%2BLRjMQ%3D%3D%22%2C%22pridsd%22%3A%222026-03-09T13%3A46%3A25.0396309%2B00%3A00%22%2C%22consent%22%3A%7B%22dnt%22%3Afalse%7D%2C%22pv%22%3A%222026.2.13.1%22%7D",
    "afx_profile_da": "1",
    "afx_profile_hs": "%7B%22acid%22%3A%221453175ce84559ff66b51fae4e92ce30eb793203%22%2C%22prid%22%3A%22W7M1SNgdpafXL53SRwo2fe5ArnKdxtgOu7qyI1aQMZ9UjeoigxDpOL2jNd8wjnsdgUqqa7bpNa5%2FOEuL1ilq00%2Ft5czf0oAaocnTwqJ88n4MfGh6VZPGxzO1EfKl58H7%2B3jqTYaj%2FYj%2FGaF3iOVGrGWxs7EyI8T9df%2FNIF8X9i2ly6f4mX9GtPEZ9HKQXoOnppsN6jq4bqJcCmUr%2BLRjMQ%3D%3D%22%2C%22pridsd%22%3A%222026-03-09T13%3A46%3A33.4190126%2B00%3A00%22%2C%22consent%22%3A%7B%22dnt%22%3Afalse%7D%2C%22pv%22%3A%222026.2.13.1%22%7D",
    "afx_ptpce": "1",
    "afx_syndr": "1773063995005",
    "AMCV_72FF3B526128B17A0A495F9B%40AdobeOrg": "MCMID|17712920789034764902887733948544731721",
    "cto_bundle": "7l38H19aVVBsMHliUkxPJTJGdmJPdDZ2SENSU1pGOG01NXd4ZVhuN1VoTFRCVzElMkZ3dVhYbTdLaE1SUiUyQmtNWTclMkJQSHh5Z3RYaThuJTJGdjZHRnp5dTNpMXdqRWVZcjFvT3V5ZW54ZGhvV0JsUUw4dDI0dUNsOHJNb2JDclN1bFZEMVJveDZ5QkZPRU9YR0pnJTJGSUozRFBmaEpwJTJCS25zQSUzRCUzRA",
    "datadome": "js9v5eAkkQKoTY_BFgAnlul7FCNKctjM87kRXZ3TMal5Y2TPr1cIdOkAgF3TkdN8Uf0Wm83f1GUPihVOvush9Y~F1HSH8X5NFzLXmYOSaKCtPlOZ_7OCtT1WlAzMp8rC",
    "FCCDCF": "%5Bnull%2Cnull%2Cnull%2Cnull%2Cnull%2Cnull%2C%5B%5B32%2C%22%5B%5C%225af3011e-2e54-4696-9853-fee81dd3ba19%5C%22%2C%5B1773063989%2C942000000%5D%5D%22%5D%5D%5D",
    "FCNEC": "%5B%5B%22AKsRol9TIta3Qe6tAXRNquzqt2PdnItj9G0HcSTTPTKjE_ldDIPH6TixRU1eqsHBOG0huGEynEmphqUHUbf5Oq6sKH_4VyAJIPr_ree2igx7X3G-NLIlTypfmssCtNnE1a1kbEUuyiDB7rB0wI-YxZkKqZTiMjL5PA%3D%3D%22%5D%5D",
    "kndctr_72FF3B526128B17A0A495F9B_AdobeOrg_cluster": "va6",
    "kndctr_72FF3B526128B17A0A495F9B_AdobeOrg_identity": "CiYxNzcxMjkyMDc4OTAzNDc2NDkwMjg4NzczMzk0ODU0NDczMTcyMVISCJbG5ZbNMxABGAEqA1ZBNjAA8AGWxuWWzTM=",
    "aws-waf-token": "46b0aa3b-d794-40d9-8771-1344ad874e61:EAoAZaNgOyAwAAAA:QvjB8dPPHcZcQQiAfiC9PzdQltomzRny1XgFrCG1dvmcGmz7eIJNnAKI5oaEULCUA8J+1LbLWvOKqzScEJU+95Bo1F0hKBwgPFoeNRm3ZpNVKRGFtiQuIwDrgu1Ue1zB86zclgyq8Ozpqo+dAMvmYTMYfU4mDhjHUAWoYs0oMRjv7WWAgdzGGC16MSx66TDohYv5kVqb6Lf3gpKsCf/LfWHOwzC5RPWqN0201aSREdrxy8yJ7h5RWq3y",
    "usprivacy": "1---",
    "adpbp": "w",
    "afxcarsid": "1453175ce84559ff66b51fae4e92ce30eb793203",
    "cidgenerated": "1",
    "csn.bi": "1773064000268",
    "csncidcf": "39F5DD75-EB80-47F9-9BE0-6556FDAFE8B0",
    "csnclientid": "B779762E-831B-44FB-B37D-54DFDD19564C",
    "csnSessionId": "spdGlpaF+kefuLrO1pMzRgAAF4jOljpk4ZnUIAaxzQ9znoA9urmcis9jy+LSsLHSJxZ++Q=="
}
headers_base["cookie"] = "; ".join([f"{k}={v}" for k, v in cookies_manual.items()])

# Configuración de logging
logging.basicConfig(level=logging.INFO)

# Función para realizar solicitudes con sesión asíncrona
async def realizar_solicitud_con_sesion_async(url, session, semaphore=None, cookies=None, max_reintentos=5):
    # Ya no forzamos el semaphore aquí si llega como None, dejamos la concurrencia a la Cola.
    intentos = 0
    while intentos < max_reintentos:
        try:
            # Copia de los headers base
            headers = headers_base.copy()
            # Formatear las cookies si existen
            if cookies:
                headers["cookie"] = "; ".join([f"{k}={v}" for k, v in cookies.items()])
            
            # Realizar la solicitud cruda con Timeout de 4s
            response = await session.get(url, headers=headers, impersonate="chrome99_android", timeout=4)
            
            if response.status_code == 200:
                return response
            elif response.status_code in [403, 429]:
                tiempo_bloqueado = random.uniform(30, 60)
                logging.warning(f"Bloqueo Datadome {response.status_code} en {url}. Durmiendo {tiempo_bloqueado:.1f}s...")
                await asyncio.sleep(tiempo_bloqueado)
            else:
                logging.warning(f"Error {response.status_code} al acceder a {url}. Reintentando...")
        except Exception as e:
            logging.error(f"Error en la solicitud a {url}: {e}")
        
        await asyncio.sleep(random.uniform(2,5))
        intentos += 1
    
    logging.error(f"No se pudo acceder a {url} después de {max_reintentos} intentos.")
    return None

def limpiar_texto(texto):
    # Normalizar el texto a Unicode NFKD para eliminar caracteres especiales
    texto = unicodedata.normalize('NFKD', texto)
    texto = texto.encode('ascii', 'ignore').decode('utf-8')  # Eliminar caracteres no ASCII

    # Eliminar caracteres no deseados excepto letras, números y espacios
    texto = re.sub(r'[^\w\sáéíóúñÁÉÍÓÚÑ]', '', texto)
    
    # Reemplazar espacios no visibles (incluyendo \u00A0, \t, etc.) por un espacio normal
    texto = re.sub(r'\s+', ' ', texto)
    
    # Eliminar comillas especiales o caracteres invisibles
    texto = re.sub(r'[\u201C\u201D\u2018\u2019]', '', texto)
    
    texto.replace('"', '')
    texto = texto.strip()
    
    return texto




def obtener_cookies_iniciales(session):
    pass

async def obtener_cookies_iniciales_async(session):
    # Realiza una solicitud inicial para obtener las cookies necesarias
    url = "https://www.chileautos.cl/mobiapi/chileautos/v1/stock/listing?p=TipoVeh%C3%ADculo.Autos.&pg=1&ni=18"
    headers = headers_base.copy()
    response = await session.get(url, headers=headers, impersonate="chrome99_android", timeout=6)
    if response.status_code == 200:
        logging.info("Cookies iniciales obtenidas con éxito.")
        
        return response.cookies.get_dict()
    else:
        logging.error(f"Error al obtener cookies iniciales: {response.status_code}")
        return None

def extraer_vehiculos(html):
    soup = html.json()
    
    # Encontrar todos los contenedores de vehículos
    vehiculos = soup.get("result")
    
    # Lista para almacenar la información de los vehículos
    lista_vehiculos = []
    
    for vehiculo in vehiculos:
        try:
            año = int(vehiculo.get("saveItemAction").get("tracking").get("fb").get("attributes").get("year"))
            moneda = vehiculo.get("displayPrice").get("preInfo")
            auto = vehiculo.get("saveItemAction").get("tracking").get("fb").get("attributes").get("make") 
            precio = int(limpiar_texto(vehiculo.get("displayPrice").get("price").split(" ")[0
            
            ]))
            condicion = vehiculo.get("saveItemAction").get("tracking").get("fb").get("attributes").get("condition_of_vehicle")
            data = vehiculo.get("keyDetails")
            combustible = vehiculo.get("saveItemAction").get("tracking").get("fb").get("attributes").get("fuel_type")
            id = vehiculo.get("saveItemAction").get("tracking").get("fb").get("attributes").get("content_ids")
            transmision = ""
            kilometraje = 0
            modelo = vehiculo.get("saveItemAction").get("tracking").get("fb").get("attributes").get("model")
            variantes = vehiculo.get("displayTitle").split(" ")
            variante = ""
            traccion = "4x2"
            
            for a in range(3, len(variantes)):
                if a != len(variantes):
                      variante += variantes[a]
                      variante += " "
                else:
                    variante += variantes[a]
            variante = variante[:-1]
                
            for x in data:
                try:
                
                    if x.split(" ")[1] == 'km':
                        kilometraje =  int(limpiar_texto(x.split(" ")[0]))
                        
                        
                
                except:
                    None
                    
                try:
                    if x == 'Automática':
                      transmision = "automatica"
                    elif x == 'Manual':
                      transmision = "manual"
                    else:
                        transmision = ""
                except:
                    transmision = ""
            for palabra in variantes:
                if "4x4" == palabra.lower():
                    traccion = "4x4"  
                if "4wd" == palabra.lower():        
                    traccion = "4x4"
                    
                else:
                    None
            
            entidad = vehiculo.get("siloTypeFriendlyName")
            
            
            detalles_url = f"https://www.chileautos.cl{vehiculo.get('webDetailsUrl')}"
            
            # Se ha eliminado el print coloreado para evitar spam en la consola

            
            lista_vehiculos.append({
                "fecha_scrapeo": datetime.now().strftime("%d-%m-%Y %H:%M"),
                "año": año,
                "modelo": modelo,
                "version": variante.lower(),
                "condicion": condicion,
                "traccion": traccion,
                'marca': auto,
                'precio': precio,
                'kilometraje': kilometraje,
                'transmision': transmision,
                'combustible': combustible,
                'moneda': moneda,
                'id':id,
                'entidad': entidad,
                'url_detalles': detalles_url
            })
            
        except AttributeError as e:
            
            continue
    
    return lista_vehiculos

async def fetch_and_process(url, session, cookies, productos_totales):
    response = await realizar_solicitud_con_sesion_async(url, session, None, cookies)
    if response:
        # Extraer vehículos y agregarlos a la lista global
        vehiculos = extraer_vehiculos(response)
        productos_totales.extend(vehiculos)


async def obtener_productos_con_sesion_async():
    # -----------------------------------------------------------------
    # NÚMERO DE PÁGINAS A RECORRER
    # Normal (todo el catálogo): 3500
    # PRUEBA rápida (segundos):  poné 20
    # Cambiá SOLO el número de acá abajo.
    # -----------------------------------------------------------------
    TOTAL_PAGINAS = 3500  # catalogo COMPLETO (antes 20: valor de prueba que quedo pegado)

    paginas = [
        ['https://www.chileautos.cl/mobiapi/chileautos/v1/stock/listing?p=TipoVeh%C3%ADculo.Autos.&pg=', TOTAL_PAGINAS],
    ]
    productos_totales = []
    
    async with AsyncSession() as session:
        cookies = await obtener_cookies_iniciales_async(session)
        if cookies:
            headers_base["cookie"] = "; ".join([f"{k}={v}" for k, v in cookies.items()])
        else:
            logging.error("No se pudieron obtener cookies iniciales.")
            
        print(f"Trabajando con {max_workers} trabajadores (Async Queue)")
        
        # 1. Crear Cola de páginas
        queue_paginas = asyncio.Queue()
        total_paginas_a_visitar = 0
        for base_url, total_paginas in paginas:
            total_paginas_a_visitar += total_paginas
            for pagina in range(total_paginas):
                url = f"{base_url}{pagina}&ni=18"
                queue_paginas.put_nowait(url)
                
        # Variables de progreso y ETA
        paginas_completadas = 0
        start_time = time.time()
        
        async def trabajador_paginas():
            nonlocal paginas_completadas
            while True:
                try:
                    url = queue_paginas.get_nowait()
                except asyncio.QueueEmpty:
                    break
                    
                await fetch_and_process(url, session, cookies, productos_totales)
                
                queue_paginas.task_done()
                paginas_completadas += 1
                
                # Imprimir progreso cada 20 páginas completadas
                if paginas_completadas % 20 == 0 or paginas_completadas == total_paginas_a_visitar:
                    elapsed = time.time() - start_time
                    vel = paginas_completadas / elapsed if elapsed > 0 else 0
                    restantes = total_paginas_a_visitar - paginas_completadas
                    eta_sec = restantes / vel if vel > 0 else 0
                    eta_str = time.strftime('%H:%M:%S', time.gmtime(eta_sec))
                    
                    sys.stdout.write(f"\rProgreso Autos (Páginas): {paginas_completadas}/{total_paginas_a_visitar} | Autos hallados: {len(productos_totales)} | ETA: {eta_str}   ")
                    sys.stdout.flush()
                    
        # Iniciar trabajadores
        workers = [asyncio.create_task(trabajador_paginas()) for _ in range(max_workers)]
        await asyncio.gather(*workers)
        
        print() # Salto de línea limpio tras el flush
    
    print(f"\n✅ Análisis finalizado. Se extrajeron {len(productos_totales)} vehículos en total.")
    return productos_totales

def obtener_productos_con_sesion():
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    return asyncio.run(obtener_productos_con_sesion_async())

def crear_base_datos(nombre_archivo, tienda):
    productos = tienda()
    with open(nombre_archivo, 'w', encoding='utf-8') as archivo:
        json.dump(productos, archivo, ensure_ascii=False, indent=4)

def formato_excel(json):
    fecha = []
    marca = []
    año = []
    modelo = []
    version = []
    condicion = []
    precio = []
    kilometraje = []
    transmision = []
    combustible = []
    moneda = []
    url_detalles = []
    traccion =[]
    id = []
    entidad = []
    telefono = []
    whatsapp = []
    nombre_whatsapp = []

    for datos in json:
        fecha.append(datos["fecha_scrapeo"])
        marca.append(datos["marca"])
        año.append(datos["año"])
        modelo.append(datos["modelo"])
        version.append(datos["version"])
        traccion.append(datos["traccion"])
        condicion.append(datos["condicion"])
        precio.append(datos["precio"])
        kilometraje.append(datos["kilometraje"])
        transmision.append(datos["transmision"])
        combustible.append(datos["combustible"])
        moneda.append(datos["moneda"])
        url_detalles.append(datos.get("url_detalles", ""))
        id.append(datos.get("id", ""))
        entidad.append(datos.get("entidad", ""))
        telefono.append(datos.get("telefono", "No disponible"))
        whatsapp.append(datos.get("whatsapp", "No disponible"))
        nombre_whatsapp.append(datos.get("nombre_whatsapp", ""))
    
    vehiculos_dict = {
    "fecha": fecha,
    "marca": marca,
    "modelo": modelo,
    "version": version,
    "año": año,
    "traccion": traccion,
    "condicion": condicion,
    "precio": precio,
    "kilometraje": kilometraje,
    "transmision": transmision,
    "combustible": combustible,
    "id": id,
    "moneda": moneda,
    "entidad": entidad,
    "url_detalles": url_detalles,
    "telefono": telefono,
    "whatsapp": whatsapp,
    "nombre_whatsapp": nombre_whatsapp,
    
    }
    
    return pd.DataFrame(vehiculos_dict)


import pandas as pd
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def guardar_excel_con_autoajuste(dataframe):
    try:
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)

        ruta_archivo = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialfile="Chileautos",
            title="Guardar archivo Excel report como..."
        )

        if ruta_archivo:
            # Primero guarda el archivo base con pandas
            dataframe.to_excel(ruta_archivo, index=False)

            # Usar openpyxl para agregar los "superpoderes" visuales
            wb = load_workbook(ruta_archivo)
            ws = wb.active
            
            # --- 1. DEFINIR TEMA Y COLORES CORE ---
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul Marino Corporativo
            row_fill_par = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Gris Claro cebra
            row_fill_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Blanco cebra
            
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            link_font = Font(name="Segoe UI", size=10, color="0563C1", underline="single")
            item_font = Font(name="Segoe UI", size=10, color="333333")
            
            thin_border = Border(left=Side(style='thin', color='D9D9D9'), 
                                 right=Side(style='thin', color='D9D9D9'), 
                                 top=Side(style='thin', color='D9D9D9'), 
                                 bottom=Side(style='thin', color='D9D9D9'))
                                 
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            
            # --- 2. UBICAR COLUMNAS ESPECIALES ---
            url_col_idx = None
            price_col_idx = None
            kms_col_idx = None
            
            # Formatear la Fila 1 (Cabeceras)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Aprovechar para sacar índices clave e incluso renombrar en la marcha
                if str(cell.value).lower() == "url_detalles":
                    url_col_idx = cell.column
                    cell.value = "Enlace Web"
                elif str(cell.value).lower() == "precio":
                    price_col_idx = cell.column
                elif str(cell.value).lower() == "kilometraje":
                    kms_col_idx = cell.column
                    
            # --- 3. APLICAR ESTILOS A LOS DATOS (FILAS 2 EN ADELANTE) ---
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                es_par = row[0].row % 2 == 0
                fondo_actual = row_fill_par if es_par else row_fill_impar
                
                for cell in row:
                    cell.fill = fondo_actual
                    cell.border = thin_border
                    cell.font = item_font
                    cell.alignment = left_alignment
                    
                    # Convertir enlaces web de URL de texto a botones Clickables
                    if cell.column == url_col_idx and cell.value:
                        link = str(cell.value)
                        cell.hyperlink = link
                        cell.value = "Abrir Publicación 🔗" # El texto bonito que se verá
                        cell.font = link_font
                        cell.alignment = center_alignment
                        
                    # Formato Dinero ($) para la columna PRECIO
                    if cell.column == price_col_idx and cell.value != None:
                        try:
                            # Asegurar que sea int/float
                            val = int(str(cell.value).replace(".","").replace(",",""))
                            cell.value = val
                            cell.number_format = '$ * #,##0'
                            cell.alignment = center_alignment
                        except:
                            pass
                            
                    # Formato Kilómetros para la columna KILOMETRAJE
                    if cell.column == kms_col_idx and cell.value != None:
                        try:
                            val = int(str(cell.value).replace(".","").replace(",",""))
                            cell.value = val
                            cell.number_format = '#,##0 "km"'
                            cell.alignment = center_alignment
                        except:
                            pass

            # --- 4. AUTO-AJUSTE DINÁMICO DE ANCHOS DE COLUMNA ---
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Ajuste final de la columna (con márgenes y topes de estética)
                adjusted_width = max_length + 2
                # Forzar columna de Enlace a un tamaño fijo para que no sobresalga el icono
                if column == get_column_letter(url_col_idx) if url_col_idx else None:
                    adjusted_width = 22
                
                ws.column_dimensions[column].width = min(adjusted_width, 50)

            # --- 5. DETALLES FINALES UX (Filtro Autofilter + Panel Congelado) ---
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            # Guardar Todo
            wb.save(ruta_archivo)
            print(Fore.GREEN + f"✅ ¡Excel Profesional y Funcional Guardado en: {ruta_archivo}!" + Style.RESET_ALL)
        else:
            print(Fore.RED + "❌ Operación de guardado cancelada por el usuario." + Style.RESET_ALL)
    except Exception as e:
        print(Fore.RED + f"⚠️ Error al guardar el reporte estilizado: {e}" + Style.RESET_ALL)



def mostrar_banner():
    banner = Fore.MAGENTA + Style.BRIGHT + r"""
                    ,--,                                
            ,--. ,---.'|                                
        ,--/  /| |   | :                     ,----..    
        ,---,': / ' :   : |             ,--,   /   /   \   
        :   : '/ /  |   ' :           ,'_ /|  |   :     :  
        |   '   ,   ;   ; '      .--. |  | :  .   |  ;. /  
        '   |  /    '   | |__  ,'_ /| :  . |  .   ; /--`   
        |   ;  ;    |   | :.'| |  ' | |  . .  ;   | ;  __  
        :   '   \   '   :    ; |  | ' |  | |  |   : |.' .' 
        |   |    '  |   |  ./  :  | | :  ' ;  .   | '_.' : 
        '   : |.  \ ;   : ;    |  ; ' |  | '  '   ; : \  | 
        |   | '_\.' |   ,/     :  | : ;  ; |  '   | '/  .' 
        '   : |     '---'      '  :  `--'   \ |   :    /   
        ;   |,'                :  ,      .-./  \   \ .'    
        '---'                   `--`----'       `---`      
    """ + Fore.YELLOW + "              CHILEAUTOS SCRAPER by Nicolás Stade\n"
    print(banner)

if __name__ == "__main__":
    
    mostrar_banner()
    time.sleep(2)
    print(Fore.YELLOW + Style.BRIGHT + "\n⚠️  Atención:" + Style.RESET_ALL)
    print("Es normal que el script se detenga momentáneamente si la API se satura.")
    print("Por favor, espera. El sistema reintentará automáticamente y continuará el proceso.")
    print(Fore.CYAN + "Al finalizar, se abrirá una ventana para guardar el archivo Excel." + Style.RESET_ALL)

    time.sleep(7)
    productos = obtener_productos_con_sesion()
    
    
    df = formato_excel(productos)
    
    print("Datos obtenidos y formateados correctamente.")
    guardar_excel_con_autoajuste(df)

